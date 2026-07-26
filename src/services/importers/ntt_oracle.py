"""
Importador do relatório Oracle/NTT.

Este módulo deverá ser chamado pelo scheduler de importações através de algo como:
    src.services.import_scheduling.IMPORT_DISPATCHER["NTTOracle"]

Objetivos principais:
- Ler o Excel Oracle/NTT em modo edição (openpyxl, read/write).
- Processar linhas em chunks (lotes) para reduzir pico de memória.
- Usar caches em memória para reduzir lookups repetidos no banco.
- Atualizar status de progresso na tbImportControl (ImportControlRepository).
- Resolver product_id, customer_id, asset_id (quando definirmos o layout).
- Inserir/atualizar contratos em tbContractVendorAsset (ou tabela específica).
- Processar o arquivo de entrada IN PLACE, removendo linhas importadas com sucesso.
  Ao final (ou em caso de interrupção), o arquivo ficará apenas com:
    * linhas não lidas; e
    * linhas lidas mas sem sucesso na importação.
"""

import logging
from pathlib import Path
from typing import Dict, Optional, Any, List, Tuple
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.domain.import_validator import validate_import_columns
from src.infrastructure.database.repositories.import_log_repository import (
    ImportLogRepository,
)
from src.infrastructure.database.repositories.import_control_repository import (
    ImportControlRepository,
    ImportStatus,
)

logger = logging.getLogger(__name__)

# Diretório padrão de entrada (seguindo padrão do projeto)
BASE_INPUT_PATH = Path("/home/bridgeadoption/storage/input")

# Config de performance
CHUNK_ROWS = 2000       # tamanho do lote de linhas
STATUS_EVERY = 500      # atualizar status a cada N linhas
NTT_VENDOR_ID = 2       # TODO: ajustar para o vendor_id real da NTT/Oracle


# =============================================================================
# STATUS / PROGRESSO (tbImportControl)
# =============================================================================
def _fmt_hms(seconds: float) -> str:
    if seconds is None or seconds < 0:
        return "--:--:--"
    seconds = int(seconds)
    h = seconds // 3600
    m = (seconds % 3600) // 60
    s = seconds % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _build_progress_message(
    filename: str,
    read_rows: int,
    ok_rows: int,
    fail_rows: int,
    total_rows: int,
    started_at: datetime,
    eta_dt: Optional[datetime],
) -> str:
    now = datetime.now()
    elapsed = (now - started_at).total_seconds()
    pct = (read_rows / total_rows * 100.0) if total_rows else 0.0
    remaining = max(total_rows - read_rows, 0)
    eta_str = eta_dt.strftime("%Y-%m-%d %H:%M:%S") if eta_dt else "N/A"

    return (
        f"[NTTOracle] {filename} | "
        f"lidas={read_rows}/{total_rows} ({pct:6.2f}%) | "
        f"sucesso={ok_rows} | "
        f"falha_parcial={fail_rows} | "
        f"restante_estimado={remaining} | "
        f"elapsed={_fmt_hms(elapsed)} | ETA={eta_str}"
    )


def _calc_eta(
    started_at: datetime,
    read_rows: int,
    total_rows: int,
) -> Optional[datetime]:
    if read_rows <= 0:
        return None
    elapsed = (datetime.now() - started_at).total_seconds()
    rate = elapsed / max(read_rows, 1)  # sec/row
    remaining = max(total_rows - read_rows, 0)
    eta_sec = remaining * rate
    return datetime.now() + timedelta(seconds=eta_sec)


# =============================================================================
# VALIDAÇÃO DE COLUNAS
# =============================================================================
def validate_ntt_oracle_file_columns(file_columns: list) -> Dict[str, list]:
    """
    Usa o domínio de importação para validar as colunas esperadas
    do layout Oracle/NTT.

    É necessário configurar o tipo "NTTOracle" em src/domain/import_schemas.py
    para que esta função funcione corretamente.
    """
    return validate_import_columns(
        import_type="NTTOracle",
        file_columns=file_columns,
        case_sensitive=False,
    )


# =============================================================================
# EXCEL READ/WRITE (IN PLACE)
# =============================================================================
def _open_workbook_rw(path_xls: Path) -> Tuple[List[str], Worksheet]:
    """
    Abre o workbook em modo de leitura/escrita (in place) e retorna:
        (headers, worksheet)
    A primeira linha é considerada header.
    """
    wb = load_workbook(str(path_xls), read_only=False, data_only=True)
    ws = wb.active

    # headers na primeira linha
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    return headers, ws


def _ws_max_data_row(ws: Worksheet) -> int:
    """
    Retorna o índice da última linha com dados (ignorando headers).
    """
    return max(ws.max_row, 1)


# =============================================================================
# PROCESSAMENTO EM CHUNKS (IN PLACE)
# =============================================================================
def _process_chunk_inplace(
    ws: Worksheet,
    start_row: int,
    end_row: int,
    headers: List[str],
    file_name: str,
    log_repo: ImportLogRepository,
) -> Tuple[int, int, int]:
    """
    Processa um chunk de linhas do Excel IN PLACE (linha por linha).

    - Linhas com sucesso serão REMOVIDAS da worksheet.
    - Linhas com falha permanecem no arquivo.

    Parâmetros:
        ws        : worksheet aberta em modo editável
        start_row : primeira linha de dados (>= 2)
        end_row   : última linha de dados a considerar (inclusive)
        headers   : lista de headers (1ª linha)
        file_name : nome do arquivo (para logs)

    Retorna:
        (qtd_linhas_lidas, qtd_sucesso, qtd_falha_no_chunk)

    OBS: Nesta primeira versão, ainda não há lógica de persistência;
    todas as linhas são tratadas como "falha" até definirmos o layout
    Oracle/NTT e o mapeamento para o banco.
    """
    idx_map = {h: i for i, h in enumerate(headers)}

    linhas_lidas = 0
    linhas_sucesso = 0
    linhas_falha = 0

    for row_idx in range(end_row, start_row - 1, -1):
        excel_row_number = row_idx  # número real da linha no Excel
        linhas_lidas += 1

        row_cells = ws[row_idx]
        raw_row = [cell.value for cell in row_cells]

        row_dict = {
            h: (raw_row[idx_map[h]] if idx_map[h] < len(raw_row) else None)
            for h in headers
        }

        # TODO: aqui entra a lógica de:
        # - resolver customer_id
        # - resolver/crear product_id
        # - resolver/crear asset_id
        # - montar dados de contrato / vendorasset
        # - inserir/atualizar no banco
        #
        # Por enquanto, geramos um log simples de "não implementado"
        # e consideramos a linha como falha (não removemos).
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=excel_row_number,
            message="Processamento NTT Oracle ainda não implementado para esta linha.",
        )
        linhas_falha += 1

        # Quando implementarmos a lógica real:
        # if linha_sucesso:
        #     ws.delete_rows(row_idx, 1)
        #     linhas_sucesso += 1
        # else:
        #     linhas_falha += 1

    return linhas_lidas, linhas_sucesso, linhas_falha


# =============================================================================
# FUNÇÃO PRINCIPAL
# =============================================================================
def run_import(file_name: str, user_id: Optional[str]) -> Dict[str, str]:
    """
    Função de entrada chamada pelo scheduler de importações.

    Comportamento (IN PLACE) planejado:
    - O arquivo de entrada é processado diretamente.
    - Cada linha processada com sucesso é removida do arquivo.
    - Linhas com falha permanecem no arquivo.
    - Se o processo for interrompido (timeout, kill, etc.),
      o arquivo conterá:
        * linhas não lidas; e
        * linhas lidas mas sem sucesso.

    Além disso, o status na tbImportControl é atualizado ao longo da execução
    com informações de progresso.
    """
    logger.info("Iniciando import_ntt_oracle (in place) para arquivo: %s", file_name)
    start_time = datetime.now()

    path = BASE_INPUT_PATH / file_name

    log_repo = ImportLogRepository()
    import_ctrl_repo = ImportControlRepository()

    # Encontrar importctrl_id para este arquivo
    importctrl_id = import_ctrl_repo.get_id_by_file(file_name)
    if not importctrl_id:
        msg = f"Registro de controle não encontrado para arquivo: {file_name}"
        logger.error(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )
        return {"status": "FAILED", "message": msg}

    # 1) Validação de existência e extensão
    if not path.exists():
        msg = f"Arquivo não encontrado: {path}"
        logger.error(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )
        import_ctrl_repo.fail_import(importctrl_id, message=msg)
        return {"status": "FAILED", "message": msg}

    if path.suffix.lower() != ".xlsx":
        msg = "Arquivo deve possuir extensão .xlsx"
        logger.error(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )
        import_ctrl_repo.fail_import(importctrl_id, message=msg)
        return {"status": "FAILED", "message": msg}

    # Marca início da importação
    import_ctrl_repo.start_import(
        importctrl_id,
        message=f"Iniciando processamento IN PLACE do arquivo {file_name}",
    )

    # 2) Abre workbook para leitura/escrita (IN PLACE) e valida colunas
    try:
        headers, ws = _open_workbook_rw(path)
        wb = ws.parent
    except Exception as e:
        msg = f"Erro ao abrir Excel em modo edição (in place): {e}"
        logger.exception(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )
        import_ctrl_repo.fail_import(importctrl_id, message=msg)
        return {"status": "FAILED", "message": msg}

    # 3) Validação de colunas via domínio
    validation = validate_ntt_oracle_file_columns(headers)
    missing_required = validation.get("missing_required") or []

    if missing_required:
        msg = f"Colunas obrigatórias ausentes: {', '.join(missing_required)}"
        logger.error(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )
        try:
            wb.close()
        except Exception:
            pass
        import_ctrl_repo.fail_import(importctrl_id, message=msg)
        return {"status": "FAILED", "message": msg}

    # 4) Determina total de linhas de dados (excluindo header)
    try:
        max_row = _ws_max_data_row(ws)
        total_rows = max(max_row - 1, 0)
    except Exception:
        total_rows = 0

    read_rows = 0
    ok_rows = 0
    fail_rows = 0
    last_status_update = 0

    # status inicial na tbImportControl
    try:
        eta = _calc_eta(start_time, read_rows, total_rows)
        progress_msg = _build_progress_message(
            filename=file_name,
            read_rows=read_rows,
            ok_rows=ok_rows,
            fail_rows=fail_rows,
            total_rows=total_rows,
            started_at=start_time,
            eta_dt=eta,
        )
        import_ctrl_repo.update_status(
            importctrl_id=importctrl_id,
            status=ImportStatus.RUNNING,
            message=progress_msg,
        )
    except Exception:
        pass

    try:
        # Loop enquanto houver linhas de dados (linha 2 em diante)
        while True:
            max_row = _ws_max_data_row(ws)
            first_data_row = 2
            last_data_row = max_row

            if last_data_row < first_data_row:
                # Sem mais linhas para processar
                break

            # Definimos o intervalo do chunk
            chunk_start = last_data_row - CHUNK_ROWS + 1
            if chunk_start < first_data_row:
                chunk_start = first_data_row
            chunk_end = last_data_row

            processed, success, fail = _process_chunk_inplace(
                ws=ws,
                start_row=chunk_start,
                end_row=chunk_end,
                headers=headers,
                file_name=file_name,
                log_repo=log_repo,
            )

            read_rows += processed
            ok_rows += success
            fail_rows += fail

            # Atualiza status periódico
            if read_rows - last_status_update >= STATUS_EVERY:
                try:
                    eta = _calc_eta(start_time, read_rows, total_rows)
                    progress_msg = _build_progress_message(
                        filename=file_name,
                        read_rows=read_rows,
                        ok_rows=ok_rows,
                        fail_rows=fail_rows,
                        total_rows=total_rows,
                        started_at=start_time,
                        eta_dt=eta,
                    )
                    import_ctrl_repo.update_status(
                        importctrl_id=importctrl_id,
                        status=ImportStatus.RUNNING,
                        message=progress_msg,
                    )
                except Exception:
                    pass
                last_status_update = read_rows

        # Salva modificações IN PLACE
        wb.save(str(path))

    except Exception as e:
        msg = (
            f"Erro durante processamento IN PLACE em chunks "
            f"(linhas_lidas_ate_agora={read_rows}, sucesso={ok_rows}, falha={fail_rows}): {e}"
        )
        logger.exception(msg)
        log_repo.create(
            source="IMPORT_NTT_ORACLE",
            file=file_name,
            row=0,
            message=msg,
        )

        try:
            wb.save(str(path))
        except Exception as e_save:
            logger.exception("Erro ao salvar workbook IN PLACE após falha: %s", e_save)
            log_repo.create(
                source="IMPORT_NTT_ORACLE",
                file=file_name,
                row=0,
                message=f"Erro ao salvar Excel (in place) após falha: {e_save}",
            )

        try:
            wb.close()
        except Exception:
            pass

        fail_msg = (
            f"Importação NTT Oracle (IN PLACE) falhou após ler {read_rows} "
            f"linhas (sucesso={ok_rows}, falha={fail_rows}): {e}. "
            f"Arquivo de entrada mantido com apenas linhas não importadas ou com falha."
        )
        import_ctrl_repo.fail_import(importctrl_id, message=fail_msg)
        return {"status": "FAILED", "message": fail_msg}
    finally:
        try:
            wb.close()
        except Exception:
            pass

    tempo_execucao = (datetime.now() - start_time).total_seconds()
    linhas_restantes = max(_ws_max_data_row(ws) - 1, 0)

    # 7) Atualizar status final na tbImportControl
    try:
        if linhas_restantes > 0:
            final_msg = (
                f"Importação NTT Oracle (IN PLACE) concluída com pendências. "
                f"Total linhas estimadas: {total_rows}, "
                f"Lidas: {read_rows}, "
                f"Sucesso: {ok_rows}, "
                f"Falha: {fail_rows}, "
                f"Restantes_no_arquivo={linhas_restantes}, "
                f"Tempo: {tempo_execucao:.2f}s. "
                f"Arquivo de entrada atualizado IN PLACE contendo apenas linhas "
                f"não lidas ou com falha."
            )
            import_ctrl_repo.finish_import(
                importctrl_id=importctrl_id,
                message=final_msg,
            )
            return {"status": "PENDING", "message": final_msg}

        final_msg = (
            f"Importação NTT Oracle (IN PLACE) concluída com sucesso. "
            f"Total processado (estimado): {total_rows or read_rows}. "
            f"Sucesso: {ok_rows}, Falha: {fail_rows}. "
            f"Tempo: {tempo_execucao:.2f}s. "
            f"Todas as linhas foram removidas do arquivo de entrada "
            f"(nenhuma pendência restante)."
        )
        import_ctrl_repo.finish_import(
            importctrl_id=importctrl_id,
            message=final_msg,
        )
        return {"status": "FINISHED", "message": final_msg}
    except Exception as e:
        fail_msg = (
            f"Falha ao atualizar status final da importação (IN PLACE): {e}. "
            f"Arquivo de entrada mantém apenas linhas não processadas ou com falha."
        )
        logger.exception(fail_msg)
        import_ctrl_repo.fail_import(importctrl_id, message=fail_msg)
        return {"status": "FAILED", "message": fail_msg}
