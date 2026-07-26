"""
cisco_enterprise_agreement_usage_fetcher.py

Importador Cisco Enterprise Agreement (EA)
Padrão operacional alinhado ao cisco_smart_account_usage_fetcher.py

Características:
- Log operacional em storage/logs/<arquivo>.log
- Arquivo fixo de falhas em storage/output/<arquivo>_failed_rows.xlsx
- Toda linha lida sai do arquivo original
- Arquivo original é regravado apenas com linhas remanescentes
- Lock simples por arquivo
"""

from __future__ import annotations

import os
import re
import traceback
import unicodedata
from datetime import datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from src.infrastructure.database.repositories.product_repository import ProductRepository
from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository
from src.infrastructure.database.repositories.cisco_ea_repository import CiscoEARepository
from src.infrastructure.database.repositories.task_repository import TaskRepository

try:
    from src.infrastructure.database.repositories.import_control_repository import (
        ImportControlRepository,
        ImportStatus,
    )
except Exception:
    ImportControlRepository = None
    ImportStatus = None


# =============================================================================
# PATHS
# =============================================================================

BASE_STORAGE_PATH = Path("storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"
BASE_LOCKS_PATH = BASE_STORAGE_PATH / "locks"

IMPORT_SOURCE = "CiscoEnterpriseAgreementUsageFetcher"

repo_ea = CiscoEARepository()
repo_product = ProductRepository()
repo_company = CompanyListNameRepository()
repo_import_log = ImportLogRepository()
repo_task = TaskRepository()
repo_import_control = ImportControlRepository() if ImportControlRepository else None


# =============================================================================
# DIRECTORY / LOG / LOCK
# =============================================================================

def _ensure_directories():
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOCKS_PATH.mkdir(parents=True, exist_ok=True)


def _timestamp():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _build_log_path(source_file: Path):
    return BASE_LOGS_PATH / f"{source_file.stem}.log"


def _append_log(log_path: Path, message: str):
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{_timestamp()}] {message}\n")


def _report_progress(
    import_control_id: Optional[int],
    total_rows: int,
    success_rows: int,
    failed_rows: int,
    remaining_rows: int,
    log_path: Path,
) -> None:
    if repo_import_control is None or import_control_id is None:
        return

    lidas = total_rows - remaining_rows
    msg = (
        f"Arquivo processado. "
        f"Total linhas origem={total_rows} | "
        f"lidas={lidas} | "
        f"sucesso={success_rows} | "
        f"falhas={failed_rows} | "
        f"restantes={remaining_rows}"
    )

    try:
        repo_import_control.update_status(
            importctrl_id=import_control_id,
            status=ImportStatus.RUNNING,
            message=msg,
        )
    except Exception as e:
        _append_log(log_path, f"WARN failed updating import control progress error={str(e)[:500]}")


def _build_lock_path(source_file: Path):
    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", source_file.name)
    return BASE_LOCKS_PATH / f"{safe_name}.lock"


def _acquire_lock(lock_path: Path):
    fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
    os.close(fd)


def _release_lock(lock_path: Path):
    if lock_path.exists():
        lock_path.unlink(missing_ok=True)


# =============================================================================
# HELPERS
# =============================================================================

def _normalize_str(value: Any) -> str:
    return str(value).strip() if value else ""


def _remove_accent_upper(text: str) -> str:
    normalized = unicodedata.normalize("NFKD", text)
    no_accent = "".join(c for c in normalized if not unicodedata.combining(c))
    return no_accent.upper()


def _to_int(value: Any) -> int:
    try:
        return int(float(str(value).replace(",", ".")))
    except Exception:
        return 0


def _to_date(value: Any):
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(value), fmt).date().isoformat()
        except Exception:
            continue
    return None


# =============================================================================
# RESOLUTION
# =============================================================================

def _resolve_customer_id(name: str, cache: Dict[str, Optional[int]]):
    key = _remove_accent_upper(_normalize_str(name))
    if key in cache:
        return cache[key]
    customer_id = repo_company.get_company_id_by_name(key)
    cache[key] = customer_id
    return customer_id


def _resolve_product_id(sku: str, cache: Dict[str, Optional[int]]):
    sku = _normalize_str(sku)
    if sku in cache:
        return cache[sku]

    existing = repo_product.find_ids_by({
        "product_vendor_id": 1,
        "product_name": sku
    })

    if existing:
        product_id = existing[0]
    else:
        product_id = repo_product.insert({
            "product_vendor_id": 1,
            "product_name": sku,
            "product_part_number": sku
        })

    cache[sku] = product_id
    return product_id


# =============================================================================
# PROCESSING
# =============================================================================

def _read_excel(source_file: Path):
    wb = load_workbook(source_file, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    header = [str(h).strip() for h in next(rows)]
    data = []
    for idx, values in enumerate(rows, start=2):
        row = {header[i]: values[i] if i < len(values) else None for i in range(len(header))}
        if any(v not in (None, "") for v in row.values()):
            data.append((idx, row))
    wb.close()
    return header, data


def _write_remaining(source_file: Path, header, remaining_rows):
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = Path(tmp.name)

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in remaining_rows:
        ws.append([row.get(col) for col in header])
    wb.save(temp_path)
    wb.close()

    # Em alguns ambientes (Linux + mount diferente) os.replace gera:
    # OSError: Invalid cross-device link
    # Portanto utilizamos shutil.move, que é seguro entre devices.
    import shutil
    shutil.move(str(temp_path), str(source_file))


def _build_failed_file(source_file: Path):
    return BASE_OUTPUT_PATH / f"{source_file.stem}_failed_rows.xlsx"


def _append_failed(failed_file: Path, header, failed_rows):
    if not failed_rows:
        return

    wb = Workbook()
    ws = wb.active
    failed_header = header + ["import_error", "import_original_row"]
    ws.append(failed_header)

    for row in failed_rows:
        ws.append([row.get(col) for col in header] + [row["error"], row["row_number"]])

    wb.save(failed_file)
    wb.close()


# =============================================================================
# MAIN
# =============================================================================

def run_import(file_name: Optional[str] = None, user_id: Optional[str] = None, import_control_id: Optional[int] = None):

    _ensure_directories()

    if not file_name:
        return {"status": "WARNING", "message": "Arquivo não informado"}

    source_file = BASE_INPUT_PATH / file_name
    if not source_file.exists():
        return {"status": "ERROR", "message": "Arquivo não encontrado"}

    log_path = _build_log_path(source_file)
    lock_path = _build_lock_path(source_file)

    if import_control_id is None and repo_import_control is not None:
        try:
            import_control_id = repo_import_control.get_id_by_file(source_file.name)
        except Exception:
            import_control_id = None

    _acquire_lock(lock_path)

    try:
        _append_log(log_path, f"Iniciando processamento de {file_name}")

        header, rows = _read_excel(source_file)

        _append_log(log_path, f"Linhas lidas do Excel: {len(rows)}")

        # Validação mínima de colunas obrigatórias
        required_columns = {"Client", "SKU", "Balance"}
        missing_columns = required_columns - set(header)
        if missing_columns:
            raise ValueError(f"Colunas obrigatórias ausentes no arquivo: {missing_columns}")

        customer_cache = {}
        product_cache = {}
        failed_rows = []
        success_rows = set()
        inserted_count = 0

        for row_number, row in rows:
            try:
                client_id = _resolve_customer_id(row["Client"], customer_cache)
                if not client_id:
                    raise ValueError("Cliente não encontrado")

                product_id = _resolve_product_id(row["SKU"], product_cache)
                if not product_id:
                    raise ValueError("Produto não encontrado")

                # Payload completo para atender possíveis constraints NOT NULL
                payload = {
                    "mcea_client_id": client_id,
                    "mcea_client": _normalize_str(row.get("Client")),
                    "mcea_domain": _normalize_str(row.get("Domain")),
                    "mcea_virtual_account": _normalize_str(row.get("Virtual Account")),
                    "mcea_subscription": _normalize_str(row.get("Subscription")),
                    "mcea_status": _normalize_str(row.get("Status")),
                    "mcea_suite_name": _normalize_str(row.get("Suite Name")),
                    "mcea_calculation_method": _normalize_str(row.get("Calculation Method")),
                    "mcea_product_id": product_id,
                    "mcea_sku": _normalize_str(row.get("SKU")),
                    "mcea_purchased": _to_int(row.get("Purchased")),
                    "mcea_growth_allowance": _to_int(row.get("Growth Allowance")),
                    "mcea_total_purchased": _to_int(row.get("Total Purchased")),
                    "mcea_generated": _to_int(row.get("Generated")),
                    "mcea_balance": _to_int(row.get("Balance")),
                    "mcea_pre_ea": _to_int(row.get("Pre EA")),
                    "mcea_license_migrated": _to_int(row.get("License Migrated")),
                    "mcea_ntf_date": _to_date(row.get("NTF Date")),
                    "mcea_start_date": _to_date(row.get("Start Date")),
                    "mcea_end_date": _to_date(row.get("End Date")),
                    "mcea_update": datetime.now().date().isoformat(),
                    "mcea_track": 0,
                }

                identity = {k: v for k, v in payload.items() if k != "mcea_update"}
                existing = repo_ea.find_metering_first_by(identity)

                if not existing:
                    insert_id = repo_ea.insert_metering(payload)
                    if not insert_id:
                        raise RuntimeError("Insert não retornou ID válido")
                    inserted_count += 1

                success_rows.add(row_number)

            except Exception as e:
                failed_rows.append({
                    **row,
                    "error": str(e),
                    "row_number": row_number
                })

        # falhas
        failed_file = _build_failed_file(source_file)
        _append_failed(failed_file, header, failed_rows)

        # regrava original com não lidas
        read_rows = success_rows | {r["row_number"] for r in failed_rows}
        remaining = [row for row_number, row in rows if row_number not in read_rows]
        _write_remaining(source_file, header, remaining)

        summary = (
            f"Arquivo processado. "
            f"Total linhas origem={len(rows)} | "
            f"lidas={len(read_rows)} | "
            f"sucesso={len(success_rows)} | "
            f"falhas={len(failed_rows)} | "
            f"restantes={len(remaining)}"
        )

        _append_log(log_path, summary)

        _report_progress(
            import_control_id=import_control_id,
            total_rows=len(rows),
            success_rows=len(success_rows),
            failed_rows=len(failed_rows),
            remaining_rows=len(remaining),
            log_path=log_path,
        )

        repo_import_log.create(
            source=IMPORT_SOURCE,
            file=file_name,
            row=0,
            message=summary
        )

        return {"status": "FINISHED", "message": summary}

    except Exception as e:
        _append_log(log_path, f"Erro fatal: {str(e)}")
        _append_log(log_path, traceback.format_exc())
        repo_import_log.create(
            source=IMPORT_SOURCE,
            file=file_name,
            row=0,
            message=str(e)
        )
        raise

    finally:
        _release_lock(lock_path)
