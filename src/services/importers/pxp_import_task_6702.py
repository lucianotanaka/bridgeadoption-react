"""
pxp_import_task_6702.py

Importador PXP do processo 6702.

Objetivo
--------
Processar arquivo XLSX de entrada contendo linhas PXP e refletir as regras
de negócio no banco de dados, com foco em criação, atualização e encerramento
de tasks e activities, além de geração de histórico e arquivos de falha.

Padrão adotado
--------------
Este importador segue a mesma linha operacional observada em
`cisco_subscription_ccw.py`:

- lê o arquivo XLSX diretamente;
- processa em chunks;
- remove do arquivo original toda linha lida;
- escreve linhas com erro em um arquivo fixo de falhas;
- gera log texto de execução;
- grava logs funcionais em tbImportLog quando disponível;
- registra histórico em tbTaskRecord.

Observações importantes
-----------------------
1. Por precaução:
   - usa adapter defensivo para a parte de UseCase / Solution Domain;
   - evita depender rigidamente de uma assinatura única externa.

2. Fallback como segurança:
   - solução/arquitetura via UseCaseRepository são opcionais;
   - owner segue default 93 quando não houver regra melhor.

3. Regras consolidadas incorporadas:
   - Incentive Eligibility Status = Eligible:
       * se não existir task -> cria task com task_status = 1
       * se existir task aberta -> não muda task_status, apenas atualiza campos comparáveis
   - Incentive Eligibility Status = Not Eligible:
       * se não existir task -> não cria task, não faz nada
       * se existir task aberta -> task_status = 5 e activities abertas/ativas viram 4
   - Incentive Eligibility Status = Expired:
       * se não existir task -> cria task com task_status = 6
       * se existir task aberta -> task_status = 6 e activities abertas/ativas viram 6

4. Task considerada encerrada:
   - task_status IN (4, 5, 6, 10)

5. Activity considerada encerrada:
   - activity_status IN (4, 5, 6, 10)

6. Sanitização numérica:
   - task_value e task_booking_amount são validados contra o range suportado
     pelo banco para DECIMAL(30,6)
   - quando o valor numérico vier inválido ou fora do range:
       * a linha NÃO é rejeitada
       * o campo é gravado como None
       * é gerado log funcional

7. Regra de Deal Id x Deal ID:
   - task_ws deve vir obrigatoriamente da coluna 'Deal Id'
   - task_deal_id deve vir da coluna 'Deal ID'
   - o código diferencia ambas as colunas de forma case sensitive
   - WS deve seguir o padrão 'WS-<número>'
   - sem WS válido a linha deve falhar
   - a busca de task existente também depende obrigatoriamente de WS

8. Regra de task type:
   - novas tasks criadas por este importador devem ser sempre task_tasktype_id = 22
   - na busca de task existente por WS, considerar task types 21 e 22
   - prioridade sempre para type 22 sobre 21
"""

from __future__ import annotations

import logging
import re
import traceback
from dataclasses import dataclass
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.database.repositories.task_repository import TaskRepository
from src.infrastructure.database.repositories.task_activity_repository import TaskActivityRepository
from src.infrastructure.database.repositories.task_history_repository import TaskHistoryRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository

try:
    from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
except Exception:
    ImportLogRepository = None

try:
    from src.infrastructure.database.repositories.use_case_repository import UseCaseRepository
except Exception:
    UseCaseRepository = None

try:
    from src.infrastructure.database.repositories.import_control_repository import (
        ImportControlRepository,
        ImportStatus,
    )
except Exception:
    ImportControlRepository = None
    ImportStatus = None


# ======================================================================================
# CONFIGURAÇÃO BASE
# ======================================================================================

BASE_STORAGE_PATH = Path("/home/bridgeadoption/storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"

IMPORT_SOURCE = "PxpImportTask6702"

ERROR_EXTRA_COLUMNS = [
    "import_error_message",
    "import_error_column",
    "import_error_value",
    "import_original_row",
    "import_processed_at",
    "import_source",
    "import_original_file",
]

logger = logging.getLogger(__name__)

repo_task = TaskRepository()
repo_activity = TaskActivityRepository()
repo_history = TaskHistoryRepository()
repo_company = CompanyListNameRepository()
repo_log = ImportLogRepository() if ImportLogRepository else None
repo_use_case = UseCaseRepository() if UseCaseRepository else None
repo_import_control = ImportControlRepository() if ImportControlRepository else None


# ======================================================================================
# CONSTANTES DE DOMÍNIO
# ======================================================================================

DEFAULT_TASK_OWNER_ID = 93
CHUNK_SIZE = 1000

# A cada quantos chunks o progresso é reportado na tbImportControl
IMPORT_CONTROL_PROGRESS_INTERVAL_CHUNKS = 1

SOURCE_STATUS_NOT_ELIGIBLE = 0
SOURCE_STATUS_ELIGIBLE = 1
SOURCE_STATUS_EXPIRED = 6

TASK_STATUS_OPEN = 1
TASK_STATUS_IN_PROGRESS = 2
TASK_STATUS_ON_HOLD = 3
TASK_STATUS_CANCELLED = 4
TASK_STATUS_DECLINED = 5
TASK_STATUS_EXPIRED = 6
TASK_STATUS_SUBMITTED = 7
TASK_STATUS_RESUBMITTED = 8
TASK_STATUS_APPROVED_TO_CLOSE = 9
TASK_STATUS_COMPLETED = 10

ACTIVITY_STATUS_CANCELLED = 4
ACTIVITY_STATUS_DECLINED = 5
ACTIVITY_STATUS_EXPIRED = 6

# Valor sentinela para Lifecycle Opt-In Status = "Opted Out"
OPT_IN_FLAG_OPTED_OUT = -1

CLOSED_TASK_STATUSES = {
    TASK_STATUS_CANCELLED,
    TASK_STATUS_DECLINED,
    TASK_STATUS_EXPIRED,
    TASK_STATUS_COMPLETED,
}

CLOSED_ACTIVITY_STATUSES = {4, 5, 6, 10}

VALID_EXISTING_TASK_TYPE_IDS = {21, 22}
PRIORITY_TASK_TYPE_ID = 22

REQUIRED_COLUMNS = [
    "Plan Name",
    "Deal Name",
    "CR Party Name",
    "Incentive Eligibility Status",
    "Track",
    "Sub-Track",
    "Deal Id",
    "Deal ID",
]

OPTIONAL_COLUMNS = [
    "WS",
    "Web Order ID",
    "Sales Order ID",
    "CR Party ID",
    "Eligible Incentive Activity Description and Estimated Incentive Value",
    "Deal Incentive Expiry Date",
    "Max Estimated Incentives Amount",
    "Max Estimated Incentives Amount (Currency)",
    "EA Flag",
    "Telemetry Flag",
    "Booking Date",
    "Lifecycle Opt-In Status",
    "Lifecycle Start Date",
    "Booking Amount - Net to Cisco",
    "Fund Type",
]

EXPECTED_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS

DECIMAL_30_6_MAX = Decimal("999999999999999999999999.999999")
DECIMAL_30_6_MIN = Decimal("-999999999999999999999999.999999")
DECIMAL_30_6_QUANTIZER = Decimal("0.000001")


# ======================================================================================
# DATACLASSES
# ======================================================================================

@dataclass
class RowProcessResult:
    success: bool
    error_message: Optional[str] = None
    error_column: Optional[str] = None
    error_value: Optional[Any] = None
    ignored: bool = False
    created: bool = False
    updated: bool = False
    cancelled: bool = False
    activities_cancelled: int = 0


# ======================================================================================
# HELPERS DE DIRETÓRIO / LOG
# ======================================================================================

def _ensure_directories() -> None:
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)


def _build_execution_log_path(input_path: Path) -> Path:
    return BASE_LOGS_PATH / f"{input_path.stem}.log"


def _report_progress(
    import_control_id: Optional[int],
    total_rows: int,
    processed_success: int,
    failed_rows: int,
    ignored_rows: int,
    remaining_rows: int,
    execution_log_path: Optional[Path] = None,
) -> None:
    """
    Atualiza importctrl_message na tbImportControl com o progresso atual.

    Chamada periodicamente durante o processamento em chunks para que
    quem monitora a tabela possa acompanhar o andamento sem esperar o fim.
    """
    if repo_import_control is None or import_control_id is None:
        return

    lidas = total_rows - remaining_rows
    msg = (
        f"Arquivo processado. "
        f"Total linhas origem={total_rows} | "
        f"lidas={lidas} | "
        f"sucesso={processed_success} | "
        f"erros={failed_rows} | "
        f"ignoradas={ignored_rows} | "
        f"restantes={remaining_rows}"
    )

    try:
        repo_import_control.update_status(
            importctrl_id=import_control_id,
            status=ImportStatus.RUNNING,
            message=msg,
        )
    except Exception as e:
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"WARN failed updating import control progress error={str(e)[:500]}",
            )


def _append_execution_log(log_path: Path, message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def _safe_log(
    file_path: str,
    row_number: int,
    message: str,
    column_name: Optional[str] = None,
    value: Optional[Any] = None,
) -> None:
    if repo_log is None:
        return

    try:
        repo_log.create(
            IMPORT_SOURCE,
            file_path,
            row_number,
            message,
            column_name,
            value,
        )
    except Exception as log_ex:
        logger.exception(
            "Falha ao gravar tbImportLog. file=%s row=%s column=%s value=%s error=%s",
            file_path,
            row_number,
            column_name,
            value,
            log_ex,
        )


# ======================================================================================
# HELPERS DE NORMALIZAÇÃO
# ======================================================================================

def _normalize_str(value: Any, max_length: int = 1000) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    return text[:max_length]


def _normalize_stage(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=100)
    if text is None:
        return None

    normalized = re.sub(r"\s+", "", text).lower()
    return normalized


def _normalize_ws(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=100)
    if not text:
        return None

    normalized = text.strip().upper()

    if not re.fullmatch(r"WS-\d+", normalized):
        return None

    return normalized


def _to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None

    if isinstance(value, Decimal):
        return value

    text = str(value).strip()
    if not text:
        return None

    text = re.sub(r"[^\d,\.\-+]", "", text)

    if not text or text in {"", ".", "-", "+", "-.", "+."}:
        return None

    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "")
            text = text.replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "")
        text = text.replace(",", ".")

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _to_float(value: Any) -> Optional[float]:
    dec = _to_decimal(value)
    return float(dec) if dec is not None else None


def _to_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()[:50]

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _parse_yes_no_flag(value: Any) -> int:
    text = _normalize_str(value, max_length=50)
    if not text:
        return 0

    normalized = re.sub(r"\s+", "", text).lower()

    if normalized in {"yes", "y", "true", "1"}:
        return -1

    if normalized in {"no", "n", "false", "0"}:
        return 0

    return 0


def _parse_opt_in_flag(value: Any) -> int:
    text = _normalize_str(value, max_length=100)
    if not text:
        return 0

    normalized = re.sub(r"\s+", "", text).lower()

    if normalized in {"optedin", "optedin.", "opted-in", "opted_in"}:
        return 1

    if normalized in {"optedout", "opted-out", "opted_out"}:
        return OPT_IN_FLAG_OPTED_OUT

    if normalized in {"pre-approved", "preapproved", "pending", "notoptedin"}:
        return 0

    return 0


def _clean_currency(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=50)
    if not text:
        return None

    text = text.replace("($)", "").replace("$", "").strip()
    return text or None


def _resolve_eligibility_status_to_source_status(status_value: Any) -> int:
    normalized = _normalize_stage(status_value)

    mapping = {
        "eligible": SOURCE_STATUS_ELIGIBLE,
        "noteligible": SOURCE_STATUS_NOT_ELIGIBLE,
        "expired": SOURCE_STATUS_EXPIRED,
    }

    if normalized not in mapping:
        raise ValueError(f"Incentive Eligibility Status inválido: {status_value}")

    return mapping[normalized]


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default

    if isinstance(value, bool):
        return default

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value) if value.is_integer() else default

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    if text.lstrip("+-").isdigit():
        try:
            return int(text)
        except Exception:
            return default

    return default


def _sanitize_decimal_30_6(value: Any) -> Optional[Decimal]:
    dec = _to_decimal(value)
    if dec is None:
        return None

    try:
        dec = dec.quantize(DECIMAL_30_6_QUANTIZER, rounding=ROUND_HALF_UP)
    except Exception:
        return None

    if dec < DECIMAL_30_6_MIN or dec > DECIMAL_30_6_MAX:
        return None

    return dec


def _sanitize_task_value(value: Any) -> Optional[Decimal]:
    return _sanitize_decimal_30_6(value)


def _sanitize_task_booking_amount(value: Any) -> Optional[Decimal]:
    return _sanitize_decimal_30_6(value)


def _normalize_compare_string(value: Any) -> Optional[str]:
    return _normalize_str(value, max_length=4000)


def _normalize_compare_decimal(value: Any) -> Optional[Decimal]:
    dec = _to_decimal(value)
    if dec is None:
        return None

    try:
        return dec.normalize()
    except Exception:
        return dec


def _normalize_compare_date(value: Any) -> Optional[date]:
    return _to_date(value)


def _values_different(field: str, old_val: Any, new_val: Any) -> bool:
    int_fields = {
        "task_status",
        "task_customer_id",
        "task_cr_party_id",
        "task_opt_in_flag",
        "task_telemetry_flag",
        "task_ea_flag",
        "task_owner_id",
    }

    decimal_fields = {
        "task_value",
        "task_booking_amount",
    }

    date_fields = {
        "task_start",
        "task_start_performed",
        "task_end",
        "task_booking_date",
    }

    string_fields = {
        "task_reference",
        "task_cr_party_name",
        "task_currency",
        "task_track",
        "task_subtrack",
        "task_description",
        "task_architecture",
        "task_solution_domain",
        "task_deal_id",
        "task_ws",
        "task_eligible",
    }

    if old_val is None and new_val is None:
        return False

    if field in int_fields:
        return _safe_int(old_val, default=0) != _safe_int(new_val, default=0)

    if field in decimal_fields:
        return _normalize_compare_decimal(old_val) != _normalize_compare_decimal(new_val)

    if field in date_fields:
        return _normalize_compare_date(old_val) != _normalize_compare_date(new_val)

    if field in string_fields:
        return _normalize_compare_string(old_val) != _normalize_compare_string(new_val)

    return old_val != new_val


# ======================================================================================
# HELPERS DE ARQUIVO EXCEL
# ======================================================================================

def _resolve_input_path(file_path: str) -> Path:
    path = Path(file_path)
    if path.is_absolute():
        return path
    return BASE_INPUT_PATH / path.name


def _build_failed_output_path(input_path: Path) -> Path:
    return BASE_OUTPUT_PATH / f"{input_path.stem}_failed_rows.xlsx"


def _open_workbook_rw(file_path: str) -> Tuple[Any, Worksheet, Dict[str, int], List[str]]:
    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx_map = {h: i for i, h in enumerate(headers)}
    return wb, ws, idx_map, headers


def _ensure_failed_workbook(
    failed_path: Path,
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
) -> None:
    """
    Cria (ou recria) o arquivo de falhas com apenas o header.

    Sempre sobrescreve o arquivo se já existir, garantindo que não haja
    resíduos de execuções anteriores. A criação só será feita de fato
    se esta função for chamada (ou seja, somente quando houver falhas
    a serem escritas).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "failed_rows"
    ws.append(original_headers + ERROR_EXTRA_COLUMNS)
    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(execution_log_path, f"INFO failed file created (overwritten if existed): {str(failed_path)}")


def _append_failed_rows(
    failed_path: Path,
    rows_to_append: List[List[Any]],
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
) -> None:
    """
    Acrescenta linhas de falha ao arquivo de falhas.

    - Se rows_to_append estiver vazio, não faz nada e não cria arquivo.
    - Se o arquivo ainda não existir, ele é criado sobrescrevendo
      qualquer conteúdo anterior (via _ensure_failed_workbook).
    """
    if not rows_to_append:
        return

    # Garante criação/sobrescrita do arquivo apenas quando houver linhas de falha
    _ensure_failed_workbook(failed_path, original_headers, execution_log_path)

    wb = load_workbook(str(failed_path))
    ws = wb.active

    for row in rows_to_append:
        ws.append(row)

    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO appended failed rows count={len(rows_to_append)} to file={str(failed_path)}",
        )


def _row_dict_to_failed_excel_row(
    headers: List[str],
    row_dict: Dict[str, Any],
    original_row_number: int,
    input_file_path: str,
    result: RowProcessResult,
) -> List[Any]:
    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    original_values = [row_dict.get(h) for h in headers]

    return original_values + [
        result.error_message,
        result.error_column,
        str(result.error_value)[:2000] if result.error_value is not None else None,
        original_row_number,
        processed_at,
        IMPORT_SOURCE,
        input_file_path,
    ]


# ======================================================================================
# RESOLUÇÃO DE NEGÓCIO
# ======================================================================================

def _validate_required_columns(headers: List[str]) -> List[str]:
    return [col for col in REQUIRED_COLUMNS if col not in headers]


def _normalize_row(row_dict: Dict[str, Any]) -> Dict[str, Any]:
    deal_name = _normalize_str(row_dict.get("Deal Name"))
    explicit_deal_id = _normalize_str(row_dict.get("Deal ID"))
    eligibility_status_raw = _normalize_str(row_dict.get("Incentive Eligibility Status"))

    normalized = {
        "plan_name": _normalize_str(row_dict.get("Plan Name")),
        "ws": _normalize_ws(row_dict.get("Deal Id")),
        "deal_id": explicit_deal_id,
        "deal_name": deal_name,
        "web_order_id": _normalize_str(row_dict.get("Web Order ID")),
        "sales_order_id": _normalize_str(row_dict.get("Sales Order ID")),
        "cr_party_id": _normalize_str(row_dict.get("CR Party ID")),
        "cr_party_name": _normalize_str(row_dict.get("CR Party Name")),
        "eligibility_status_raw": eligibility_status_raw,
        "description": _normalize_str(
            row_dict.get("Eligible Incentive Activity Description and Estimated Incentive Value"),
            max_length=2000,
        ),
        "end_date": _to_date(row_dict.get("Deal Incentive Expiry Date")),
        "amount": _to_decimal(row_dict.get("Max Estimated Incentives Amount")),
        "currency": _clean_currency(row_dict.get("Max Estimated Incentives Amount (Currency)")),
        "track": _normalize_str(row_dict.get("Track")),
        "subtrack": _normalize_str(row_dict.get("Sub-Track")),
        "ea_flag": _parse_yes_no_flag(row_dict.get("EA Flag")),
        "telemetry_flag": _parse_yes_no_flag(row_dict.get("Telemetry Flag")),
        "booking_date": _to_date(row_dict.get("Booking Date")),
        "opt_in_flag": _parse_opt_in_flag(row_dict.get("Lifecycle Opt-In Status")),
        "lifecycle_start_date": _to_date(row_dict.get("Lifecycle Start Date")),
        "booking_amount": _to_decimal(row_dict.get("Booking Amount - Net to Cisco")),
        "fund_type": _normalize_str(row_dict.get("Fund Type")),
        "source_status": _resolve_eligibility_status_to_source_status(row_dict.get("Incentive Eligibility Status")),
    }

    return normalized


def _validate_minimum_required_fields(data: Dict[str, Any]) -> Tuple[bool, Optional[str], Optional[Any]]:
    """
    Valida os campos obrigatórios para que o processamento da linha possa prosseguir.

    IMPORTANTE: 'Deal ID' (coluna task_deal_id) NÃO é validado aqui porque a busca
    de task existente pelo WS (coluna 'Deal Id') deve ocorrer primeiro.
    'Deal ID' só é obrigatório quando não existe task na base e uma nova task
    precisa ser criada — essa validação é feita em _validate_required_for_creation.
    """
    required_map = {
        "Plan Name": data.get("plan_name"),
        "CR Party Name": data.get("cr_party_name"),
        "Incentive Eligibility Status": data.get("eligibility_status_raw"),
        "Track": data.get("track"),
        "Sub-Track": data.get("subtrack"),
        "Deal Id": data.get("ws"),
    }

    for col, value in required_map.items():
        if value is None or str(value).strip() == "":
            return False, col, value

    return True, None, None


def _validate_required_for_creation(data: Dict[str, Any]) -> Tuple[bool, Optional[str], Optional[Any]]:
    """
    Valida campos que são obrigatórios especificamente para CRIAÇÃO de nova task.

    'Deal ID' é exigido na criação porque task_deal_id precisa ser populado.
    Quando a task já existe (buscada pelo WS), este campo pode estar ausente
    no Excel sem impedir o processamento.
    """
    deal_id = data.get("deal_id")
    if deal_id is None or str(deal_id).strip() == "":
        return False, "Deal ID", deal_id

    return True, None, None


def _build_task_reference(data: Dict[str, Any]) -> Optional[str]:
    parts: List[str] = []

    if data.get("web_order_id"):
        parts.append(str(data["web_order_id"]).strip())

    if data.get("sales_order_id"):
        parts.append(str(data["sales_order_id"]).strip())

    if data.get("cr_party_id"):
        parts.append(f"CR Party ID: {str(data['cr_party_id']).strip()}")

    if not parts:
        return None

    return "; ".join(parts)


def _resolve_company_id(company_name: Optional[str]) -> int:
    if not company_name:
        return 0

    try:
        company_id = repo_company.get_company_id_by_name(company_name)
        return _safe_int(company_id, default=0)
    except Exception:
        return 0


def _resolve_task_type_id(data: Dict[str, Any]) -> int:
    return PRIORITY_TASK_TYPE_ID


def _resolve_solution_and_architecture(subtrack: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    if not subtrack or repo_use_case is None:
        return None, None

    try:
        if not hasattr(repo_use_case, "get_solution_by_use_case"):
            return None, None

        result = repo_use_case.get_solution_by_use_case(subtrack)

        if result is None:
            return None, None

        if isinstance(result, dict):
            architecture = (
                result.get("uc_architecture")
                or result.get("architecture")
                or result.get("task_architecture")
            )
            solution_domain = (
                result.get("uc_solution_domain")
                or result.get("solution_domain")
                or result.get("task_solution_domain")
                or result.get("solution")
            )
            return _normalize_str(solution_domain), _normalize_str(architecture)

        if isinstance(result, (list, tuple)) and result:
            first = result[0]
            if isinstance(first, dict):
                architecture = (
                    first.get("uc_architecture")
                    or first.get("architecture")
                    or first.get("task_architecture")
                )
                solution_domain = (
                    first.get("uc_solution_domain")
                    or first.get("solution_domain")
                    or first.get("task_solution_domain")
                    or first.get("solution")
                )
                return _normalize_str(solution_domain), _normalize_str(architecture)

            if isinstance(first, str):
                return _normalize_str(first), None

        if isinstance(result, str):
            return _normalize_str(result), None

    except Exception:
        return None, None

    return None, None


def _get_task_columns_for_match(task_id: int) -> Optional[Dict[str, Any]]:
    required_columns = [
        "task_id",
        "task_tasktype_id",
        "task_status",
        "task_owner_id",
        "task_opt_in_flag",
        "task_telemetry_flag",
        "task_ea_flag",
        "task_reference",
        "task_customer_id",
        "task_cr_party_id",
        "task_cr_party_name",
        "task_end",
        "task_value",
        "task_currency",
        "task_track",
        "task_subtrack",
        "task_description",
        "task_booking_date",
        "task_booking_amount",
        "task_architecture",
        "task_solution_domain",
        "task_deal_id",
        "task_ws",
        "task_start",
        "task_start_performed",
        "task_eligible",
    ]

    try:
        if hasattr(repo_task, "get_columns_by_task_id"):
            row = repo_task.get_columns_by_task_id(task_id, required_columns, as_df=False)

            if isinstance(row, dict):
                return row

            if isinstance(row, list) and row:
                if isinstance(row[0], dict):
                    return row[0]

        rows = repo_task.get_task(task_id=task_id, as_df=False)
        if isinstance(rows, list) and rows:
            if isinstance(rows[0], dict):
                return rows[0]

    except Exception:
        return None

    return None


def _resolve_existing_task(data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    ws = data.get("ws")

    if not ws:
        return None

    try:
        found_ids = repo_task.find_ids_by({"task_ws": ws})
    except Exception:
        return None

    if not found_ids:
        return None

    matched_tasks: List[Dict[str, Any]] = []

    for found_id in found_ids:
        task_id = _safe_int(found_id, default=0)
        if not task_id:
            continue

        task_row = _get_task_columns_for_match(task_id)
        if not task_row:
            continue

        task_type_id = _safe_int(task_row.get("task_tasktype_id"), default=0)

        if task_type_id in VALID_EXISTING_TASK_TYPE_IDS:
            matched_tasks.append(task_row)

    if len(matched_tasks) == 0:
        return None

    type_22_tasks = [
        t for t in matched_tasks
        if _safe_int(t.get("task_tasktype_id"), default=0) == 22
    ]

    type_21_tasks = [
        t for t in matched_tasks
        if _safe_int(t.get("task_tasktype_id"), default=0) == 21
    ]

    if len(type_22_tasks) == 1:
        return type_22_tasks[0]

    if len(type_22_tasks) > 1:
        raise ValueError(
            f"Mais de uma task tipo 22 encontrada para o mesmo WS {ws}. "
            f"task_ids={[t.get('task_id') for t in type_22_tasks]}"
        )

    if len(type_21_tasks) == 1:
        return type_21_tasks[0]

    if len(type_21_tasks) > 1:
        raise ValueError(
            f"Mais de uma task tipo 21 encontrada para o mesmo WS {ws}. "
            f"task_ids={[t.get('task_id') for t in type_21_tasks]}"
        )

    raise ValueError(
        f"Mais de uma task encontrada para o mesmo WS {ws}. "
        f"task_ids={[t.get('task_id') for t in matched_tasks]}"
    )


def _build_task_payload(
    data: Dict[Any, Any],
    company_id: int,
    task_type_id: int,
    force_status: Optional[int] = None,
    force_status_justification: Optional[str] = None,
) -> Dict[str, Any]:
    task_start = data.get("booking_date")
    if data.get("opt_in_flag") == 1 and data.get("lifecycle_start_date"):
        task_start = data.get("lifecycle_start_date")

    solution_domain, architecture = _resolve_solution_and_architecture(data.get("subtrack"))

    if force_status is not None:
        initial_status = force_status
    else:
        initial_status = TASK_STATUS_OPEN
        if data.get("source_status") == SOURCE_STATUS_EXPIRED:
            initial_status = TASK_STATUS_EXPIRED
        elif data.get("source_status") == SOURCE_STATUS_ELIGIBLE and data.get("opt_in_flag") == 1:
            initial_status = TASK_STATUS_IN_PROGRESS

    payload = {
        "task_tasktype_id": task_type_id,
        "task_reference": _build_task_reference(data),
        "task_owner_id": DEFAULT_TASK_OWNER_ID,
        "task_temp_owner_id": None,
        "task_customer_id": company_id,
        "task_cr_party_id": _safe_int(data.get("cr_party_id"), default=0),
        "task_cr_party_name": data.get("cr_party_name"),
        "task_created_in": datetime.now(),
        "task_created_by": 0,
        "task_priority": "LOW",
        "task_project_id": 0,
        "task_status": initial_status,
        "task_status_justification": force_status_justification,
        "task_start": task_start,
        "task_end": data.get("end_date"),
        "task_start_performed": task_start,
        "task_end_performed": data.get("end_date"),
        "task_value": data.get("amount") if data.get("amount") is not None else Decimal("0.000000"),
        "task_forecast": 0,
        "task_backlog": 0,
        "task_rate": 1,
        "task_currency": data.get("currency") or "USD",
        "task_ws": data.get("ws"),
        "task_deal_id": data.get("deal_id"),
        "task_track": data.get("track"),
        "task_subtrack": data.get("subtrack"),
        "task_highlight": 0,
        "task_remark": None,
        "task_description": data.get("description"),
        "task_ea_flag": data.get("ea_flag", 0),
        "task_telemetry_flag": data.get("telemetry_flag", 0),
        "task_opt_in_flag": data.get("opt_in_flag", 0),
        "task_completed": 0,
        "task_architecture": architecture,
        "task_solution_domain": solution_domain,
        "task_eligible": "Y" if data.get("source_status") != SOURCE_STATUS_NOT_ELIGIBLE else "N",
        "task_end_fy": None,
        "task_booking_date": data.get("booking_date"),
        "task_booking_amount": data.get("booking_amount"),
    }

    return payload


def _insert_task_history(
    task_id: int,
    remark: str,
    updated_by: str = "System BA",
    activity_id: int = 0,
    next_followup: Optional[date] = None,
) -> None:
    history = {
        "taskrecord_task_id": task_id,
        "taskrecord_activity_id": activity_id,
        "taskrecord_remark": remark,
        "taskrecord_updated_by": updated_by,
    }

    if next_followup is not None:
        history["taskrecord_next_followup"] = next_followup.strftime("%Y-%m-%d")

    repo_history.insert(history)


def _create_task_from_payload(payload: Dict[str, Any]) -> int:
    task_id = repo_task.insert(payload)
    return _safe_int(task_id, default=0)


def _close_open_activities_with_status(
    task_id: int,
    new_activity_status: int,
    history_remark: str,
    updated_by: str = "System BA",
) -> int:
    changed = 0

    try:
        activities = repo_activity.get_activity(task_id=task_id, as_df=False)
    except Exception:
        return 0

    for activity in activities:
        activity_id = _safe_int(activity.get("activity_id"), default=0)
        current_status = _safe_int(activity.get("activity_status"), default=0)

        if not activity_id:
            continue

        if current_status in CLOSED_ACTIVITY_STATUSES:
            continue

        try:
            repo_activity.update(
                data={"activity_status": new_activity_status},
                where={"activity_id": activity_id},
            )
            changed += 1

            _insert_task_history(
                task_id=task_id,
                activity_id=activity_id,
                remark=history_remark,
                updated_by=updated_by,
            )
        except Exception:
            continue

    return changed


def _close_task_as_not_eligible(task_id: int, updated_by: str = "System BA") -> int:
    try:
        repo_task.update(
            data={
                "task_status": TASK_STATUS_DECLINED,
                "task_eligible": "N",
                "task_forecast": 0,
                "task_status_justification": "TASK NOT ELIGIBLE BY VENDOR",
            },
            where={"task_id": task_id},
        )

        _insert_task_history(
            task_id=task_id,
            activity_id=0,
            remark="Task status changed to Declined; Task marked as not eligible by Cisco",
            updated_by=updated_by,
        )

        return _close_open_activities_with_status(
            task_id=task_id,
            new_activity_status=ACTIVITY_STATUS_CANCELLED,
            history_remark="Activity status changed to Cancelled because task changed to Not Eligible by Cisco",
            updated_by=updated_by,
        )
    except Exception:
        return 0


def _close_task_as_opted_out(
    task_id: int,
    opt_out_remark: str,
    updated_by: str = "System BA",
) -> int:
    """
    Fecha a task com task_status = 4 (CANCELLED) e
    task_status_justification = 'AN OPT-OUT WAS PERFORMED FOR THE TASK'.

    Também fecha todas as activities abertas com activity_status = 4 (CANCELLED),
    registrando histórico individual em cada uma.

    Retorna o número de activities canceladas.
    """
    try:
        repo_task.update(
            data={
                "task_status": TASK_STATUS_CANCELLED,
                "task_forecast": 0,
                "task_status_justification": "AN OPT-OUT WAS PERFORMED FOR THE TASK",
            },
            where={"task_id": task_id},
        )

        _insert_task_history(
            task_id=task_id,
            activity_id=0,
            remark=opt_out_remark,
            updated_by=updated_by,
        )

        return _close_open_activities_with_status(
            task_id=task_id,
            new_activity_status=ACTIVITY_STATUS_CANCELLED,
            history_remark=opt_out_remark,
            updated_by=updated_by,
        )
    except Exception:
        return 0


def _close_task_as_expired(task_id: int, updated_by: str = "System BA") -> int:
    try:
        repo_task.update(
            data={
                "task_status": TASK_STATUS_EXPIRED,
                "task_forecast": 0,
            },
            where={"task_id": task_id},
        )

        _insert_task_history(
            task_id=task_id,
            activity_id=0,
            remark="Task status changed to Expired",
            updated_by=updated_by,
        )

        return _close_open_activities_with_status(
            task_id=task_id,
            new_activity_status=ACTIVITY_STATUS_EXPIRED,
            history_remark="Activity status changed to Expired because task changed to Expired",
            updated_by=updated_by,
        )
    except Exception:
        return 0


def _update_existing_task(
    task_id: int,
    existing_task: Dict[str, Any],
    payload: Dict[str, Any],
    updated_by: str = "System BA",
    source_status: int = SOURCE_STATUS_ELIGIBLE,
    opt_in_flag: int = 0,
) -> bool:
    updates: Dict[str, Any] = {}
    remarks: List[str] = []

    current_status = _safe_int(existing_task.get("task_status"), default=0)

    if current_status in CLOSED_TASK_STATUSES:
        return False

    # Regra: Eligible + Opted In + task_status == 1 (Open) → promover para 2 (In Progress)
    if (
        source_status == SOURCE_STATUS_ELIGIBLE
        and opt_in_flag == 1
        and current_status == TASK_STATUS_OPEN
    ):
        updates["task_status"] = TASK_STATUS_IN_PROGRESS
        remarks.append("Change Task Status to In Progress (Opted In)")

    current_opt_in = _safe_int(existing_task.get("task_opt_in_flag"), default=0)
    current_telemetry = _safe_int(existing_task.get("task_telemetry_flag"), default=0)
    current_ea = _safe_int(existing_task.get("task_ea_flag"), default=0)
    current_owner = _safe_int(existing_task.get("task_owner_id"), default=0)

    new_opt_in = payload.get("task_opt_in_flag", current_opt_in)
    if _values_different("task_opt_in_flag", current_opt_in, new_opt_in):
        updates["task_opt_in_flag"] = new_opt_in
        remarks.append(f"Change Opt-In Flag to {new_opt_in}")

        if payload.get("task_start") and _values_different(
            "task_start",
            existing_task.get("task_start"),
            payload.get("task_start"),
        ):
            updates["task_start"] = payload.get("task_start")
            remarks.append(f"Change Task Start to {payload.get('task_start')}")

        if payload.get("task_start") and _values_different(
            "task_start_performed",
            existing_task.get("task_start_performed"),
            payload.get("task_start"),
        ):
            updates["task_start_performed"] = payload.get("task_start")
            remarks.append(f"Change Task Start Performed to {payload.get('task_start')}")

    new_telemetry = payload.get("task_telemetry_flag", current_telemetry)
    if _values_different("task_telemetry_flag", current_telemetry, new_telemetry):
        updates["task_telemetry_flag"] = new_telemetry
        remarks.append(f"Change Telemetry Flag to {new_telemetry}")

    new_ea = payload.get("task_ea_flag", current_ea)
    if _values_different("task_ea_flag", current_ea, new_ea):
        updates["task_ea_flag"] = new_ea
        remarks.append(f"Change EA Flag to {new_ea}")

    payload_owner = _safe_int(payload.get("task_owner_id"), default=0)
    if current_owner == 0 and payload_owner != 0:
        if _values_different("task_owner_id", current_owner, payload_owner):
            updates["task_owner_id"] = payload.get("task_owner_id")
            remarks.append(f"Change Task Owner to {payload.get('task_owner_id')}")

    simple_compare_fields = [
        "task_reference",
        "task_customer_id",
        "task_cr_party_id",
        "task_cr_party_name",
        "task_end",
        "task_value",
        "task_currency",
        "task_track",
        "task_subtrack",
        "task_description",
        "task_booking_date",
        "task_booking_amount",
        "task_architecture",
        "task_solution_domain",
        "task_deal_id",
        "task_ws",
    ]

    for field in simple_compare_fields:
        old_val = existing_task.get(field)
        new_val = payload.get(field)

        if new_val is None:
            continue

        if _values_different(field, old_val, new_val):
            updates[field] = new_val
            remarks.append(f"Change {field} to {new_val}")

    if not updates:
        return False

    repo_task.update(data=updates, where={"task_id": task_id})

    if remarks:
        _insert_task_history(
            task_id=task_id,
            remark="; ".join(remarks),
            updated_by=updated_by,
        )

    return True


def _log_numeric_sanitization_if_needed(
    file_path: str,
    row_number: int,
    original_value: Any,
    sanitized_value: Optional[Decimal],
    column_name: str,
    field_name: str,
) -> None:
    if original_value is None or str(original_value).strip() == "":
        return

    if sanitized_value is not None:
        return

    _safe_log(
        file_path,
        row_number,
        f"Valor inválido ou fora do range permitido para {field_name} DECIMAL(30,6). Campo será gravado como NULL.",
        column_name,
        original_value,
    )


# ======================================================================================
# PROCESSAMENTO DE LINHA
# ======================================================================================

def _process_single_row(
    row_dict: Dict[str, Any],
    file_path: str,
    row_number: int,
    execution_log_path: Optional[Path] = None,
) -> RowProcessResult:
    try:
        data = _normalize_row(row_dict)

        data["amount"] = _sanitize_task_value(data.get("amount"))
        data["booking_amount"] = _sanitize_task_booking_amount(data.get("booking_amount"))

        _log_numeric_sanitization_if_needed(
            file_path=file_path,
            row_number=row_number,
            original_value=row_dict.get("Max Estimated Incentives Amount"),
            sanitized_value=data.get("amount"),
            column_name="Max Estimated Incentives Amount",
            field_name="task_value",
        )

        _log_numeric_sanitization_if_needed(
            file_path=file_path,
            row_number=row_number,
            original_value=row_dict.get("Booking Amount - Net to Cisco"),
            sanitized_value=data.get("booking_amount"),
            column_name="Booking Amount - Net to Cisco",
            field_name="task_booking_amount",
        )

        valid, invalid_col, invalid_val = _validate_minimum_required_fields(data)
        if not valid:
            message = f"Campo obrigatório ausente ou inválido: {invalid_col}"
            _safe_log(file_path, row_number, message, invalid_col, invalid_val)
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column=invalid_col,
                error_value=invalid_val,
            )

        company_id = _resolve_company_id(data.get("cr_party_name"))
        if not company_id:
            message = "Customer não encontrado a partir de 'CR Party Name'"
            _safe_log(file_path, row_number, message, "CR Party Name", data.get("cr_party_name"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="CR Party Name",
                error_value=data.get("cr_party_name"),
            )

        task_type_id = _resolve_task_type_id(data)
        if not task_type_id:
            message = "Task type não resolvido"
            _safe_log(file_path, row_number, message, None, None)
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column=None,
                error_value=None,
            )

        existing_task = _resolve_existing_task(data)
        source_status = data.get("source_status")
        opt_in_flag = data.get("opt_in_flag", 0)
        lifecycle_start_date = data.get("lifecycle_start_date")

        # ------------------------------------------------------------------
        # Regra: Lifecycle Opt-In Status = Opted Out
        # ------------------------------------------------------------------
        if opt_in_flag == OPT_IN_FLAG_OPTED_OUT:
            opt_out_date_str = (
                lifecycle_start_date.strftime("%Y-%b-%d") if lifecycle_start_date else None
            )
            opt_out_remark = (
                f"Opt-Out was performed for this task on {opt_out_date_str}"
                if opt_out_date_str
                else "Opt-Out was performed for this task"
            )

            if existing_task is None:
                # Valida 'Deal ID' para criação
                valid_creation, invalid_col_c, invalid_val_c = _validate_required_for_creation(data)
                if not valid_creation:
                    message = f"Campo obrigatório ausente ou inválido para criação de task: {invalid_col_c}"
                    _safe_log(file_path, row_number, message, invalid_col_c, invalid_val_c)
                    return RowProcessResult(
                        success=False,
                        error_message=message,
                        error_column=invalid_col_c,
                        error_value=invalid_val_c,
                    )

                payload = _build_task_payload(
                    data=data,
                    company_id=company_id,
                    task_type_id=task_type_id,
                    force_status=TASK_STATUS_CANCELLED,
                    force_status_justification="AN OPT-OUT WAS PERFORMED FOR THE TASK",
                )

                if not payload.get("task_ws"):
                    message = "WS obrigatório ausente para criação da task"
                    _safe_log(file_path, row_number, message, "Deal Id", row_dict.get("Deal Id"))
                    return RowProcessResult(
                        success=False,
                        error_message=message,
                        error_column="Deal Id",
                        error_value=row_dict.get("Deal Id"),
                    )

                task_id = _create_task_from_payload(payload)
                if not task_id:
                    raise ValueError("Falha ao criar task (opted out)")

                _insert_task_history(
                    task_id=task_id,
                    activity_id=0,
                    remark=opt_out_remark,
                )

                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} task created as opted-out task_id={task_id} ws={payload.get('task_ws')} deal_id={payload.get('task_deal_id')}",
                    )

                return RowProcessResult(
                    success=True,
                    created=True,
                    cancelled=True,
                )

            # Task existente — fecha activities e task
            task_id = _safe_int(existing_task.get("task_id"), default=0)
            if not task_id:
                raise ValueError("task_id inválido na task existente (opted out)")

            current_status = _safe_int(existing_task.get("task_status"), default=0)
            if current_status in CLOSED_TASK_STATUSES:
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} ignored because task already closed (opted out) task_id={task_id} current_status={current_status} ws={data.get('ws')}",
                    )
                return RowProcessResult(success=True, ignored=True)

            activities_cancelled = _close_task_as_opted_out(
                task_id=task_id,
                opt_out_remark=opt_out_remark,
            )

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task opted-out task_id={task_id} ws={data.get('ws')} activities_cancelled={activities_cancelled}",
                )

            return RowProcessResult(
                success=True,
                cancelled=True,
                updated=True,
                activities_cancelled=activities_cancelled,
            )

        if existing_task is None:
            # Valida 'Deal ID' apenas agora, quando não há task existente e será necessário criar uma.
            valid_creation, invalid_col_c, invalid_val_c = _validate_required_for_creation(data)
            if not valid_creation:
                message = f"Campo obrigatório ausente ou inválido para criação de task: {invalid_col_c}"
                _safe_log(file_path, row_number, message, invalid_col_c, invalid_val_c)
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column=invalid_col_c,
                    error_value=invalid_val_c,
                )

            if source_status == SOURCE_STATUS_NOT_ELIGIBLE:
                payload = _build_task_payload(
                    data=data,
                    company_id=company_id,
                    task_type_id=task_type_id,
                    force_status=TASK_STATUS_DECLINED,
                    force_status_justification="TASK NOT ELIGIBLE BY VENDOR",
                )

                if not payload.get("task_ws"):
                    message = "WS obrigatório ausente para criação da task"
                    _safe_log(file_path, row_number, message, "Deal Id", row_dict.get("Deal Id"))
                    return RowProcessResult(
                        success=False,
                        error_message=message,
                        error_column="Deal Id",
                        error_value=row_dict.get("Deal Id"),
                    )

                task_id = _create_task_from_payload(payload)
                if not task_id:
                    raise ValueError("Falha ao criar task")

                _insert_task_history(
                    task_id=task_id,
                    remark="Task created as not eligible by Cisco",
                )

                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} task created as not eligible task_id={task_id} ws={payload.get('task_ws')} deal_id={payload.get('task_deal_id')} tasktype_id={payload.get('task_tasktype_id')} source_status={source_status}",
                    )

                return RowProcessResult(
                    success=True,
                    created=True,
                    ignored=True,
                )

            payload = _build_task_payload(data=data, company_id=company_id, task_type_id=task_type_id)

            if not payload.get("task_ws"):
                message = "WS obrigatório ausente para criação da task"
                _safe_log(file_path, row_number, message, "Deal Id", row_dict.get("Deal Id"))
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Deal Id",
                    error_value=row_dict.get("Deal Id"),
                )

            task_id = _create_task_from_payload(payload)
            if not task_id:
                raise ValueError("Falha ao criar task")

            if source_status == SOURCE_STATUS_EXPIRED:
                _insert_task_history(
                    task_id=task_id,
                    remark=f"Task created at {datetime.now().strftime('%Y-%b-%d')} with status Expired",
                )
            else:
                _insert_task_history(
                    task_id=task_id,
                    remark=f"Task created at {datetime.now().strftime('%Y-%b-%d')}",
                )

            if source_status == SOURCE_STATUS_ELIGIBLE and data.get("opt_in_flag") == 1 and payload.get("task_start"):
                _insert_task_history(
                    task_id=task_id,
                    remark=f"Task started at {payload['task_start'].strftime('%Y-%b-%d')}",
                    next_followup=payload["task_start"] + timedelta(days=5),
                )

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task created task_id={task_id} ws={payload.get('task_ws')} deal_id={payload.get('task_deal_id')} tasktype_id={payload.get('task_tasktype_id')} source_status={source_status}",
                )

            return RowProcessResult(
                success=True,
                created=True,
            )

        task_id = _safe_int(existing_task.get("task_id"), default=0)
        if not task_id:
            raise ValueError("task_id inválido na task existente")

        current_status = _safe_int(existing_task.get("task_status"), default=0)
        if current_status in CLOSED_TASK_STATUSES:
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} ignored because task already closed task_id={task_id} current_status={current_status} ws={data.get('ws')}",
                )
            return RowProcessResult(
                success=True,
                ignored=True,
            )

        if source_status == SOURCE_STATUS_NOT_ELIGIBLE:
            activities_cancelled = _close_task_as_not_eligible(task_id=task_id)

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task declined task_id={task_id} ws={data.get('ws')} activities_cancelled={activities_cancelled}",
                )

            return RowProcessResult(
                success=True,
                cancelled=True,
                updated=True,
                activities_cancelled=activities_cancelled,
            )

        if source_status == SOURCE_STATUS_EXPIRED:
            activities_cancelled = _close_task_as_expired(task_id=task_id)

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task expired task_id={task_id} ws={data.get('ws')} activities_closed={activities_cancelled}",
                )

            return RowProcessResult(
                success=True,
                cancelled=True,
                updated=True,
                activities_cancelled=activities_cancelled,
            )

        payload = _build_task_payload(data=data, company_id=company_id, task_type_id=task_type_id)

        updated = _update_existing_task(
            task_id=task_id,
            existing_task=existing_task,
            payload=payload,
            source_status=source_status,
            opt_in_flag=data.get("opt_in_flag", 0),
        )

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO row={row_number} task existing processed task_id={task_id} ws={payload.get('task_ws')} tasktype_id_existing={existing_task.get('task_tasktype_id')} updated={updated}",
            )

        return RowProcessResult(
            success=True,
            updated=updated,
            ignored=not updated,
        )

    except Exception as ex:
        message = f"Unexpected error: {ex}"
        _safe_log(file_path, row_number, message, None, None)

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"ERROR row={row_number} unexpected error={str(ex)[:1000]} traceback={traceback.format_exc()[:2000]}",
            )

        return RowProcessResult(
            success=False,
            error_message=message,
            error_column=None,
            error_value=None,
        )


# ======================================================================================
# PROCESSAMENTO DE CHUNK
# ======================================================================================

def _process_chunk(
    ws: Worksheet,
    idx_map: Dict[str, int],
    headers: List[str],
    start_row: int,
    end_row: int,
    file_path: str,
    failed_output_path: Path,
    execution_log_path: Optional[Path] = None,
    chunk_number: int = 0,
) -> Dict[str, int]:
    metrics = {
        "success": 0,
        "error": 0,
        "ignored": 0,
        "created": 0,
        "updated": 0,
        "cancelled": 0,
        "activities_cancelled": 0,
    }

    failed_rows_buffer: List[List[Any]] = []

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_start chunk={chunk_number} start_row={start_row} end_row={end_row}",
        )

    for r in range(end_row, start_row - 1, -1):
        row_cells = ws[r]
        raw_values = [cell.value for cell in row_cells]
        row_dict: Dict[str, Any] = {}

        for h, idx in idx_map.items():
            row_dict[h] = raw_values[idx] if idx < len(raw_values) else None

        result = _process_single_row(
            row_dict=row_dict,
            file_path=file_path,
            row_number=r,
            execution_log_path=execution_log_path,
        )

        if result.success:
            metrics["success"] += 1
            if result.ignored:
                metrics["ignored"] += 1
            if result.created:
                metrics["created"] += 1
            if result.updated:
                metrics["updated"] += 1
            if result.cancelled:
                metrics["cancelled"] += 1
            metrics["activities_cancelled"] += result.activities_cancelled
        else:
            metrics["error"] += 1
            failed_rows_buffer.append(
                _row_dict_to_failed_excel_row(
                    headers=headers,
                    row_dict=row_dict,
                    original_row_number=r,
                    input_file_path=file_path,
                    result=result,
                )
            )

        ws.delete_rows(r, 1)

    if failed_rows_buffer:
        failed_rows_buffer.reverse()
        _append_failed_rows(
            failed_path=failed_output_path,
            rows_to_append=failed_rows_buffer,
            original_headers=headers,
            execution_log_path=execution_log_path,
        )

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_finish chunk={chunk_number} success={metrics['success']} error={metrics['error']} ignored={metrics['ignored']} created={metrics['created']} updated={metrics['updated']} cancelled={metrics['cancelled']} activities_cancelled={metrics['activities_cancelled']}",
        )

    return metrics


# ======================================================================================
# FUNÇÃO PÚBLICA PRINCIPAL
# ======================================================================================

def run_import(
    file_path: str,
    user_id: Optional[str] = None,
    import_control_id: Optional[int] = None,
) -> Dict[str, Any]:
    _ensure_directories()

    input_path = _resolve_input_path(file_path)
    failed_output_path = _build_failed_output_path(input_path)
    execution_log_path = _build_execution_log_path(input_path)

    # Se não foi passado explicitamente, tenta descobrir pelo nome do arquivo
    if import_control_id is None and repo_import_control is not None:
        try:
            import_control_id = repo_import_control.get_id_by_file(input_path.name)
        except Exception:
            import_control_id = None

    started_at = datetime.now()
    total_rows = 0
    chunk_number = 0

    summary_metrics = {
        "processed_success": 0,
        "failed_rows": 0,
        "ignored_rows": 0,
        "tasks_created": 0,
        "tasks_updated": 0,
        "tasks_cancelled": 0,
        "activities_cancelled": 0,
    }

    _append_execution_log(execution_log_path, "START import")
    _append_execution_log(execution_log_path, f"INFO import_source={IMPORT_SOURCE}")
    _append_execution_log(execution_log_path, f"INFO user_id={user_id}")
    _append_execution_log(execution_log_path, f"INFO input_file={str(input_path)}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")

    try:
        wb, ws, idx_map, headers = _open_workbook_rw(str(input_path))
    except Exception as e:
        msg = f"Erro ao abrir arquivo Excel: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR failed opening workbook error={str(e)[:1000]}")
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": 0,
                "processed_success": 0,
                "failed_rows": 0,
                "execution_log_path": str(execution_log_path),
            },
        }

    total_rows = max(ws.max_row - 1, 0)
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO chunk_size={CHUNK_SIZE}")

    missing_cols = _validate_required_columns(headers)
    if missing_cols:
        msg = "Missing columns: " + ", ".join(missing_cols)
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR missing_columns={', '.join(missing_cols)}")
        wb.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": 0,
                "failed_rows": total_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    # Removido _ensure_failed_workbook aqui para que o arquivo de falhas
    # só seja criado quando realmente houver linhas com erro (dentro de _append_failed_rows).

    try:
        while ws.max_row > 1:
            chunk_number += 1
            chunk_end = min(1 + CHUNK_SIZE, ws.max_row)

            chunk_metrics = _process_chunk(
                ws=ws,
                idx_map=idx_map,
                headers=headers,
                start_row=2,
                end_row=chunk_end,
                file_path=str(input_path),
                failed_output_path=failed_output_path,
                execution_log_path=execution_log_path,
                chunk_number=chunk_number,
            )

            summary_metrics["processed_success"] += chunk_metrics["success"]
            summary_metrics["failed_rows"] += chunk_metrics["error"]
            summary_metrics["ignored_rows"] += chunk_metrics["ignored"]
            summary_metrics["tasks_created"] += chunk_metrics["created"]
            summary_metrics["tasks_updated"] += chunk_metrics["updated"]
            summary_metrics["tasks_cancelled"] += chunk_metrics["cancelled"]
            summary_metrics["activities_cancelled"] += chunk_metrics["activities_cancelled"]

            wb.save(str(input_path))

            remaining_rows_current = max(ws.max_row - 1, 0)

            _append_execution_log(
                execution_log_path,
                f"INFO checkpoint_after_chunk chunk={chunk_number} success_total={summary_metrics['processed_success']} error_total={summary_metrics['failed_rows']} ignored_total={summary_metrics['ignored_rows']} remaining_rows_current={remaining_rows_current}",
            )

            # Reporta progresso na tbImportControl a cada IMPORT_CONTROL_PROGRESS_INTERVAL_CHUNKS chunks
            if chunk_number % IMPORT_CONTROL_PROGRESS_INTERVAL_CHUNKS == 0:
                _report_progress(
                    import_control_id=import_control_id,
                    total_rows=total_rows,
                    processed_success=summary_metrics["processed_success"],
                    failed_rows=summary_metrics["failed_rows"],
                    ignored_rows=summary_metrics["ignored_rows"],
                    remaining_rows=remaining_rows_current,
                    execution_log_path=execution_log_path,
                )

    except Exception as e:
        msg = f"Erro durante o processamento da importação: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR processing failed error={str(e)[:1000]}")

        try:
            wb.save(str(input_path))
            _append_execution_log(execution_log_path, "INFO workbook saved after processing exception")
        except Exception as save_ex:
            _append_execution_log(
                execution_log_path,
                f"ERROR failed saving workbook after exception error={str(save_ex)[:1000]}",
            )

        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "tasks_created": summary_metrics["tasks_created"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "tasks_cancelled": summary_metrics["tasks_cancelled"],
                "activities_cancelled": summary_metrics["activities_cancelled"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        wb.save(str(input_path))
        _append_execution_log(execution_log_path, "INFO workbook saved successfully at end")
    except Exception as e:
        msg = f"Erro ao salvar arquivo Excel após processamento: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR final workbook save failed error={str(e)[:1000]}")
        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "tasks_created": summary_metrics["tasks_created"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "tasks_cancelled": summary_metrics["tasks_cancelled"],
                "activities_cancelled": summary_metrics["activities_cancelled"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    wb.close()

    remaining_rows_in_input = 0
    try:
        wb_check = load_workbook(str(input_path), read_only=True, data_only=True)
        ws_check = wb_check.active
        remaining_rows_in_input = max(ws_check.max_row - 1, 0)
        wb_check.close()
    except Exception as e:
        remaining_rows_in_input = -1
        _append_execution_log(
            execution_log_path,
            f"WARN failed checking remaining rows error={str(e)[:1000]}",
        )

    finished_at = datetime.now()
    duration_seconds = int((finished_at - started_at).total_seconds())

    _append_execution_log(execution_log_path, f"INFO total_chunks={chunk_number}")
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO processed_success={summary_metrics['processed_success']}")
    _append_execution_log(execution_log_path, f"INFO failed_rows={summary_metrics['failed_rows']}")
    _append_execution_log(execution_log_path, f"INFO ignored_rows={summary_metrics['ignored_rows']}")
    _append_execution_log(execution_log_path, f"INFO tasks_created={summary_metrics['tasks_created']}")
    _append_execution_log(execution_log_path, f"INFO tasks_updated={summary_metrics['tasks_updated']}")
    _append_execution_log(execution_log_path, f"INFO tasks_cancelled={summary_metrics['tasks_cancelled']}")
    _append_execution_log(execution_log_path, f"INFO activities_cancelled={summary_metrics['activities_cancelled']}")
    _append_execution_log(execution_log_path, f"INFO remaining_rows_in_input={remaining_rows_in_input}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")
    _append_execution_log(execution_log_path, f"INFO duration_seconds={duration_seconds}")

    if summary_metrics["failed_rows"] > 0:
        status = "FAILED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com falhas. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"criadas={summary_metrics['tasks_created']}, atualizadas={summary_metrics['tasks_updated']}, "
            f"canceladas={summary_metrics['tasks_cancelled']}, activities_cancelled={summary_metrics['activities_cancelled']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )
    else:
        status = "FINISHED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com sucesso. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"criadas={summary_metrics['tasks_created']}, atualizadas={summary_metrics['tasks_updated']}, "
            f"canceladas={summary_metrics['tasks_cancelled']}, activities_cancelled={summary_metrics['activities_cancelled']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )

    _append_execution_log(
        execution_log_path,
        f"FINISH status={status} success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} ignored={summary_metrics['ignored_rows']} remaining_rows_in_input={remaining_rows_in_input} duration_seconds={duration_seconds}",
    )

    return {
        "status": status,
        "message": msg,
        "summary": {
            "file_path": str(input_path),
            "total_rows": total_rows,
            "processed_success": summary_metrics["processed_success"],
            "failed_rows": summary_metrics["failed_rows"],
            "ignored_rows": summary_metrics["ignored_rows"],
            "tasks_created": summary_metrics["tasks_created"],
            "tasks_updated": summary_metrics["tasks_updated"],
            "tasks_cancelled": summary_metrics["tasks_cancelled"],
            "activities_cancelled": summary_metrics["activities_cancelled"],
            "failed_file_path": str(failed_output_path),
            "remaining_rows_in_input": remaining_rows_in_input,
            "execution_log_path": str(execution_log_path),
            "duration_seconds": duration_seconds,
            "total_chunks": chunk_number,
        },
    }
