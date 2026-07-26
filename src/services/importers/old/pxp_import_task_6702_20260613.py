"""
pxp_import_task_6702.py

Importador PXP do processo 6702.

Objetivo
--------
Processar arquivo XLSX de entrada contendo linhas PXP e refletir as regras
de negócio no banco de dados, com foco em criação, atualização e cancelamento
de tasks e activities, além de geração de histórico e arquivos de falha.

Padrão adotado
--------------
Este importador segue a mesma linha operacional observada em
`cisco_subscription_ccw.py`:

- lê o arquivo XLSX diretamente;
- processa em chunks;
- remove do arquivo original toda linha lida;
- escreve linhas com erro em um arquivo fixo de falhas;
- gera log texto de execução;
- grava logs funcionais em tbImportLog quando disponível;
- registra histórico em tbTaskRecord.

Observações importantes
-----------------------
1. Este arquivo foi entregue na "opção 2" solicitada:
   - usa adapter defensivo para a parte de UseCase / Solution Domain;
   - evita depender rigidamente de uma assinatura única externa.

2. Como nem todos os repositories auxiliares foram enviados, algumas partes
   foram implementadas com fallback seguro:
   - resolução de task type por nome usa tbTaskType via TaskRepository;
   - solution domain / architecture via UseCaseRepository são opcionais;
   - owner segue default 93 quando não houver regra melhor.

3. Regras consolidadas incorporadas:
   - origem/status elegível -> pode criar ou atualizar task
   - origem/status not eligible -> se task existir e não estiver fechada:
       * task_status = 5
       * task_eligible = 'N'
       * cancela activities abertas com activity_status = 4
   - se não existir task e linha vier not eligible:
       * linha é ignorada com sucesso operacional, sem criar task

4. O mapeamento de stage foi assumido conforme o desenho anterior:
   - eligible     -> 1
   - expired      -> 6
   - noteligible  -> 0

5. Quando status de origem = 0, isso NÃO significa task_status = 0.
   A regra correta aplicada é:
   - task_status = 5
   - task_eligible = 'N'

Dependências esperadas
----------------------
- openpyxl
- pandas
- repositories do projeto:
    * TaskRepository
    * TaskActivityRepository
    * TaskHistoryRepository
    * CompanyListNameRepository
    * ImportLogRepository (opcional)
    * UseCaseRepository (opcional)
"""

from __future__ import annotations

import logging
import re
import traceback
from dataclasses import dataclass
from datetime import datetime, timedelta, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.database.repositories.task_repository import TaskRepository
from src.infrastructure.database.repositories.task_activity_repository import TaskActivityRepository
from src.infrastructure.database.repositories.task_history_repository import TaskHistoryRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository

try:
    from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
except Exception:
    ImportLogRepository = None

try:
    from src.infrastructure.database.repositories.use_case_repository import UseCaseRepository
except Exception:
    UseCaseRepository = None


# ======================================================================================
# CONFIGURAÇÃO BASE
# ======================================================================================

BASE_STORAGE_PATH = Path("/home/bridgeadoption/storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"

IMPORT_SOURCE = "PxpImportTask6702"

ERROR_EXTRA_COLUMNS = [
    "import_error_message",
    "import_error_column",
    "import_error_value",
    "import_original_row",
    "import_processed_at",
    "import_source",
    "import_original_file",
]

logger = logging.getLogger(__name__)

repo_task = TaskRepository()
repo_activity = TaskActivityRepository()
repo_history = TaskHistoryRepository()
repo_company = CompanyListNameRepository()
repo_log = ImportLogRepository() if ImportLogRepository else None
repo_use_case = UseCaseRepository() if UseCaseRepository else None


# ======================================================================================
# CONSTANTES DE DOMÍNIO
# ======================================================================================

DEFAULT_TASK_OWNER_ID = 93
CHUNK_SIZE = 1000

SOURCE_STATUS_NOT_ELIGIBLE = 0
SOURCE_STATUS_ELIGIBLE = 1

TASK_STATUS_OPEN = 1
TASK_STATUS_IN_PROGRESS = 2
TASK_STATUS_ON_HOLD = 3
TASK_STATUS_CANCELLED = 4
TASK_STATUS_DECLINED = 5
TASK_STATUS_EXPIRED = 6
TASK_STATUS_SUBMITTED = 7
TASK_STATUS_RESUBMITTED = 8
TASK_STATUS_APPROVED_TO_CLOSE = 9
TASK_STATUS_COMPLETED = 10

ACTIVITY_STATUS_CANCELLED = 4

CLOSED_TASK_STATUSES = {
    TASK_STATUS_CANCELLED,
    TASK_STATUS_DECLINED,
    TASK_STATUS_EXPIRED,
    TASK_STATUS_COMPLETED,
}

CLOSED_ACTIVITY_STATUSES = {4, 5, 6, 10}

REQUIRED_COLUMNS = [
    "Plan Name",
    "Deal Name",
    "CR Party Name",
    "Stage",
    "Track",
    "Sub-Track",
]

OPTIONAL_COLUMNS = [
    "WS",
    "Deal Id",
    "Deal ID",
    "Web Order ID",
    "Sales Order ID",
    "CR Party ID",
    "Eligible Incentive Activity Description and Estimated Incentive Value",
    "Deal Incentive Expiry Date",
    "Max Estimated Incentives Amount",
    "Max Estimated Incentives Amount (Currency)",
    "EA Flag",
    "Telemetry Flag",
    "Booking Date",
    "Lifecycle Opt-In Status",
    "Lifecycle Start Date",
    "Booking Amount - Net to Cisco",
    "Fund Type",
]

EXPECTED_COLUMNS = REQUIRED_COLUMNS + OPTIONAL_COLUMNS


# ======================================================================================
# DATACLASSES
# ======================================================================================

@dataclass
class RowProcessResult:
    success: bool
    error_message: Optional[str] = None
    error_column: Optional[str] = None
    error_value: Optional[Any] = None
    ignored: bool = False
    created: bool = False
    updated: bool = False
    cancelled: bool = False
    activities_cancelled: int = 0


# ======================================================================================
# HELPERS DE DIRETÓRIO / LOG
# ======================================================================================

def _ensure_directories() -> None:
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)


def _build_execution_log_path(input_path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return BASE_LOGS_PATH / f"{input_path.stem}__{IMPORT_SOURCE}__{timestamp}.log"


def _append_execution_log(log_path: Path, message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def _safe_log(
    file_path: str,
    row_number: int,
    message: str,
    column_name: Optional[str] = None,
    value: Optional[Any] = None,
) -> None:
    if repo_log is None:
        return

    try:
        repo_log.create(
            IMPORT_SOURCE,
            file_path,
            row_number,
            message,
            column_name,
            value,
        )
    except Exception as log_ex:
        logger.exception(
            "Falha ao gravar tbImportLog. file=%s row=%s column=%s value=%s error=%s",
            file_path,
            row_number,
            column_name,
            value,
            log_ex,
        )


# ======================================================================================
# HELPERS DE NORMALIZAÇÃO
# ======================================================================================

def _normalize_str(value: Any, max_length: int = 1000) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    return text[:max_length]


def _normalize_stage(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=100)
    if text is None:
        return None

    normalized = re.sub(r"\s+", "", text).lower()
    return normalized


def _to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", ".")
    text = re.sub(r"[^0-9\.\-+]", "", text)

    if text in {"", ".", "-", "+", "-.", "+."}:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _to_float(value: Any) -> Optional[float]:
    dec = _to_decimal(value)
    return float(dec) if dec is not None else None


def _to_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()[:50]

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _parse_yes_no_flag(value: Any) -> int:
    text = _normalize_str(value, max_length=50)
    if not text:
        return 0

    normalized = re.sub(r"\s+", "", text).lower()

    if normalized in {"yes", "y", "true", "1"}:
        return -1

    if normalized in {"no", "n", "false", "0"}:
        return 0

    return 0


def _parse_opt_in_flag(value: Any) -> int:
    text = _normalize_str(value, max_length=100)
    if not text:
        return 0

    normalized = re.sub(r"\s+", "", text).lower()

    if normalized in {"optedin", "optedin.", "opted-in", "opted_in"}:
        return -1

    if normalized in {"pre-approved", "preapproved", "pending", "notoptedin"}:
        return 0

    return 0


def _clean_currency(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=50)
    if not text:
        return None

    text = text.replace("($)", "").replace("$", "").strip()
    return text or None


def _extract_deal_id_from_deal_name(deal_name: Optional[str]) -> Optional[str]:
    if not deal_name:
        return None

    match = re.search(r"Deal\s+(\d+)", deal_name, flags=re.IGNORECASE)
    if match:
        return match.group(1)

    return None


def _resolve_stage_to_source_status(stage_value: Any) -> int:
    normalized = _normalize_stage(stage_value)

    mapping = {
        "eligible": SOURCE_STATUS_ELIGIBLE,
        "expired": TASK_STATUS_EXPIRED,
        "noteligible": SOURCE_STATUS_NOT_ELIGIBLE,
    }

    if normalized not in mapping:
        raise ValueError(f"Stage inválido: {stage_value}")

    return mapping[normalized]


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default

    if isinstance(value, bool):
        return default

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value) if value.is_integer() else default

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    if text.lstrip("+-").isdigit():
        try:
            return int(text)
        except Exception:
            return default

    return default


# ======================================================================================
# HELPERS DE ARQUIVO EXCEL
# ======================================================================================

def _resolve_input_path(file_path: str) -> Path:
    path = Path(file_path)
    if path.is_absolute():
        return path
    return BASE_INPUT_PATH / path.name


def _build_failed_output_path(input_path: Path) -> Path:
    return BASE_OUTPUT_PATH / f"{input_path.stem}__failed_rows.xlsx"


def _open_workbook_rw(file_path: str) -> Tuple[Any, Worksheet, Dict[str, int], List[str]]:
    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx_map = {h: i for i, h in enumerate(headers)}
    return wb, ws, idx_map, headers


def _ensure_failed_workbook(
    failed_path: Path,
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
) -> None:
    if failed_path.exists():
        if execution_log_path:
            _append_execution_log(execution_log_path, f"INFO failed file reused: {str(failed_path)}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "failed_rows"
    ws.append(original_headers + ERROR_EXTRA_COLUMNS)
    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(execution_log_path, f"INFO failed file created: {str(failed_path)}")


def _append_failed_rows(
    failed_path: Path,
    rows_to_append: List[List[Any]],
    execution_log_path: Optional[Path] = None,
) -> None:
    if not rows_to_append:
        return

    wb = load_workbook(str(failed_path))
    ws = wb.active

    for row in rows_to_append:
        ws.append(row)

    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO appended failed rows count={len(rows_to_append)} to file={str(failed_path)}",
        )


def _row_dict_to_failed_excel_row(
    headers: List[str],
    row_dict: Dict[str, Any],
    original_row_number: int,
    input_file_path: str,
    result: RowProcessResult,
) -> List[Any]:
    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    original_values = [row_dict.get(h) for h in headers]

    return original_values + [
        result.error_message,
        result.error_column,
        str(result.error_value)[:2000] if result.error_value is not None else None,
        original_row_number,
        processed_at,
        IMPORT_SOURCE,
        input_file_path,
    ]


# ======================================================================================
# RESOLUÇÃO DE NEGÓCIO
# ======================================================================================

def _validate_required_columns(headers: List[str]) -> List[str]:
    return [col for col in REQUIRED_COLUMNS if col not in headers]


def _normalize_row(row_dict: Dict[str, Any]) -> Dict[str, Any]:
    deal_name = _normalize_str(row_dict.get("Deal Name"))
    explicit_deal_id = _normalize_str(row_dict.get("Deal ID")) or _normalize_str(row_dict.get("Deal Id"))

    normalized = {
        "plan_name": _normalize_str(row_dict.get("Plan Name")),
        "ws": _normalize_str(row_dict.get("WS")),
        "deal_id": explicit_deal_id or _extract_deal_id_from_deal_name(deal_name),
        "deal_name": deal_name,
        "web_order_id": _normalize_str(row_dict.get("Web Order ID")),
        "sales_order_id": _normalize_str(row_dict.get("Sales Order ID")),
        "cr_party_id": _normalize_str(row_dict.get("CR Party ID")),
        "cr_party_name": _normalize_str(row_dict.get("CR Party Name")),
        "stage_raw": _normalize_str(row_dict.get("Stage")),
        "description": _normalize_str(
            row_dict.get("Eligible Incentive Activity Description and Estimated Incentive Value"),
            max_length=2000,
        ),
        "end_date": _to_date(row_dict.get("Deal Incentive Expiry Date")),
        "amount": _to_float(row_dict.get("Max Estimated Incentives Amount")),
        "currency": _clean_currency(row_dict.get("Max Estimated Incentives Amount (Currency)")),
        "track": _normalize_str(row_dict.get("Track")),
        "subtrack": _normalize_str(row_dict.get("Sub-Track")),
        "ea_flag": _parse_yes_no_flag(row_dict.get("EA Flag")),
        "telemetry_flag": _parse_yes_no_flag(row_dict.get("Telemetry Flag")),
        "booking_date": _to_date(row_dict.get("Booking Date")),
        "opt_in_flag": _parse_opt_in_flag(row_dict.get("Lifecycle Opt-In Status")),
        "lifecycle_start_date": _to_date(row_dict.get("Lifecycle Start Date")),
        "booking_amount": _to_float(row_dict.get("Booking Amount - Net to Cisco")),
        "fund_type": _normalize_str(row_dict.get("Fund Type")),
        "source_status": _resolve_stage_to_source_status(row_dict.get("Stage")),
    }

    return normalized


def _validate_minimum_required_fields(data: Dict[str, Any]) -> Tuple[bool, Optional[str], Optional[Any]]:
    required_map = {
        "Plan Name": data.get("plan_name"),
        "CR Party Name": data.get("cr_party_name"),
        "Stage": data.get("stage_raw"),
        "Track": data.get("track"),
        "Sub-Track": data.get("subtrack"),
    }

    for col, value in required_map.items():
        if value is None or str(value).strip() == "":
            return False, col, value

    return True, None, None


def _build_task_reference(data: Dict[str, Any]) -> Optional[str]:
    parts: List[str] = []

    if data.get("web_order_id"):
        parts.append(str(data["web_order_id"]).strip())

    if data.get("sales_order_id"):
        parts.append(str(data["sales_order_id"]).strip())

    if data.get("cr_party_id"):
        parts.append(f"CR Party ID: {str(data['cr_party_id']).strip()}")

    if not parts:
        return None

    return "; ".join(parts)


def _resolve_company_id(company_name: Optional[str]) -> int:
    if not company_name:
        return 0

    try:
        company_id = repo_company.get_company_id_by_name(company_name)
        return _safe_int(company_id, default=0)
    except Exception:
        return 0


def _resolve_task_type_id_from_task_types(name_value: Optional[str]) -> int:
    if not name_value:
        return 0

    try:
        task_types = repo_task.get_task_type_by_ids(as_df=False)
    except Exception:
        return 0

    normalized_target = str(name_value).strip().lower()

    for row in task_types:
        tasktype_name = str(row.get("tasktype_name") or "").strip().lower()
        if tasktype_name == normalized_target:
            return _safe_int(row.get("tasktype_id"), default=0)

    return 0


def _resolve_task_type_id(data: Dict[str, Any]) -> int:
    plan_name = data.get("plan_name")
    fund_type = data.get("fund_type")

    candidate_names: List[str] = []

    if plan_name and "|" in plan_name:
        last_part = plan_name.split("|")[-1].strip()
        if last_part:
            candidate_names.append(last_part)

    if fund_type:
        candidate_names.append(fund_type)

    for candidate in candidate_names:
        found_id = _resolve_task_type_id_from_task_types(candidate)
        if found_id:
            return found_id

    if fund_type and fund_type.strip().lower() == "lifecycle incentives":
        return 22

    return 0


def _resolve_solution_and_architecture(subtrack: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    """
    Adapter defensivo para integração com UseCaseRepository.

    Estratégia:
    - se o repository não existir, retorna (None, None)
    - se o método não existir, retorna (None, None)
    - se o retorno for dict, tenta mapear:
        * architecture
        * solution_domain
      com nomes alternativos
    - se o retorno for string, trata como solution_domain
    """
    if not subtrack or repo_use_case is None:
        return None, None

    try:
        if not hasattr(repo_use_case, "get_solution_by_use_case"):
            return None, None

        result = repo_use_case.get_solution_by_use_case(subtrack)

        if result is None:
            return None, None

        if isinstance(result, dict):
            architecture = (
                result.get("uc_architecture")
                or result.get("architecture")
                or result.get("task_architecture")
            )
            solution_domain = (
                result.get("uc_solution_domain")
                or result.get("solution_domain")
                or result.get("task_solution_domain")
                or result.get("solution")
            )
            return _normalize_str(solution_domain), _normalize_str(architecture)

        if isinstance(result, (list, tuple)) and result:
            first = result[0]
            if isinstance(first, dict):
                architecture = (
                    first.get("uc_architecture")
                    or first.get("architecture")
                    or first.get("task_architecture")
                )
                solution_domain = (
                    first.get("uc_solution_domain")
                    or first.get("solution_domain")
                    or first.get("task_solution_domain")
                    or first.get("solution")
                )
                return _normalize_str(solution_domain), _normalize_str(architecture)

            if isinstance(first, str):
                return _normalize_str(first), None

        if isinstance(result, str):
            return _normalize_str(result), None

    except Exception:
        return None, None

    return None, None


def _resolve_existing_task(data: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    """
    Regra operacional:
    1. tenta por WS se houver
    2. fallback por ws + track + subtrack, usando tbTask diretamente
    """
    ws = data.get("ws")
    track = data.get("track")
    subtrack = data.get("subtrack")

    try:
        if ws:
            found_ids = repo_task.find_ids_by({"task_ws": ws})
            if found_ids:
                task_id = _safe_int(found_ids[0], default=0)
                if task_id:
                    rows = repo_task.get_task(task_id=task_id, as_df=False)
                    return rows[0] if rows else {"task_id": task_id}

        if ws and track and subtrack:
            found_ids = repo_task.find_ids_by(
                {
                    "task_ws": ws,
                    "task_track": track,
                    "task_subtrack": subtrack,
                }
            )
            if found_ids:
                task_id = _safe_int(found_ids[0], default=0)
                if task_id:
                    rows = repo_task.get_task(task_id=task_id, as_df=False)
                    return rows[0] if rows else {"task_id": task_id}

    except Exception:
        return None

    return None


def _build_task_payload(data: Dict[str, Any], company_id: int, task_type_id: int) -> Dict[str, Any]:
    task_start = data.get("booking_date")
    if data.get("opt_in_flag") == -1 and data.get("lifecycle_start_date"):
        task_start = data.get("lifecycle_start_date")

    solution_domain, architecture = _resolve_solution_and_architecture(data.get("subtrack"))

    payload = {
        "task_tasktype_id": task_type_id,
        "task_reference": _build_task_reference(data),
        "task_owner_id": DEFAULT_TASK_OWNER_ID,
        "task_temp_owner_id": None,
        "task_customer_id": company_id,
        "task_cr_party_id": _safe_int(data.get("cr_party_id"), default=0),
        "task_cr_party_name": data.get("cr_party_name"),
        "task_created_in": datetime.now(),
        "task_created_by": 0,
        "task_priority": "LOW",
        "task_project_id": 0,
        "task_status": TASK_STATUS_OPEN if data.get("source_status") != TASK_STATUS_EXPIRED else TASK_STATUS_EXPIRED,
        "task_status_justification": None,
        "task_start": task_start,
        "task_end": data.get("end_date"),
        "task_start_performed": task_start,
        "task_end_performed": data.get("end_date"),
        "task_value": data.get("amount") if data.get("amount") is not None else 0,
        "task_forecast": 0,
        "task_backlog": 0,
        "task_rate": 1,
        "task_currency": data.get("currency") or "USD",
        "task_ws": data.get("ws"),
        "task_deal_id": data.get("deal_id"),
        "task_track": data.get("track"),
        "task_subtrack": data.get("subtrack"),
        "task_highlight": 0,
        "task_remark": None,
        "task_description": data.get("description"),
        "task_ea_flag": data.get("ea_flag", 0),
        "task_telemetry_flag": data.get("telemetry_flag", 0),
        "task_opt_in_flag": data.get("opt_in_flag", 0),
        "task_completed": 0,
        "task_architecture": architecture,
        "task_solution_domain": solution_domain,
        "task_eligible": "Y",
        "task_end_fy": None,
        "task_booking_date": data.get("booking_date"),
        "task_booking_amount": data.get("booking_amount"),
    }

    if data.get("opt_in_flag") == -1:
        payload["task_status"] = TASK_STATUS_IN_PROGRESS

    return payload


def _insert_task_history(
    task_id: int,
    remark: str,
    updated_by: str = "System BA",
    activity_id: int = 0,
    next_followup: Optional[date] = None,
) -> None:
    history = {
        "taskrecord_task_id": task_id,
        "taskrecord_activity_id": activity_id,
        "taskrecord_remark": remark,
        "taskrecord_updated_by": updated_by,
    }

    if next_followup is not None:
        history["taskrecord_next_followup"] = next_followup.strftime("%Y-%m-%d")

    repo_history.insert(history)


def _create_task_from_payload(payload: Dict[str, Any]) -> int:
    task_id = repo_task.insert(payload)
    return _safe_int(task_id, default=0)


def _cancel_open_activities(task_id: int, updated_by: str = "System BA") -> int:
    cancelled = 0

    try:
        activities = repo_activity.get_activity(task_id=task_id, as_df=False)
    except Exception:
        return 0

    for activity in activities:
        activity_id = _safe_int(activity.get("activity_id"), default=0)
        current_status = _safe_int(activity.get("activity_status"), default=0)

        if not activity_id:
            continue

        if current_status in CLOSED_ACTIVITY_STATUSES:
            continue

        try:
            repo_activity.update(
                data={"activity_status": ACTIVITY_STATUS_CANCELLED},
                where={"activity_id": activity_id},
            )
            cancelled += 1

            _insert_task_history(
                task_id=task_id,
                activity_id=activity_id,
                remark="Activity cancelled because task changed to not eligible by Cisco; Forecast value $0.00",
                updated_by=updated_by,
            )
        except Exception:
            continue

    return cancelled


def _cancel_task_as_not_eligible(task_id: int, updated_by: str = "System BA") -> int:
    try:
        repo_task.update(
            data={
                "task_status": TASK_STATUS_DECLINED,
                "task_eligible": "N",
                "task_forecast": 0,
            },
            where={"task_id": task_id},
        )

        _insert_task_history(
            task_id=task_id,
            activity_id=0,
            remark="Task not eligible by Cisco",
            updated_by=updated_by,
        )

        return _cancel_open_activities(task_id=task_id, updated_by=updated_by)
    except Exception:
        return 0


def _update_existing_task(
    task_id: int,
    existing_task: Dict[str, Any],
    payload: Dict[str, Any],
    updated_by: str = "System BA",
) -> bool:
    updates: Dict[str, Any] = {}
    remarks: List[str] = []

    current_status = _safe_int(existing_task.get("task_status"), default=0)
    current_opt_in = _safe_int(existing_task.get("task_opt_in_flag"), default=0)
    current_telemetry = _safe_int(existing_task.get("task_telemetry_flag"), default=0)
    current_ea = _safe_int(existing_task.get("task_ea_flag"), default=0)
    current_owner = _safe_int(existing_task.get("task_owner_id"), default=0)

    new_status = _safe_int(payload.get("task_status"), default=current_status)

    if new_status > current_status and current_status not in {4, 6, 7, 10}:
        updates["task_status"] = new_status
        remarks.append(f"Change Task Status to {new_status}")

    if payload.get("task_opt_in_flag", current_opt_in) != current_opt_in:
        updates["task_opt_in_flag"] = payload.get("task_opt_in_flag", current_opt_in)
        remarks.append(f"Change Opt-In Flag to {payload.get('task_opt_in_flag', current_opt_in)}")

        if payload.get("task_start"):
            updates["task_start"] = payload.get("task_start")
            updates["task_start_performed"] = payload.get("task_start")
            remarks.append(f"Change Task Start to {payload.get('task_start')}")

    if payload.get("task_telemetry_flag", current_telemetry) != current_telemetry:
        updates["task_telemetry_flag"] = payload.get("task_telemetry_flag", current_telemetry)
        remarks.append(f"Change Telemetry Flag to {payload.get('task_telemetry_flag', current_telemetry)}")

    if payload.get("task_ea_flag", current_ea) != current_ea:
        updates["task_ea_flag"] = payload.get("task_ea_flag", current_ea)
        remarks.append(f"Change EA Flag to {payload.get('task_ea_flag', current_ea)}")

    if current_owner == 0 and _safe_int(payload.get("task_owner_id"), default=0) != 0:
        updates["task_owner_id"] = payload.get("task_owner_id")
        remarks.append(f"Change Task Owner to {payload.get('task_owner_id')}")

    simple_compare_fields = [
        "task_reference",
        "task_customer_id",
        "task_cr_party_id",
        "task_cr_party_name",
        "task_end",
        "task_value",
        "task_currency",
        "task_track",
        "task_subtrack",
        "task_description",
        "task_booking_date",
        "task_booking_amount",
        "task_architecture",
        "task_solution_domain",
        "task_deal_id",
        "task_ws",
    ]

    for field in simple_compare_fields:
        old_val = existing_task.get(field)
        new_val = payload.get(field)

        if old_val != new_val and new_val is not None:
            updates[field] = new_val

    if not updates:
        return False

    repo_task.update(data=updates, where={"task_id": task_id})

    if remarks:
        _insert_task_history(
            task_id=task_id,
            remark="; ".join(remarks),
            updated_by=updated_by,
        )

    return True


# ======================================================================================
# PROCESSAMENTO DE LINHA
# ======================================================================================

def _process_single_row(
    row_dict: Dict[str, Any],
    file_path: str,
    row_number: int,
    execution_log_path: Optional[Path] = None,
) -> RowProcessResult:
    try:
        data = _normalize_row(row_dict)

        valid, invalid_col, invalid_val = _validate_minimum_required_fields(data)
        if not valid:
            message = f"Campo obrigatório ausente: {invalid_col}"
            _safe_log(file_path, row_number, message, invalid_col, invalid_val)
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column=invalid_col,
                error_value=invalid_val,
            )

        company_id = _resolve_company_id(data.get("cr_party_name"))
        if not company_id:
            message = "Customer não encontrado a partir de 'CR Party Name'"
            _safe_log(file_path, row_number, message, "CR Party Name", data.get("cr_party_name"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="CR Party Name",
                error_value=data.get("cr_party_name"),
            )

        task_type_id = _resolve_task_type_id(data)
        if not task_type_id:
            message = "Task type não resolvido a partir de 'Plan Name' / 'Fund Type'"
            _safe_log(file_path, row_number, message, "Plan Name", data.get("plan_name"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Plan Name",
                error_value=data.get("plan_name"),
            )

        existing_task = _resolve_existing_task(data)

        if data.get("source_status") == SOURCE_STATUS_NOT_ELIGIBLE:
            if existing_task is None:
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} ignored not eligible without existing task",
                    )
                return RowProcessResult(
                    success=True,
                    ignored=True,
                )

            task_id = _safe_int(existing_task.get("task_id"), default=0)
            current_status = _safe_int(existing_task.get("task_status"), default=0)

            if not task_id:
                message = "Task existente inválida para cancelamento"
                _safe_log(file_path, row_number, message, "WS", data.get("ws"))
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="WS",
                    error_value=data.get("ws"),
                )

            if current_status in CLOSED_TASK_STATUSES:
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} ignored not eligible because task already closed task_id={task_id}",
                    )
                return RowProcessResult(
                    success=True,
                    ignored=True,
                )

            activities_cancelled = _cancel_task_as_not_eligible(task_id=task_id)

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task declined task_id={task_id} activities_cancelled={activities_cancelled}",
                )

            return RowProcessResult(
                success=True,
                cancelled=True,
                updated=True,
                activities_cancelled=activities_cancelled,
            )

        payload = _build_task_payload(data=data, company_id=company_id, task_type_id=task_type_id)

        if existing_task is None:
            task_id = _create_task_from_payload(payload)
            if not task_id:
                raise ValueError("Falha ao criar task")

            _insert_task_history(
                task_id=task_id,
                remark=f"Task created at {datetime.now().strftime('%Y-%b-%d')}",
            )

            if data.get("opt_in_flag") == -1 and payload.get("task_start"):
                _insert_task_history(
                    task_id=task_id,
                    remark=f"Task started at {payload['task_start'].strftime('%Y-%b-%d')}",
                    next_followup=payload["task_start"] + timedelta(days=5),
                )

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task created task_id={task_id}",
                )

            return RowProcessResult(
                success=True,
                created=True,
            )

        task_id = _safe_int(existing_task.get("task_id"), default=0)
        if not task_id:
            raise ValueError("task_id inválido na task existente")

        updated = _update_existing_task(
            task_id=task_id,
            existing_task=existing_task,
            payload=payload,
        )

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO row={row_number} task existing processed task_id={task_id} updated={updated}",
            )

        return RowProcessResult(
            success=True,
            updated=updated,
            ignored=not updated,
        )

    except Exception as ex:
        message = f"Unexpected error: {ex}"
        _safe_log(file_path, row_number, message, None, None)

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"ERROR row={row_number} unexpected error={str(ex)[:1000]} traceback={traceback.format_exc()[:2000]}",
            )

        return RowProcessResult(
            success=False,
            error_message=message,
            error_column=None,
            error_value=None,
        )


# ======================================================================================
# PROCESSAMENTO DE CHUNK
# ======================================================================================

def _process_chunk(
    ws: Worksheet,
    idx_map: Dict[str, int],
    headers: List[str],
    start_row: int,
    end_row: int,
    file_path: str,
    failed_output_path: Path,
    execution_log_path: Optional[Path] = None,
    chunk_number: int = 0,
) -> Dict[str, int]:
    metrics = {
        "success": 0,
        "error": 0,
        "ignored": 0,
        "created": 0,
        "updated": 0,
        "cancelled": 0,
        "activities_cancelled": 0,
    }

    failed_rows_buffer: List[List[Any]] = []

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_start chunk={chunk_number} start_row={start_row} end_row={end_row}",
        )

    for r in range(end_row, start_row - 1, -1):
        row_cells = ws[r]
        raw_values = [cell.value for cell in row_cells]
        row_dict: Dict[str, Any] = {}

        for h, idx in idx_map.items():
            row_dict[h] = raw_values[idx] if idx < len(raw_values) else None

        result = _process_single_row(
            row_dict=row_dict,
            file_path=file_path,
            row_number=r,
            execution_log_path=execution_log_path,
        )

        if result.success:
            metrics["success"] += 1
            if result.ignored:
                metrics["ignored"] += 1
            if result.created:
                metrics["created"] += 1
            if result.updated:
                metrics["updated"] += 1
            if result.cancelled:
                metrics["cancelled"] += 1
            metrics["activities_cancelled"] += result.activities_cancelled
        else:
            metrics["error"] += 1
            failed_rows_buffer.append(
                _row_dict_to_failed_excel_row(
                    headers=headers,
                    row_dict=row_dict,
                    original_row_number=r,
                    input_file_path=file_path,
                    result=result,
                )
            )

        ws.delete_rows(r, 1)

    if failed_rows_buffer:
        failed_rows_buffer.reverse()
        _append_failed_rows(
            failed_path=failed_output_path,
            rows_to_append=failed_rows_buffer,
            execution_log_path=execution_log_path,
        )

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_finish chunk={chunk_number} success={metrics['success']} error={metrics['error']} ignored={metrics['ignored']} created={metrics['created']} updated={metrics['updated']} cancelled={metrics['cancelled']} activities_cancelled={metrics['activities_cancelled']}",
        )

    return metrics


# ======================================================================================
# FUNÇÃO PÚBLICA PRINCIPAL
# ======================================================================================

def run_import(file_path: str, user_id: Optional[str] = None) -> Dict[str, Any]:
    """
    Executa a importação do processo 6702.

    Fluxo:
    1. abre arquivo Excel
    2. valida colunas obrigatórias
    3. processa em chunks
    4. remove toda linha processada do original
    5. grava linhas com erro no arquivo de falhas
    6. retorna resumo consolidado
    """
    _ensure_directories()

    input_path = _resolve_input_path(file_path)
    failed_output_path = _build_failed_output_path(input_path)
    execution_log_path = _build_execution_log_path(input_path)

    started_at = datetime.now()
    total_rows = 0
    chunk_number = 0

    summary_metrics = {
        "processed_success": 0,
        "failed_rows": 0,
        "ignored_rows": 0,
        "tasks_created": 0,
        "tasks_updated": 0,
        "tasks_cancelled": 0,
        "activities_cancelled": 0,
    }

    _append_execution_log(execution_log_path, "START import")
    _append_execution_log(execution_log_path, f"INFO import_source={IMPORT_SOURCE}")
    _append_execution_log(execution_log_path, f"INFO user_id={user_id}")
    _append_execution_log(execution_log_path, f"INFO input_file={str(input_path)}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")

    try:
        wb, ws, idx_map, headers = _open_workbook_rw(str(input_path))
    except Exception as e:
        msg = f"Erro ao abrir arquivo Excel: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR failed opening workbook error={str(e)[:1000]}")
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": 0,
                "processed_success": 0,
                "failed_rows": 0,
                "execution_log_path": str(execution_log_path),
            },
        }

    total_rows = max(ws.max_row - 1, 0)
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO chunk_size={CHUNK_SIZE}")

    missing_cols = _validate_required_columns(headers)
    if missing_cols:
        msg = "Missing columns: " + ", ".join(missing_cols)
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR missing_columns={', '.join(missing_cols)}")
        wb.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": 0,
                "failed_rows": total_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        _ensure_failed_workbook(failed_output_path, headers, execution_log_path)
    except Exception as e:
        msg = f"Erro ao preparar arquivo XLSX de falhas: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR failed preparing error workbook error={str(e)[:1000]}")
        wb.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": 0,
                "failed_rows": 0,
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        while ws.max_row > 1:
            chunk_number += 1
            chunk_end = min(1 + CHUNK_SIZE, ws.max_row)

            chunk_metrics = _process_chunk(
                ws=ws,
                idx_map=idx_map,
                headers=headers,
                start_row=2,
                end_row=chunk_end,
                file_path=str(input_path),
                failed_output_path=failed_output_path,
                execution_log_path=execution_log_path,
                chunk_number=chunk_number,
            )

            summary_metrics["processed_success"] += chunk_metrics["success"]
            summary_metrics["failed_rows"] += chunk_metrics["error"]
            summary_metrics["ignored_rows"] += chunk_metrics["ignored"]
            summary_metrics["tasks_created"] += chunk_metrics["created"]
            summary_metrics["tasks_updated"] += chunk_metrics["updated"]
            summary_metrics["tasks_cancelled"] += chunk_metrics["cancelled"]
            summary_metrics["activities_cancelled"] += chunk_metrics["activities_cancelled"]

            wb.save(str(input_path))

            _append_execution_log(
                execution_log_path,
                f"INFO checkpoint_after_chunk chunk={chunk_number} success_total={summary_metrics['processed_success']} error_total={summary_metrics['failed_rows']} ignored_total={summary_metrics['ignored_rows']} remaining_rows_current={max(ws.max_row - 1, 0)}",
            )

    except Exception as e:
        msg = f"Erro durante o processamento da importação: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR processing failed error={str(e)[:1000]}")

        try:
            wb.save(str(input_path))
            _append_execution_log(execution_log_path, "INFO workbook saved after processing exception")
        except Exception as save_ex:
            _append_execution_log(
                execution_log_path,
                f"ERROR failed saving workbook after exception error={str(save_ex)[:1000]}",
            )

        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "tasks_created": summary_metrics["tasks_created"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "tasks_cancelled": summary_metrics["tasks_cancelled"],
                "activities_cancelled": summary_metrics["activities_cancelled"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        wb.save(str(input_path))
        _append_execution_log(execution_log_path, "INFO workbook saved successfully at end")
    except Exception as e:
        msg = f"Erro ao salvar arquivo Excel após processamento: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR final workbook save failed error={str(e)[:1000]}")
        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "tasks_created": summary_metrics["tasks_created"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "tasks_cancelled": summary_metrics["tasks_cancelled"],
                "activities_cancelled": summary_metrics["activities_cancelled"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    wb.close()

    remaining_rows_in_input = 0
    try:
        wb_check = load_workbook(str(input_path), read_only=True, data_only=True)
        ws_check = wb_check.active
        remaining_rows_in_input = max(ws_check.max_row - 1, 0)
        wb_check.close()
    except Exception as e:
        remaining_rows_in_input = -1
        _append_execution_log(
            execution_log_path,
            f"WARN failed checking remaining rows error={str(e)[:1000]}",
        )

    finished_at = datetime.now()
    duration_seconds = int((finished_at - started_at).total_seconds())

    _append_execution_log(execution_log_path, f"INFO total_chunks={chunk_number}")
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO processed_success={summary_metrics['processed_success']}")
    _append_execution_log(execution_log_path, f"INFO failed_rows={summary_metrics['failed_rows']}")
    _append_execution_log(execution_log_path, f"INFO ignored_rows={summary_metrics['ignored_rows']}")
    _append_execution_log(execution_log_path, f"INFO tasks_created={summary_metrics['tasks_created']}")
    _append_execution_log(execution_log_path, f"INFO tasks_updated={summary_metrics['tasks_updated']}")
    _append_execution_log(execution_log_path, f"INFO tasks_cancelled={summary_metrics['tasks_cancelled']}")
    _append_execution_log(execution_log_path, f"INFO activities_cancelled={summary_metrics['activities_cancelled']}")
    _append_execution_log(execution_log_path, f"INFO remaining_rows_in_input={remaining_rows_in_input}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")
    _append_execution_log(execution_log_path, f"INFO duration_seconds={duration_seconds}")

    if summary_metrics["failed_rows"] > 0:
        status = "FAILED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com falhas. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"criadas={summary_metrics['tasks_created']}, atualizadas={summary_metrics['tasks_updated']}, "
            f"canceladas={summary_metrics['tasks_cancelled']}, activities_cancelled={summary_metrics['activities_cancelled']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )
    else:
        status = "FINISHED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com sucesso. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"criadas={summary_metrics['tasks_created']}, atualizadas={summary_metrics['tasks_updated']}, "
            f"canceladas={summary_metrics['tasks_cancelled']}, activities_cancelled={summary_metrics['activities_cancelled']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )

    _append_execution_log(
        execution_log_path,
        f"FINISH status={status} success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} ignored={summary_metrics['ignored_rows']} remaining_rows_in_input={remaining_rows_in_input} duration_seconds={duration_seconds}",
    )

    return {
        "status": status,
        "message": msg,
        "summary": {
            "file_path": str(input_path),
            "total_rows": total_rows,
            "processed_success": summary_metrics["processed_success"],
            "failed_rows": summary_metrics["failed_rows"],
            "ignored_rows": summary_metrics["ignored_rows"],
            "tasks_created": summary_metrics["tasks_created"],
            "tasks_updated": summary_metrics["tasks_updated"],
            "tasks_cancelled": summary_metrics["tasks_cancelled"],
            "activities_cancelled": summary_metrics["activities_cancelled"],
            "failed_file_path": str(failed_output_path),
            "remaining_rows_in_input": remaining_rows_in_input,
            "execution_log_path": str(execution_log_path),
            "duration_seconds": duration_seconds,
            "total_chunks": chunk_number,
        },
    }
