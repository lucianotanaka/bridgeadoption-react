import logging
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple, Any
from datetime import datetime, timedelta

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.database.repositories.cisco_ea_repository import CiscoEARepository
from src.infrastructure.database.repositories.product_repository import ProductRepository
from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
from src.infrastructure.database.repositories.task_repository import TaskRepository
from src.infrastructure.database.repositories.task_activity_repository import TaskActivityRepository
from src.infrastructure.database.repositories.task_history_repository import TaskHistoryRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository
from src.infrastructure.database.repositories.cisco_web_order_repository import CiscoWebOrderRepository
from src.infrastructure.database.repositories.subscription_ignored_repository import SubscriptionIgnoredRepository
from src.services.matching.company_name_matching import generate_and_store_suggestions
from src.domain.import_schemas import IMPORT_SCHEMAS

BASE_INPUT_PATH = Path("/home/bridgeadoption/storage/input")

repo_ea = CiscoEARepository()
repo_prod = ProductRepository()
repo_log = ImportLogRepository()
repo_task = TaskRepository()
repo_activity = TaskActivityRepository()
repo_history = TaskHistoryRepository()
repo_company = CompanyListNameRepository()
repo_weborder = CiscoWebOrderRepository()
repo_subscription_ignored = SubscriptionIgnoredRepository()

IMPORT_SOURCE = "CiscoSubscriptionCCW"


@dataclass
class ImportSummary:
    file_path: str
    total_rows: int
    imported_rows: int
    error_rows: int


# Funções auxiliares (normalização, parsing)


def _normalize_str(value) -> Optional[str]:
    """
    Normaliza qualquer valor para string stripada. Retorna None para valores vazios ou NaN.
    """
    import pandas as pd
    if pd.isna(value):
        return None
    s = str(value).strip()
    return s if s else None


def _to_float(value) -> Optional[float]:
    """
    Converte números para float Python, removendo caracteres não numéricos,
    exceto o ponto decimal e o sinal.
    Retorna None para valores vazios ou inválidos.
    """
    import pandas as pd
    import re

    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    s = str(value).strip()
    if not s:
        return None

    # Mantém apenas dígitos, sinal e ponto decimal.
    # Ex.: "498717.824112002", "236954.088", "125810.784"
    cleaned = re.sub(r"[^0-9\.\-+]", "", s)
    if not cleaned or cleaned in {".", "-", "+", "-.", "+."}:
        return None

    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_date(value) -> Optional[datetime]:
    """
    Converte diferentes formatos de datas (strings Excel) em objetos datetime.
    Inclui suporte a formatos YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, etc.
    """
    from datetime import datetime, date
    import pandas as pd
    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    s = str(value).strip()[:50]
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _normalize_weborder(value: Optional[str]) -> Optional[str]:
    """
    Normaliza a coluna WebOrderID garantindo um identificador válido para consulta no banco.
    """
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if text.lower() in invalid_markers:
        return None
    if len(text) < 3:
        return None
    return text


def _get_customer_id_from_end_customer(value: Optional[str]) -> int:
    """
    Recupera o ID do cliente a partir do texto do End Customer.
    Caso não encontre, dispara o processo de matching e retorna 0 (sem bloqueio do resto do import).
    """
    if value is None:
        return 0
    name = str(value).strip()
    if not name:
        return 0
    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if name.lower() in invalid_markers:
        return 0
    if len(name) < 2:
        return 0
    try:
        company_id = repo_company.get_company_id_by_name(name)
    except Exception:
        generate_and_store_suggestions([name], user="from Cisco Subscription CCW")
        return 0
    if not company_id:
        generate_and_store_suggestions([name], user="from Cisco Subscription CCW")
        return 0
    return int(company_id)


def _get_or_create_product_id_from_offer_name(value: Optional[str]) -> int:
    """
    Busca ou cria um produto Cisco (vendor_id=1) com base na coluna Offer Name.
    Caso não consiga localizar nem inserir, retorna 0.
    """
    if value is None:
        return 0
    name = str(value).strip()
    if not name:
        return 0
    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if name.lower() in invalid_markers:
        return 0
    if len(name) < 2:
        return 0
    CISCO_VENDOR_ID = 1
    try:
        product_ids = repo_prod.find_ids_by_name_or_partnumber(name_or_part=name, vendor_id=CISCO_VENDOR_ID)
    except AttributeError:
        where_fallback = {"product_vendor_id": CISCO_VENDOR_ID}
        where_fallback["product_name"] = name
        try:
            product_ids = repo_prod.find_ids_by(where_fallback)
        except Exception:
            product_ids = []
        if not product_ids:
            where_fallback.pop("product_name", None)
            where_fallback["product_part_number"] = name
            try:
                product_ids = repo_prod.find_ids_by(where_fallback)
            except Exception:
                product_ids = []
    except Exception:
        product_ids = []

    if product_ids:
        try:
            return int(product_ids[0])
        except Exception:
            return 0
    data = {
        "product_vendor_id": CISCO_VENDOR_ID,
        "product_name": name,
        "product_part_number": name,
    }
    try:
        new_id = repo_prod.insert(data)
        return int(new_id) if new_id else 0
    except Exception:
        return 0


def _split_to_list(value: Optional[str]) -> List[str]:
    """
    Divide uma string com vírgulas em lista de valores stripados.
    Utilizada para normalizar TF Groups e Over Consumed TF Groups.
    """
    if not value:
        return []
    parts = [p.strip() for p in str(value).split(",")]
    return [p for p in parts if p]


def _build_ea_payload_from_row(row: Dict[str, Any], customer_id: int, product_id: int, weborder_str: str) -> Dict[str, Any]:
    """
    Constrói o dicionário final que será persistido em tbCiscoEA a partir da linha da planilha.
    """
    norm = _normalize_str
    to_float = _to_float
    to_dt = _to_date
    return {
        "ea_end_customer_id": customer_id,
        "ea_product_id": product_id,
        "ea_web_order_id": weborder_str,
        "ea_consumption_status": norm(row.get("Consumption Status")),
        "ea_over_consumed_tf_groups": norm(row.get("Over Consumed TF Groups")),
        "ea_tf_groups": norm(row.get("TF Groups")),
        "ea_tf_effective_date": to_dt(row.get("True Forward Effective Date")),
        "ea_next_tf": to_dt(row.get("Next True Forward")),
        "ea_subscription_id": norm(row.get("Subscription ID")),
        "ea_ccw_line_status": norm(row.get("Status")),
        "ea_start_date": to_dt(row.get("Start Date")),
        "ea_end_date": to_dt(row.get("End Date")),
        "ea_inicial_term": to_float(row.get("Initial Term")),
        "ea_renewal_date": to_dt(row.get("Renewal Date")),
        "ea_currency": norm(row.get("Currency")),
        "ea_mrc": to_float(row.get("Monthly Charge")),
        "ea_tf_overage": to_float(row.get("TF Overage")),
        "ea_po": norm(row.get("Purchase Order Number")),
        "ea_buying_program_id": norm(row.get("Buying Program ID")),
        "ea_site_url": norm(row.get("Site URL")),
        "ea_customer_success_manager": norm(row.get("Customer Success Manager")),
        "ea_sales_specialist": norm(row.get("Sales Specialist")),
        "ea_customer_success_manager_email": norm(row.get("Customer Success Manager Email")),
        "ea_sales_specialist_email": norm(row.get("Sales Specialist Email")),
        "ea_primary_billing_contact_name": norm(row.get("Primary Billing Contact Name")),
        "ea_primary_billing_contact_email": norm(row.get("Primary Billing Contact Email")),
        "ea_service_contact_name": norm(row.get("Service To Contact Name")),
        "ea_service_contact_email": norm(row.get("Service To Contact Email")),
        "ea_end_customer_contact_name": norm(row.get("End Customer Contact Name")),
        "ea_end_customer_contact_email": norm(row.get("End Customer Contact Email")),
        "ea_end_customer_contact_phone": norm(row.get("End Customer Contact Phone")),
        "ea_order_submit_date": to_dt(row.get("Order Submitted Date")),
        "ea_smart_account_name": norm(row.get("Smart Account Name")),
        "ea_renewal_manager": norm(row.get("Renewal Manager")),
        "ea_renewal_manager_email": norm(row.get("Renewal Manager Email")),
        "ea_provisioning_status": norm(row.get("Provisioning Status")),
    }


# Ajuda para abrir planilha e processar em blocos


def _open_workbook_rw(file_path: str) -> Tuple[Worksheet, Dict[str, int]]:
    """
    Abre o arquivo Excel em modo leitura/escrita e monta o índice das colunas.
    """
    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx_map = {h.strip() if h is not None else "": i for i, h in enumerate(header_row)}
    return ws, idx_map


def _process_chunk(
    ws: Worksheet,
    idx_map: Dict[str, int],
    start_row: int,
    end_row: int,
    file_path: str,
    imported_rows: int,
    error_rows: int,
) -> Tuple[int, int]:
    """
    Processa um bloco (chunk) de linhas da planilha, aplicando toda a regra de negócio.
    Mantém contador incremental de sucessos e falhas.
    IMPORTANTE: linhas importadas com sucesso são removidas da planilha,
    de forma que ao final do processo o arquivo fique apenas com as linhas com erro.
    """
    # Percorre de baixo para cima para permitir remoção segura de linhas.
    for r in range(end_row, start_row - 1, -1):
        row_cells = ws[r]
        raw_values = [cell.value for cell in row_cells]
        row_dict = {}
        for h, idx in idx_map.items():
            row_dict[h] = raw_values[idx] if idx < len(raw_values) else None
        try:
            weborder_raw = row_dict.get("WebOrderID")
            weborder_str = _normalize_weborder(weborder_raw)
            if not weborder_str:
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    r,
                    "WebOrder inválida ou ausente em 'WebOrderID'",
                    str(weborder_raw) if weborder_raw is not None else "",
                    None,
                )
                error_rows += 1
                continue

            end_customer_raw = row_dict.get("End Customer")
            end_customer_name = _normalize_str(end_customer_raw)
            end_customer_id = _get_customer_id_from_end_customer(end_customer_name)
            if not end_customer_id:
                # Customer não encontrado: logar nome da coluna do Excel + valor da célula
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    r,
                    "Customer não encontrado a partir de 'End Customer'",
                    "End Customer",        # importlog_column: cabeçalho da planilha
                    end_customer_name,     # importlog_value: valor da célula "End Customer"
                )
                error_rows += 1
                continue

            try:
                # Busca a WebOrder existente pelo par (weborder_number, weborder_customer_id)
                ciscoweborder_id = repo_weborder.find_id_by_code_and_customer(
                    code=weborder_str,
                    customer_id=end_customer_id,
                )
            except Exception as e:
                # coluna de origem: "WebOrderID"
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    r,
                    f"Erro ao buscar CiscoWebOrder: {e}",
                    "WebOrderID",          # importlog_column
                    weborder_str,          # importlog_value
                )
                ciscoweborder_id = None

            if not ciscoweborder_id:
                try:
                    ciscoweborder_id = repo_weborder.insert(
                        {
                            "weborder_number": weborder_str,
                            "weborder_customer_id": end_customer_id,
                        }
                    )
                except Exception as e:
                    # erro de gravação da WebOrder; coluna de origem continua sendo "WebOrderID"
                    repo_log.create(
                        IMPORT_SOURCE,
                        file_path,
                        r,
                        f"Falha ao criar WebOrder na tbCiscoWebOrder: {e}",
                        "WebOrderID",      # importlog_column
                        weborder_str,      # importlog_value
                    )
                    error_rows += 1
                    continue

            offer_name_raw = row_dict.get("Offer Name")
            offer_name = _normalize_str(offer_name_raw)
            product_id = _get_or_create_product_id_from_offer_name(offer_name)
            if not product_id:
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    r,
                    "Product não encontrado/criado a partir de 'Offer Name'",
                    offer_name,
                    weborder_str,
                )
                error_rows += 1
                continue

            payload = _build_ea_payload_from_row(row_dict, end_customer_id, product_id, weborder_str)

            offer_name_for_task = (offer_name or "").strip().upper()

            ea_subscription_id = payload.get("ea_subscription_id")
            ea_start_date = payload.get("ea_start_date")
            ea_end_date = payload.get("ea_end_date")
            ea_renewal_date = payload.get("ea_renewal_date")

            # Verifica se a subscription deve ser ignorada (tbSubscriptionIgnored)
            ignore_subscription = False
            if ea_subscription_id:
                try:
                    ignored_id = repo_subscription_ignored.get_id(
                        subscription_id=ea_subscription_id,
                        customer_id=end_customer_id,
                    )
                    if ignored_id:
                        ignore_subscription = True
                except Exception:
                    # Em caso de erro na checagem, por segurança NÃO considera como ignorada
                    ignore_subscription = False

            where: Dict[str, Any] = {
                "ea_web_order_id": weborder_str,
                "ea_end_customer_id": end_customer_id,
                "ea_product_id": product_id,
                "ea_start_date": ea_start_date,
                "ea_end_date": ea_end_date,
                "ea_renewal_date": ea_renewal_date,
            }
            if ea_subscription_id:
                where["ea_subscription_id"] = ea_subscription_id
            else:
                where["ea_subscription_id"] = None

            try:
                ea_ids = repo_ea.find_ids_by(where)
            except Exception as e:
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    r,
                    f"Erro ao buscar EA existente: {e}",
                    offer_name,
                    weborder_str,
                )
                error_rows += 1
                continue

            if ea_ids:
                try:
                    ea_id = int(ea_ids[0])
                    repo_ea.update(payload, where={"ea_id": ea_id})
                except Exception as e:
                    msg_err = str(e)
                    # Mapear coluna do banco para coluna de origem no Excel
                    log_column = offer_name
                    log_value = weborder_str

                    if "ea_primary_billing_contact_email" in msg_err:
                        log_column = "Primary Billing Contact Email"
                        log_value = row_dict.get("Primary Billing Contact Email")
                    elif "ea_service_contact_name" in msg_err:
                        log_column = "Service To Contact Name"
                        log_value = row_dict.get("Service To Contact Name")
                    elif "ea_over_consumed_tf_groups" in msg_err:
                        log_column = "Over Consumed TF Groups"
                        log_value = row_dict.get("Over Consumed TF Groups")
                    elif "ea_tf_groups" in msg_err:
                        log_column = "TF Groups"
                        log_value = row_dict.get("TF Groups")
                    elif "ea_smart_account_name" in msg_err:
                        log_column = "Smart Account Name"
                        log_value = row_dict.get("Smart Account Name")
                    elif "ea_end_customer_contact_name" in msg_err:
                        log_column = "End Customer Contact Name"
                        log_value = row_dict.get("End Customer Contact Name")
                    elif "ea_end_customer_contact_email" in msg_err:
                        log_column = "End Customer Contact Email"
                        log_value = row_dict.get("End Customer Contact Email")

                    repo_log.create(
                        IMPORT_SOURCE,
                        file_path,
                        r,
                        f"Erro ao atualizar EA existente (ea_id={ea_id}): {msg_err}",
                        log_column,
                        log_value,
                    )
                    error_rows += 1
                    continue

                consumption_status = payload.get("ea_consumption_status")
                if (
                    not ignore_subscription
                    and consumption_status
                    and consumption_status.upper() == "OVER CONSUMED"
                ):
                    _handle_over_consumed(
                        payload,
                        row_dict,
                        file_path,
                        r,
                        offer_name,
                        weborder_str,
                        ea_subscription_id,
                        end_customer_id,
                    )
            else:
                # Inserção de um novo EA com base nos dados da planilha.
                try:
                    ea_id_raw = repo_ea.insert(payload)
                    ea_id = int(ea_id_raw) if ea_id_raw else 0
                except Exception as e:
                    msg_err = str(e)
                    # Mapear coluna do banco para coluna de origem no Excel
                    log_column = offer_name
                    log_value = weborder_str

                    if "ea_primary_billing_contact_email" in msg_err:
                        log_column = "Primary Billing Contact Email"
                        log_value = row_dict.get("Primary Billing Contact Email")
                    elif "ea_service_contact_name" in msg_err:
                        log_column = "Service To Contact Name"
                        log_value = row_dict.get("Service To Contact Name")
                    elif "ea_over_consumed_tf_groups" in msg_err:
                        log_column = "Over Consumed TF Groups"
                        log_value = row_dict.get("Over Consumed TF Groups")
                    elif "ea_tf_groups" in msg_err:
                        log_column = "TF Groups"
                        log_value = row_dict.get("TF Groups")
                    elif "ea_smart_account_name" in msg_err:
                        log_column = "Smart Account Name"
                        log_value = row_dict.get("Smart Account Name")
                    elif "ea_end_customer_contact_name" in msg_err:
                        log_column = "End Customer Contact Name"
                        log_value = row_dict.get("End Customer Contact Name")
                    elif "ea_end_customer_contact_email" in msg_err:
                        log_column = "End Customer Contact Email"
                        log_value = row_dict.get("End Customer Contact Email")
                    elif "ea_over_consumed_tf_groups" in msg_err:
                        log_column = "Over Consumed TF Groups"
                        log_value = row_dict.get("Over Consumed TF Groups")

                    repo_log.create(
                        IMPORT_SOURCE,
                        file_path,
                        r,
                        f"Erro ao inserir novo EA: {msg_err}",
                        log_column,
                        log_value,
                    )
                    error_rows += 1
                    continue

                if not ea_id:
                    repo_log.create(
                        IMPORT_SOURCE,
                        file_path,
                        r,
                        "Falha ao obter ea_id após inserir novo EA",
                        offer_name,
                        weborder_str,
                    )
                    error_rows += 1
                    continue

                # Criação automática de tarefa tipo 1 (New EA) somente se a subscription NÃO estiver ignorada
                if (
                    not ignore_subscription
                    and offer_name_for_task in {"EA3-M", "ELA2-M", "A-FLEX", "A-FLEX-3"}
                ):
                    try:
                        owner_id = repo_task.get_last_task_owner_by_company(
                            company_id=end_customer_id,
                            user_type="CSM",
                        )
                    except Exception:
                        owner_id = 0

                    now = datetime.now()
                    task_start = now.date()
                    task_end = task_start + timedelta(days=45)

                    sub_id = ea_subscription_id
                    po_num = _normalize_str(row_dict.get("Purchase Order Number"))

                    # Monta task_reference conforme especificação:
                    # "Subscription ID: <valor>", "PO: <valor>", "WEB ORDER: <valor>"
                    ref_parts = []
                    if sub_id:
                        ref_parts.append(f"Subscription ID: {sub_id}")
                    if po_num:
                        ref_parts.append(f"PO: {po_num}")
                    if weborder_str:
                        ref_parts.append(f"WEB ORDER: {weborder_str}")

                    task_reference = ", ".join(ref_parts) if ref_parts else None

                    task_data = {
                        "task_tasktype_id": 1,
                        "task_reference": task_reference,
                        "task_owner_id": owner_id,
                        "task_temp_owner_id": None,
                        "task_customer_id": end_customer_id,
                        "task_cr_party_id": 0,
                        "task_cr_party_name": None,
                        "task_created_in": now,
                        "task_created_by": 0,
                        "task_priority": "LOW",
                        "task_project_id": 0,
                        "task_status": 1,
                        "task_status_justification": None,
                        "task_start": task_start,
                        "task_end": task_end,
                        "task_start_performed": task_start,
                        "task_end_performed": task_end,
                        "task_value": 0,
                        "task_forecast": 0,
                        "task_backlog": 0,
                        "task_rate": 1,
                        "task_currency": "USD",
                        "task_ws": sub_id if sub_id else None,
                        "task_deal_id": None,
                        "task_track": offer_name_for_task,
                        "task_subtrack": offer_name_for_task,
                        "task_highlight": 0,
                        "task_remark": None,
                        "task_description": None,
                        "task_ea_flag": 0,
                        "task_telemetry_flag": 0,
                        "task_opt_in_flag": 0,
                        "task_completed": 0,
                        "task_architecture": None,
                        "task_solution_domain": None,
                        "task_eligible": "Y",
                        "task_end_fy": None,
                        "task_booking_date": None,
                        "task_booking_amount": None,
                    }
                    try:
                        task_id_raw = repo_task.insert(task_data)
                        task_id = int(task_id_raw) if task_id_raw else 0
                    except Exception as e:
                        repo_log.create(
                            IMPORT_SOURCE,
                            file_path,
                            r,
                            f"Erro ao criar tarefa tipo 1 (New CISCO EA): {e}",
                            offer_name,
                            weborder_str,
                        )
                        error_rows += 1
                        continue

                    if task_id:
                        try:
                            repo_ea.update(
                                {"ea_new_task_id": task_id},
                                where={"ea_id": ea_id},
                            )
                        except Exception:
                            pass

                        try:
                            remark_date_str = now.strftime("%Y-%b-%d")
                            next_followup_date = (now.date() + timedelta(days=45)).strftime("%Y-%m-%d")
                            history_data = {
                                "taskrecord_task_id": task_id,
                                "taskrecord_activity_id": 0,
                                "taskrecord_remark": f"Task created at {remark_date_str}",
                                "taskrecord_next_followup": next_followup_date,
                                "taskrecord_updated_by": "System BA",
                            }
                            repo_history.insert(history_data)
                        except Exception:
                            pass

            # Sucesso: remove a linha da planilha para manter apenas registros com erro no final
            ws.delete_rows(r, 1)
            imported_rows += 1

        except Exception as ex:
            # Erro inesperado no processamento da linha.
            # Por padrão, vamos registrar a coluna "End Customer" como referência
            # e o valor correspondente daquela célula.
            repo_log.create(
                IMPORT_SOURCE,
                file_path,
                r,
                "Unexpected error: " + str(ex),
                "End Customer",      # importlog_column: coluna de origem
                end_customer_name,   # importlog_value: valor da célula
            )
            error_rows += 1

    return imported_rows, error_rows


def _handle_over_consumed(
    payload,
    row_dict,
    file_path,
    row_number,
    offer_name,
    weborder_str,
    ea_subscription_id,
    task_customer_id,
):
    """
    Implementa o PASSO 7:
    - Busca/atualiza tarefas tipo 35 para casos de Over Consumed.
    - Cria atividades correspondentes ou reabre tarefas existentes.
    - Atualiza histórico de task/activity conforme necessidade.
    """
    now = datetime.now()

    def build_non_null_ref(*args) -> Optional[str]:
        parts = [str(x).strip() for x in args if x and str(x).strip()]
        return ", ".join(parts) if parts else None

    task_references = [
        str(row_dict.get("Next True Forward")).strip() if row_dict.get("Next True Forward") else None,
        ea_subscription_id,
        row_dict.get("Purchase Order Number"),
        weborder_str,
    ]

    offer_name_val = (offer_name or "").strip().upper()
    over_consumed_tf_groups = payload.get("ea_over_consumed_tf_groups")
    tf_groups = payload.get("ea_tf_groups")

    track_keywords = [offer_name_val]
    if over_consumed_tf_groups:
        track_keywords += _split_to_list(over_consumed_tf_groups)

    subtrack_keywords = _split_to_list(tf_groups)

    found_task_info = None

    for ref in task_references:
        if not ref:
            continue
        for track_kw in track_keywords:
            if not track_kw:
                continue
            for subtrack_kw in subtrack_keywords or [None]:
                found_task_info = repo_ea.find_task_over_consumed(
                    task_customer_id=task_customer_id,
                    task_ws=ea_subscription_id,
                    task_reference=ref,
                    track_keyword=track_kw,
                    subtrack_keyword=subtrack_kw,
                )
                if found_task_info:
                    break
            if found_task_info:
                break
        if found_task_info:
            break

    def parse_date_str(value):
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value.strip(), "%Y-%m-%d").date()
            except Exception:
                return None
        return None

    true_forward_date = parse_date_str(row_dict.get("True Forward Effective Date"))
    next_true_forward_date = parse_date_str(row_dict.get("Next True Forward"))

    task_start_date = now.date()
    if true_forward_date and true_forward_date > task_start_date:
        task_end_date = true_forward_date - timedelta(days=7)
    elif next_true_forward_date:
        task_end_date = next_true_forward_date - timedelta(days=14)
    else:
        task_end_date = task_start_date + timedelta(days=30)

    monthly_charge = payload.get("ea_mrc")
    try:
        task_value = float(monthly_charge) if monthly_charge else 0.0
    except Exception:
        task_value = 0.0

    currency = payload.get("ea_currency") or "USD"

    if not found_task_info:
        # Cenário sem tarefa: criação completa da task tipo 35 e activity.
        try:
            owner_id = repo_task.get_last_task_owner_by_company(
                company_id=task_customer_id,
                user_type="CSM",
            )
        except Exception:
            owner_id = 0

        # Monta task_reference concatenando valores não nulos conforme especificação:
        # "TF Effective Date: " + True Forward Effective Date (string original)
        # "Next TF: " + Next True Forward (string original)
        # "Subscription: " + Subscription ID
        # "PO: " + Purchase Order Number
        # "WebOrderID: " + WebOrderID
        tf_effective_raw = row_dict.get("True Forward Effective Date")
        next_tf_raw = row_dict.get("Next True Forward")
        sub_id_raw = row_dict.get("Subscription ID")
        po_raw = row_dict.get("Purchase Order Number")

        parts_ref = []
        if tf_effective_raw not in (None, ""):
            parts_ref.append(f"TF Effective Date: {str(tf_effective_raw).strip()}")
        if next_tf_raw not in (None, ""):
            parts_ref.append(f"Next TF: {str(next_tf_raw).strip()}")
        if sub_id_raw not in (None, ""):
            parts_ref.append(f"Subscription: {str(sub_id_raw).strip()}")
        if po_raw not in (None, ""):
            parts_ref.append(f"PO: {str(po_raw).strip()}")
        if weborder_str:
            parts_ref.append(f"WebOrderID: {weborder_str}")

        task_reference_35 = ", ".join(parts_ref) if parts_ref else None

        task_data = {
            "task_tasktype_id": 35,
            "task_reference": task_reference_35,
            "task_owner_id": owner_id,
            "task_temp_owner_id": None,
            "task_customer_id": task_customer_id,
            "task_cr_party_id": 0,
            "task_cr_party_name": None,
            "task_created_in": now,
            "task_created_by": 0,
            "task_priority": "HIGH",
            "task_project_id": 0,
            "task_status": 1,
            "task_status_justification": None,
            "task_start": task_start_date,
            "task_end": task_end_date,
            "task_start_performed": task_start_date,
            "task_end_performed": task_end_date,
            "task_value": task_value,
            "task_forecast": 0,
            "task_backlog": 0,
            "task_rate": 1,
            "task_currency": currency,
            "task_ws": ea_subscription_id if ea_subscription_id else None,
            "task_deal_id": None,
            "task_track": build_non_null_ref(offer_name_val, over_consumed_tf_groups),
            "task_subtrack": tf_groups if tf_groups else None,
            "task_highlight": 1,
            "task_remark": None,
            "task_description": None,
            "task_ea_flag": 0,
            "task_telemetry_flag": 0,
            "task_opt_in_flag": 0,
            "task_completed": 0,
            "task_architecture": None,
            "task_solution_domain": None,
            "task_eligible": "Y",
            "task_end_fy": None,
            "task_booking_date": None,
            "task_booking_amount": None,
        }

        try:
            task_id_raw = repo_task.insert(task_data)
            task_id = int(task_id_raw) if task_id_raw else 0
        except Exception as e:
            repo_log.create(
                IMPORT_SOURCE,
                file_path,
                row_number,
                f"Erro ao criar tarefa tipo 35 (Over Consumed): {e}",
                offer_name,
                weborder_str,
            )
            return

        activity_data = {
            "activity_task_id": task_id,
            "activity_name": f"Over Consumed: {offer_name_val}",
            "activity_status": 1,
            "activity_ws": ea_subscription_id if ea_subscription_id else None,
            "activity_track": build_non_null_ref(offer_name_val, over_consumed_tf_groups),
            "activity_sub_track": tf_groups if tf_groups else None,
            "activity_start": task_start_date,
            "activity_end": task_end_date,
            "activity_start_performed": task_start_date,
            "activity_end_performed": task_end_date,
            "activity_value": task_value,
            "activity_currency": currency,
        }

        try:
            activity_id_raw = repo_activity.insert(activity_data)
            activity_id = int(activity_id_raw) if activity_id_raw else 0
        except Exception as e:
            repo_log.create(
                IMPORT_SOURCE,
                file_path,
                row_number,
                f"Erro ao criar atividade para task tipo 35: {e}",
                offer_name,
                weborder_str,
            )
            return

        try:
            remark_str = now.strftime("%Y-%b-%d")
            followup_date = (now.date() + timedelta(days=30)).strftime("%Y-%m-%d")
            history_task = {
                "taskrecord_task_id": task_id,
                "taskrecord_activity_id": 0,
                "taskrecord_remark": f"Task created at {remark_str}",
                "taskrecord_next_followup": followup_date,
                "taskrecord_updated_by": "System BA",
            }
            repo_history.insert(history_task)
        except Exception:
            pass

        try:
            history_activity = {
                "taskrecord_task_id": task_id,
                "taskrecord_activity_id": activity_id,
                "taskrecord_remark": f"Activity created at {remark_str}",
                "taskrecord_next_followup": (now.date() + timedelta(days=7)).strftime("%Y-%m-%d"),
                "taskrecord_updated_by": "System BA",
            }
            repo_history.insert(history_activity)
        except Exception:
            pass

    else:
        # Cenário com tarefa existente: aplicar análises de atualização e reabertura.
        task_id = found_task_info.get("task_id")
        task_status_id = found_task_info.get("task_status_id")
        activity_id = found_task_info.get("activity_id")
        activity_status_id = found_task_info.get("activity_status_id")

        def union_update_list(current: str, new: str) -> Optional[str]:
            current_list = [v.strip() for v in (current or "").split(",") if v.strip()]
            new_list = [v.strip() for v in (new or "").split(",") if v.strip()]
            union = current_list.copy()
            for v in new_list:
                if v not in union:
                    union.append(v)
            if set(union) != set(current_list):
                return ", ".join(union)
            return None

        task_columns = repo_task.get_columns_by_task_id(
            task_id=task_id,
            columns=["task_track", "task_subtrack"],
            as_df=False,
        )
        activity_columns = repo_activity.get_columns_by_activity_id(
            activity_id=activity_id,
            columns=["activity_status", "activity_value", "activity_track", "activity_sub_track"],
            as_df=False,
        )

        task_track = task_columns.get("task_track", "") if task_columns else ""
        task_subtrack = task_columns.get("task_subtrack", "") if task_columns else ""

        activity_track = activity_columns.get("activity_track", "") if activity_columns else ""
        activity_subtrack = activity_columns.get("activity_sub_track", "") if activity_columns else ""

        monthly_charge_val = task_value

        need_new_task = False
        need_new_activity = False
        update_task_fields = {}
        update_activity_fields = {}

        combined_track = build_non_null_ref(offer_name_val, over_consumed_tf_groups)
        combined_subtrack = tf_groups if tf_groups else ""

        new_task_track = union_update_list(task_track, combined_track)
        new_task_subtrack = union_update_list(task_subtrack, combined_subtrack)

        if set(combined_track.split(", ")) != set(task_track.split(", ")):
            need_new_task = True

        if new_task_track:
            update_task_fields["task_track"] = new_task_track

        if set(combined_subtrack.split(", ")) != set(task_subtrack.split(", ")):
            need_new_task = True

        if new_task_subtrack:
            update_task_fields["task_subtrack"] = new_task_subtrack

        new_activity_track = union_update_list(activity_track, combined_track)
        new_activity_subtrack = union_update_list(activity_subtrack, combined_subtrack)

        activity_value_current = activity_columns.get("activity_value", 0) if activity_columns else 0

        # Normaliza valor atual da activity para float seguro,
        # evitando comparação entre float e NoneType
        try:
            activity_value_current_num = (
                float(activity_value_current)
                if activity_value_current is not None
                else 0.0
            )
        except Exception:
            activity_value_current_num = 0.0

        if new_activity_track is None and monthly_charge_val > activity_value_current_num:
            need_new_activity = True

        if new_activity_track:
            update_activity_fields["activity_track"] = new_activity_track

        if new_activity_subtrack:
            update_activity_fields["activity_sub_track"] = new_activity_subtrack

        if update_task_fields:
            # Reabertura da tarefa com registro em histórico para auditoria.
            update_task_fields["task_status"] = 2
            try:
                repo_task.update(update_task_fields, where={"task_id": task_id})
                repo_history.insert(
                    {
                        "taskrecord_task_id": task_id,
                        "taskrecord_activity_id": 0,
                        "taskrecord_remark": "Task reopened with updated tracks",
                        "taskrecord_updated_by": "System BA",
                    }
                )
            except Exception:
                pass

        if task_status_id not in (4, 5, 6, 10) and activity_id == 0:
            need_new_activity = True
        elif task_status_id not in (4, 5, 6, 10) and activity_status_id not in (4, 5, 6, 10):
            need_new_activity = False
        elif task_status_id not in (4, 5, 6, 10) and activity_status_id in (0, 4, 5, 6, 10):
            if not new_activity_track:
                need_new_activity = True

        if need_new_activity:
            # Criação de nova atividade quando não há atividade aberta compatível.
            try:
                activity_name = f"Over Consumed: {offer_name_val}"
                new_activity_data = {
                    "activity_task_id": task_id,
                    "activity_name": activity_name,
                    "activity_status": 1,
                    "activity_ws": ea_subscription_id if ea_subscription_id else None,
                    "activity_track": combined_track,
                    "activity_sub_track": combined_subtrack,
                    "activity_start": task_start_date,
                    "activity_end": task_end_date,
                    "activity_start_performed": task_start_date,
                    "activity_end_performed": task_end_date,
                    "activity_value": task_value,
                    "activity_currency": currency,
                }
                new_activity_id_raw = repo_activity.insert(new_activity_data)
                new_activity_id = int(new_activity_id_raw) if new_activity_id_raw else 0

                if new_activity_id:
                    repo_history.insert(
                        {
                            "taskrecord_task_id": task_id,
                            "taskrecord_activity_id": new_activity_id,
                            "taskrecord_remark": f"Activity created at {now.strftime('%Y-%b-%d')}",
                            "taskrecord_next_followup": (now.date() + timedelta(days=7)).strftime("%Y-%m-%d"),
                            "taskrecord_updated_by": "System BA",
                        }
                    )
            except Exception as e:
                repo_log.create(
                    IMPORT_SOURCE,
                    file_path,
                    row_number,
                    f"Erro ao criar nova atividade para task tipo 35: {e}",
                    offer_name,
                    weborder_str,
                )
                return

        if update_activity_fields:
            update_activity_fields["activity_status"] = 2
            try:
                repo_activity.update(update_activity_fields, where={"activity_id": activity_id})
                repo_history.insert(
                    {
                        "taskrecord_task_id": task_id,
                        "taskrecord_activity_id": activity_id,
                        "taskrecord_remark": "Activity reopened with updated tracks",
                        "taskrecord_updated_by": "System BA",
                    }
                )
            except Exception:
                pass


def run_import(file_path: str, user_id: Optional[str] = None):
    """
    Função principal responsável por:
    1. Validar cabeçalho da planilha (schema obrigatório).
    2. Processar lotes de linhas em chunks, aplicando as regras de negócio.
    3. Criar/atualizar registros de EA, tarefas e atividades conforme PASSO 7.
    4. Registrar logs detalhados no ImportLogRepository para auditoria.

    Retorno compatível com import_scheduling:
        {
            "status": "FINISHED" | "PENDING" | "FAILED",
            "message": str,
            "summary": {
                "file_path": str,
                "total_rows": int,
                "imported_rows": int,
                "error_rows": int,
            },
        }
    """
    CHUNK_SIZE = 2000
    imported_rows = 0
    error_rows = 0

    # Resolve caminho relativo para a pasta padrão de input
    path = Path(file_path)
    if not path.is_absolute():
        path = BASE_INPUT_PATH / path.name

    # erro ao abrir arquivo
    try:
        ws, idx_map = _open_workbook_rw(str(path))
    except Exception as e:
        msg = f"Erro ao abrir arquivo Excel: {e}"
        repo_log.create(
            IMPORT_SOURCE,
            str(path),
            0,
            msg,
            None,
            None,
        )
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(path),
                "total_rows": 0,
                "imported_rows": 0,
                "error_rows": 0,
            },
        }

    total_rows = ws.max_row - 1  # Ignora linha de cabeçalho.

    required_schema = IMPORT_SCHEMAS.get("CiscoSubscriptionCCW", {})
    required_columns = required_schema.get("required", [])

    headers = list(idx_map.keys())
    missing_cols = [c for c in required_columns if c not in headers]
    if missing_cols:
        msg = "Missing columns: " + ", ".join(missing_cols)
        repo_log.create(
            IMPORT_SOURCE,
            str(path),
            0,
            msg,
            None,
            None,
        )
        ws.parent.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(path),
                "total_rows": total_rows,
                "imported_rows": 0,
                "error_rows": total_rows,
            },
        }

    # PROCESSAMENTO SEGURO:
    # Sempre reprocessa a partir da linha 2, pois as linhas
    # são removidas dinamicamente dentro de _process_chunk.
    # Isso evita que linhas sejam "puladas".
    while ws.max_row > 1:
        chunk_end = min(1 + CHUNK_SIZE, ws.max_row)

        imported_chunk, error_chunk = _process_chunk(
            ws,
            idx_map,
            2,
            chunk_end,
            str(path),
            imported_rows,
            error_rows,
        )

        imported_rows += imported_chunk
        error_rows += error_chunk

    try:
        ws.parent.save(str(path))
    except Exception as e:
        msg = f"Erro ao salvar arquivo Excel após processamento: {e}"
        repo_log.create(
            IMPORT_SOURCE,
            str(path),
            0,
            msg,
            None,
            None,
        )
        # ainda assim fecha e retorna FAILED
        ws.parent.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(path),
                "total_rows": total_rows,
                "imported_rows": imported_rows,
                "error_rows": error_rows,
            },
        }

    ws.parent.close()

    # Define status final:
    # - FINISHED: todas as linhas foram importadas sem erro
    # - FAILED: houve qualquer linha com erro
    if error_rows > 0:
        status = "FAILED"
        msg = (
            f"CiscoSubscriptionCCW import concluído com falhas. "
            f"arquivo={str(path)}, total={total_rows}, "
            f"sucesso={imported_rows}, erros={error_rows}."
        )
    else:
        status = "FINISHED"
        msg = (
            f"CiscoSubscriptionCCW import concluído com sucesso. "
            f"arquivo={str(path)}, total={total_rows}, "
            f"sucesso={imported_rows}."
        )

    return {
        "status": status,
        "message": msg,
        "summary": {
            "file_path": str(path),
            "total_rows": total_rows,
            "imported_rows": imported_rows,
            "error_rows": error_rows,
        },
    }
