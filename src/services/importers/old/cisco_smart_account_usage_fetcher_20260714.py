"""
cisco_smart_account_usage_fetcher.py

Importador de uso/licenciamento Cisco Smart Account com foco em:
- correção de regra de negócio
- consolidação global de registros do tipo quantity
- segurança operacional para arquivos grandes
- reaproveitamento do padrão operacional do cisco_subscription_ccw.py

========================================================================
VISÃO GERAL
========================================================================

Este importador processa arquivos Excel (.xlsx) colocados em:

    .storage/input/

e produz:

1) atualização do arquivo original, deixando apenas linhas ainda não lidas
2) arquivo fixo de falhas em:
       .storage/output/<nome_original>__failed_rows.xlsx
   com append
3) log operacional em:
       .storage/logs/

Toda linha lida sai do arquivo original:
- sucesso: sai do original
- falha: sai do original e vai para o arquivo de falhas

Assim, em caso de interrupção:
- o arquivo original contém somente linhas ainda não lidas
- o arquivo de falhas mantém as linhas já processadas com erro

========================================================================
REGRAS DE NEGÓCIO
========================================================================

A origem Cisco não diferencia explicitamente “quantity” e “metering”.
Essa separação é uma regra da aplicação.

1) quantity
-----------
Representa a visão contratual/consolidada de licenças.

A origem pode trazer várias linhas equivalentes que, para a aplicação,
devem ser consolidadas em um único registro quantity.

As linhas são agrupadas pela chave lógica:

- Client
- Domain
- License
- Virtual Account
- Billing
- License Type
- Subscription Id
- Start Date
- End Date

Importante:
- a vigência faz parte do agrupamento
- se Start Date ou End Date mudar, é outro grupo
- o campo Active NÃO entra na chave de agrupamento
- o campo Compliance NÃO entra na chave de agrupamento
- esses campos são carregados da linha representativa do grupo

Após o agrupamento:
- a coluna Quantity é somada
- o total é gravado em mcsa_quantity

Exemplo:
4 linhas iguais na chave acima, cada uma com Quantity = 1,
geram 1 único registro quantity com mcsa_quantity = 4.

2) metering
-----------
Representa consumo/utilização.

Uma linha pode gerar metering quando houver informação em pelo menos um dos campos:
- Available To Use
- In Use
- Balance

Metering é tratado individualmente por linha elegível.
Não há soma global como no quantity.

3) Uma mesma linha pode gerar:
-------------------------------
- apenas quantity
- apenas metering
- quantity + metering

4) Cliente e produto
--------------------
Cliente:
- resolvido a partir da coluna "Client"
- busca via CompanyListNameRepository
- se não encontrar, a linha/grupo falha

Produto:
- resolvido a partir da coluna "License"
- busca via ProductRepository
- se não encontrar, a linha/grupo falha

========================================================================
REGRA DE UPDATE / INSERT
========================================================================

1) Quantity
-----------
- consolida primeiro
- procura registro existente na tbCiscoSmartAccountMetering
- se não existir: INSERT
- se existir: UPDATE

Quando quantity existente sofrer alteração em mcsa_quantity:
- mcsa_track = 0

2) Metering
-----------
- procura registro existente
- se não existir: INSERT
- se existir: UPDATE dos campos de consumo

Observação:
caso sua regra real para metering seja “não atualizar se já existir”,
basta ajustar a função _upsert_metering_record.

========================================================================
PROTEÇÕES DE PERFORMANCE / RECURSOS
========================================================================

Este arquivo foi desenhado para reduzir custo operacional em arquivos grandes
(50 mil+ linhas), adotando as seguintes proteções:

1. leitura única do arquivo
2. consolidação global de quantity em memória enxuta
3. cache de cliente
4. cache de produto
5. sem delete_rows() linha a linha
6. regravação única do arquivo original remanescente
7. append controlado no arquivo fixo de falhas
8. lock simples por arquivo para evitar concorrência acidental
9. logs operacionais enxutos
10. validações e normalizações antes de acessar banco

========================================================================
PREMISSAS DE REPOSITÓRIO
========================================================================

Este código assume a existência de:

- src.infrastructure.database.repositories.cisco_sa_repository.CiscoSARepository
- src.infrastructure.database.repositories.product_repository.ProductRepository
- src.infrastructure.database.repositories.import_log_repository.ImportLogRepository
- src.infrastructure.database.repositories.company_list_name_repository.CompanyListNameRepository

E assume que a tabela alvo é:
- tbCiscoSmartAccountMetering

Também assume que o repositório CiscoSARepository possui ao menos:
- find_first_by(where: dict) -> dict | None
- insert(data: dict) -> int
- update(data: dict, where: dict) -> int

========================================================================
OBSERVAÇÃO IMPORTANTE SOBRE MAPEAMENTOS
========================================================================

Como os detalhes exatos de alguns repositórios podem variar no projeto real,
foram criadas funções de resolução tolerantes para cliente e produto.

Se o ProductRepository ou CompanyListNameRepository tiver nomes de método
diferentes, ajustar apenas:

- _resolve_customer_id(...)
- _resolve_product_id(...)

mantendo o restante intacto.

========================================================================
"""

from __future__ import annotations

import os
import re
import shutil
import traceback
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook

from src.infrastructure.database.repositories.cisco_sa_repository import CiscoSARepository
from src.infrastructure.database.repositories.product_repository import ProductRepository
from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository


# ============================================================================
# PATHS / CONFIG
# ============================================================================

BASE_STORAGE_PATH = Path("storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"
BASE_LOCKS_PATH = BASE_STORAGE_PATH / "locks"

IMPORT_SOURCE = "CiscoSmartAccountUsage"

MAX_ROWS_SOFT_WARNING = 150_000


# ============================================================================
# REPOSITORIES
# ============================================================================

repo_mcsa = CiscoSARepository()
repo_product = ProductRepository()
repo_import_log = ImportLogRepository()
repo_company = CompanyListNameRepository()


# ============================================================================
# DATA STRUCTURES
# ============================================================================

@dataclass
class QuantityGroup:
    """
    Representa um grupo consolidado de linhas que devem virar 1 único registro
    quantity na tbCiscoSmartAccountMetering.

    group_key:
        chave lógica consolidada

    representative_row:
        linha-base usada para carregar campos complementares não usados na chave

    original_rows:
        linhas originais do Excel que pertencem ao grupo, com:
        - excel_row_number
        - row_data

    total_quantity:
        soma consolidada da coluna Quantity
    """
    group_key: Tuple[Any, ...]
    representative_row: Dict[str, Any]
    original_rows: List[Tuple[int, Dict[str, Any]]] = field(default_factory=list)
    total_quantity: Decimal = Decimal("0")


@dataclass
class ProcessingResult:
    """
    Resultado consolidado do processamento de 1 arquivo.
    """
    success_row_numbers: set = field(default_factory=set)
    failed_rows: List[Dict[str, Any]] = field(default_factory=list)
    execution_messages: List[str] = field(default_factory=list)

    def log(self, message: str) -> None:
        self.execution_messages.append(message)


# ============================================================================
# DIRECTORY / LOG / LOCK HELPERS
# ============================================================================

def _ensure_directories() -> None:
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOCKS_PATH.mkdir(parents=True, exist_ok=True)


def _timestamp() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _build_execution_log_path(source_file: Path) -> Path:
    return BASE_LOGS_PATH / f"{source_file.stem}.log"


def _append_execution_log(log_path: Path, message: str) -> None:
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{_timestamp()}] {message}\n")


def _safe_log(log_path: Path, result: ProcessingResult, message: str) -> None:
    result.log(message)
    _append_execution_log(log_path, message)


def _build_lock_path(source_file: Path) -> Path:
    safe_name = re.sub(r"[^A-Za-z0-9_.-]", "_", source_file.name)
    return BASE_LOCKS_PATH / f"{safe_name}.lock"


def _acquire_lock(lock_path: Path) -> None:
    """
    Lock simples por criação exclusiva de arquivo.

    Objetivo:
    - evitar duas execuções simultâneas processando o mesmo arquivo

    Observação:
    - este lock é simples e suficiente para scheduler serial/controlado
    """
    try:
        fd = os.open(str(lock_path), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(f"pid={os.getpid()} created_at={_timestamp()}\n")
    except FileExistsError:
        raise RuntimeError(f"Arquivo em processamento ou lock pendente: {lock_path.name}")


def _release_lock(lock_path: Path) -> None:
    if lock_path.exists():
        lock_path.unlink(missing_ok=True)


# ============================================================================
# NORMALIZATION HELPERS
# ============================================================================

def _normalize_str(value: Any, default: str = "-") -> str:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def _normalize_nullable_str(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _normalize_identifier(value: Any) -> str:
    """
    Normaliza identificadores e textos usados como parte de chave lógica.

    Mantém legibilidade, apenas removendo espaços extremos e colapsando espaços.
    """
    text = _normalize_str(value, default="-")
    text = re.sub(r"\s+", " ", text)
    return text


def _to_decimal(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    if value is None:
        return default

    if isinstance(value, Decimal):
        return value

    if isinstance(value, (int, float)):
        return Decimal(str(value))

    text = str(value).strip()
    if not text:
        return default

    text = text.replace(".", "").replace(",", ".") if "," in text and "." in text else text.replace(",", ".")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return default


def _safe_repo_int(value: Any) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except Exception:
        return None


def _to_date(value: Any) -> Optional[date]:
    """
    Normaliza datas para date puro.

    Aceita:
    - datetime/date
    - string dd/mm/yyyy
    - string yyyy-mm-dd
    - vazio -> None
    """
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _date_to_db(value: Optional[date]) -> Optional[str]:
    return value.isoformat() if value else None


def _has_meaningful_metering(row: Dict[str, Any]) -> bool:
    return any(
        _normalize_nullable_str(row.get(col)) is not None
        for col in ("Available To Use", "In Use", "Balance")
    )


def _has_meaningful_quantity(row: Dict[str, Any]) -> bool:
    raw = row.get("Quantity")
    if raw is None:
        return False
    if isinstance(raw, (int, float, Decimal)):
        return True
    text = str(raw).strip()
    return bool(text)


# ============================================================================
# EXCEL HELPERS
# ============================================================================

def _read_source_rows(source_file: Path) -> Tuple[List[str], List[Tuple[int, Dict[str, Any]]]]:
    """
    Lê o arquivo fonte uma única vez.

    Retorna:
    - cabeçalho
    - lista de tuplas (excel_row_number, row_dict)

    Observação:
    - usa read_only=True para reduzir consumo de memória na leitura
    """
    wb = load_workbook(filename=source_file, read_only=True, data_only=True)
    ws = wb.active

    iterator = ws.iter_rows(values_only=True)
    header_row = next(iterator, None)
    if not header_row:
        wb.close()
        return [], []

    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: List[Tuple[int, Dict[str, Any]]] = []

    for excel_row_number, values in enumerate(iterator, start=2):
        row_dict = {headers[idx]: values[idx] if idx < len(values) else None for idx in range(len(headers))}
        if all(v is None or str(v).strip() == "" for v in row_dict.values()):
            continue
        rows.append((excel_row_number, row_dict))

    wb.close()
    return headers, rows


def _write_remaining_source_rows(source_file: Path, headers: List[str], remaining_rows: List[Dict[str, Any]]) -> None:
    """
    Regrava o arquivo original com apenas linhas remanescentes.

    Estratégia:
    - cria arquivo temporário
    - grava tudo novamente
    - substitui o original atomicamente quando possível
    """
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        temp_path = Path(tmp.name)

    wb = Workbook()
    ws = wb.active
    ws.append(headers)

    for row in remaining_rows:
        ws.append([row.get(col) for col in headers])

    wb.save(temp_path)
    wb.close()

    shutil.move(str(temp_path), str(source_file))


def _build_failed_file_path(source_file: Path) -> Path:
    return BASE_OUTPUT_PATH / f"{source_file.stem}_failed_rows.xlsx"


def _ensure_failed_workbook(failed_file: Path, failed_headers: List[str]) -> None:
    """
    Cria (ou recria) o arquivo de falhas com apenas o header.

    Sempre sobrescreve o arquivo se já existir, garantindo que não haja
    resíduos de execuções anteriores. A criação só será feita de fato
    se esta função for chamada (ou seja, somente quando houver falhas
    a serem escritas).
    """
    wb = Workbook()
    ws = wb.active
    ws.append(failed_headers)
    wb.save(failed_file)
    wb.close()


def _append_failed_rows(failed_file: Path, failed_headers: List[str], failed_rows: List[Dict[str, Any]]) -> None:
    """
    Faz append das falhas no arquivo fixo.

    Estrutura:
    - colunas originais
    - colunas técnicas de erro ao final

    Regras:
    - Se failed_rows estiver vazio, não faz nada e não cria arquivo.
    - Se o arquivo ainda não existir, ele é criado/recriado com header
      (sobrescrevendo conteúdo anterior) via _ensure_failed_workbook.
    """
    if not failed_rows:
        return

    _ensure_failed_workbook(failed_file, failed_headers)

    wb = load_workbook(failed_file)
    ws = wb.active

    for row in failed_rows:
        ws.append([row.get(col) for col in failed_headers])

    wb.save(failed_file)
    wb.close()


# ============================================================================
# IMPORT LOG HELPERS
# ============================================================================

def _register_import_log(
    source_file: Path,
    status: str,
    description: str
) -> None:
    """
    Tenta registrar log na estrutura existente do projeto.
    Mantido tolerante para não quebrar a execução caso o repositório tenha
    interface diferente da esperada.
    """
    payload = {
        "source": IMPORT_SOURCE,
        "file_name": source_file.name,
        "status": status,
        "description": description,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }

    candidate_methods = ("insert", "create", "save", "log_import")
    for method_name in candidate_methods:
        method = getattr(repo_import_log, method_name, None)
        if callable(method):
            try:
                method(payload)
                return
            except Exception:
                continue


# ============================================================================
# REPOSITORY RESOLUTION HELPERS
# ============================================================================

def _resolve_customer_id(client_name: Any, customer_cache: Dict[str, Optional[int]]) -> Optional[int]:
    """
    Resolve client_id via CompanyListNameRepository com cache.

    Implementação alinhada ao repository real do projeto,
    utilizando explicitamente o método:
        get_company_id_by_name()
    """
    import unicodedata

    normalized = _normalize_identifier(client_name)

    # ---------------------------------------------------------
    # Banco usa COLLATE utf8_bin (case + acento sensitive)
    # Precisamos:
    # 1) remover acentos
    # 2) padronizar para UPPER
    # ---------------------------------------------------------
    if normalized:
        no_accent = unicodedata.normalize("NFKD", normalized)
        no_accent = "".join(c for c in no_accent if not unicodedata.combining(c))
        normalized_db = no_accent.upper()
    else:
        normalized_db = normalized

    if normalized_db in customer_cache:
        return customer_cache[normalized_db]

    try:
        customer_id = repo_company.get_company_id_by_name(normalized_db)
    except Exception:
        customer_id = None

    customer_cache[normalized_db] = customer_id
    return customer_id


def _resolve_product_id(license_name: Any, product_cache: Dict[str, Optional[int]]) -> Optional[int]:
    """
    Resolve product_id via ProductRepository com cache.

    Regras específicas do projeto:
    - product_vendor_id = 1
    - product_name = nome vindo do Excel
    - product_part_number = mesmo valor do nome
    - Se não existir → criar automaticamente
    """
    normalized = _normalize_identifier(license_name)

    if not normalized:
        return None

    if normalized in product_cache:
        return product_cache[normalized]

    vendor_id = 1

    # ---------------------------------------------------------
    # 1) Buscar produto existente por vendor + name
    # ---------------------------------------------------------
    try:
        existing_ids = repo_product.find_ids_by({
            "product_vendor_id": vendor_id,
            "product_name": normalized
        })
    except Exception:
        existing_ids = []

    if existing_ids:
        product_id = _safe_repo_int(existing_ids[0])
        product_cache[normalized] = product_id
        return product_id

    # ---------------------------------------------------------
    # 2) Se não existir → criar novo produto
    # ---------------------------------------------------------
    try:
        product_id = repo_product.insert({
            "product_vendor_id": vendor_id,
            "product_name": normalized,
            "product_part_number": normalized
        })
    except Exception:
        product_id = None

    product_cache[normalized] = product_id
    return product_id


# ============================================================================
# GROUPING / PAYLOAD BUILDERS
# ============================================================================

def _build_quantity_group_key(row: Dict[str, Any], customer_id: int) -> Tuple[Any, ...]:
    """
    Chave lógica de consolidação de quantity.

    Importante:
    - datas fazem parte da chave
    - Active NÃO faz parte da chave
    - Compliance NÃO faz parte da chave
    """
    start_date = _to_date(row.get("Start Date"))
    end_date = _to_date(row.get("End Date"))

    return (
        customer_id,
        _normalize_identifier(row.get("Domain")),
        _normalize_identifier(row.get("License")),
        _normalize_identifier(row.get("Virtual Account")),
        _normalize_str(row.get("Billing")),
        _normalize_str(row.get("License Type")),
        _normalize_str(row.get("Subscription Id")),
        _date_to_db(start_date),
        _date_to_db(end_date),
    )


def _aggregate_quantity_groups(
    source_rows: List[Tuple[int, Dict[str, Any]]],
    customer_cache: Dict[str, Optional[int]],
    product_cache: Dict[str, Optional[int]],
    log_path: Path,
    result: ProcessingResult
) -> Tuple[List[QuantityGroup], List[Dict[str, Any]]]:
    """
    Consolida globalmente as linhas quantity do arquivo inteiro.

    Por que globalmente?
    - porque a regra de negócio exige soma do arquivo inteiro
    - chunk com soma parcial pode gerar resultado incorreto

    Estratégia:
    - percorre as linhas uma única vez
    - valida cliente e produto cedo
    - forma grupos em memória enxuta
    - coleta falhas imediatamente para grupos inviáveis

    Retorna:
    - lista de grupos válidos
    - lista de falhas já detectadas
    """
    grouped: Dict[Tuple[Any, ...], QuantityGroup] = {}
    failed_rows: List[Dict[str, Any]] = []

    for excel_row_number, row in source_rows:
        if not _has_meaningful_quantity(row):
            continue

        client_name = row.get("Client")
        license_name = row.get("License")

        customer_id = _resolve_customer_id(client_name, customer_cache)
        product_id = _resolve_product_id(license_name, product_cache)

        if not customer_id:
            failed_rows.append(_build_failed_row(
                row=row,
                excel_row_number=excel_row_number,
                error_message=f"Cliente não encontrado: {client_name}"
            ))
            continue

        if not product_id:
            failed_rows.append(_build_failed_row(
                row=row,
                excel_row_number=excel_row_number,
                error_message=f"Produto não encontrado para License: {license_name}"
            ))
            continue

        key = _build_quantity_group_key(row, customer_id)
        quantity_value = _to_decimal(row.get("Quantity"), default=Decimal("0"))

        if key not in grouped:
            grouped[key] = QuantityGroup(
                group_key=key,
                representative_row=row,
                original_rows=[(excel_row_number, row)],
                total_quantity=quantity_value
            )
        else:
            grouped[key].original_rows.append((excel_row_number, row))
            grouped[key].total_quantity += quantity_value

    _safe_log(
        log_path,
        result,
        f"Consolidação quantity concluída: {len(grouped)} grupos válidos e {len(failed_rows)} falhas preliminares."
    )
    return list(grouped.values()), failed_rows


def _build_quantity_payload_from_group(
    group: QuantityGroup,
    customer_id: int,
    product_id: int
) -> Dict[str, Any]:
    row = group.representative_row

    return {
        "mcsa_row_type": "quantity",
        "mcsa_client_id": customer_id,
        "mcsa_client": _normalize_identifier(row.get("Client")),
        "mcsa_domain": _normalize_identifier(row.get("Domain")),
        "mcsa_product_id": product_id,
        "mcsa_license": _normalize_identifier(row.get("License")),
        "mcsa_virtual_account": _normalize_identifier(row.get("Virtual Account")),
        "mcsa_billing": _normalize_str(row.get("Billing")),
        "mcsa_available_to_use": None,
        "mcsa_in_use": None,
        "mcsa_balance": None,
        "mcsa_compliance": _normalize_str(row.get("Compliance")),
        "mcsa_license_type": _normalize_str(row.get("License Type")),
        "mcsa_quantity": int(group.total_quantity),
        "mcsa_subscription": _normalize_str(row.get("Subscription Id")),
        "mcsa_active": _normalize_str(row.get("Active")),
        "mcsa_days_to_end": _normalize_nullable_str(row.get("Days To End")),
        "mcsa_start_date": _date_to_db(_to_date(row.get("Start Date"))),
        "mcsa_end_date": _date_to_db(_to_date(row.get("End Date"))),
        "mcsa_track": 0,
        "mcsa_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


def _build_metering_payload_from_row(
    row: Dict[str, Any],
    customer_id: int,
    product_id: int
) -> Dict[str, Any]:
    return {
        "mcsa_row_type": "metering",
        "mcsa_client_id": customer_id,
        "mcsa_client": _normalize_identifier(row.get("Client")),
        "mcsa_domain": _normalize_identifier(row.get("Domain")),
        "mcsa_product_id": product_id,
        "mcsa_license": _normalize_identifier(row.get("License")),
        "mcsa_virtual_account": _normalize_identifier(row.get("Virtual Account")),
        "mcsa_billing": _normalize_str(row.get("Billing")),
        "mcsa_available_to_use": _normalize_nullable_str(row.get("Available To Use")),
        "mcsa_in_use": _normalize_nullable_str(row.get("In Use")),
        "mcsa_balance": _normalize_nullable_str(row.get("Balance")),
        "mcsa_compliance": _normalize_str(row.get("Compliance")),
        "mcsa_license_type": _normalize_str(row.get("License Type")),
        "mcsa_quantity": None,
        "mcsa_subscription": _normalize_str(row.get("Subscription Id")),
        "mcsa_active": _normalize_str(row.get("Active")),
        "mcsa_days_to_end": _normalize_nullable_str(row.get("Days To End")),
        "mcsa_start_date": _date_to_db(_to_date(row.get("Start Date"))),
        "mcsa_end_date": _date_to_db(_to_date(row.get("End Date"))),
        "mcsa_track": 0,
        "mcsa_update": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }


# ============================================================================
# DB MATCH / UPSERT HELPERS
# ============================================================================

def _find_existing_quantity(payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    where = {
        "mcsa_row_type": "quantity",
        "mcsa_client_id": payload["mcsa_client_id"],
        "mcsa_domain": payload["mcsa_domain"],
        "mcsa_product_id": payload["mcsa_product_id"],
        "mcsa_license": payload["mcsa_license"],
        "mcsa_virtual_account": payload["mcsa_virtual_account"],
        "mcsa_billing": payload["mcsa_billing"],
        "mcsa_license_type": payload["mcsa_license_type"],
        "mcsa_subscription": payload["mcsa_subscription"],
        "mcsa_start_date": payload["mcsa_start_date"],
        "mcsa_end_date": payload["mcsa_end_date"],
    }
    return repo_mcsa.find_first_by(where)


def _find_existing_metering(payload: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    where = {
        "mcsa_row_type": "metering",
        "mcsa_client_id": payload["mcsa_client_id"],
        "mcsa_domain": payload["mcsa_domain"],
        "mcsa_product_id": payload["mcsa_product_id"],
        "mcsa_license": payload["mcsa_license"],
        "mcsa_virtual_account": payload["mcsa_virtual_account"],
        "mcsa_subscription": payload["mcsa_subscription"],
        "mcsa_start_date": payload["mcsa_start_date"],
        "mcsa_end_date": payload["mcsa_end_date"],
    }
    return repo_mcsa.find_first_by(where)


def _upsert_quantity_group(
    group: QuantityGroup,
    customer_cache: Dict[str, Optional[int]],
    product_cache: Dict[str, Optional[int]]
) -> None:
    row = group.representative_row

    customer_id = _resolve_customer_id(row.get("Client"), customer_cache)
    if not customer_id:
        raise ValueError(f"Cliente não encontrado: {row.get('Client')}")

    product_id = _resolve_product_id(row.get("License"), product_cache)
    if not product_id:
        raise ValueError(f"Produto não encontrado para License: {row.get('License')}")

    payload = _build_quantity_payload_from_group(group, customer_id, product_id)
    existing = _find_existing_quantity(payload)

    if not existing:
        repo_mcsa.insert(payload)
        return

    current_quantity = _safe_repo_int(existing.get("mcsa_quantity")) or 0
    new_quantity = _safe_repo_int(payload.get("mcsa_quantity")) or 0

    update_data = {
        "mcsa_compliance": payload["mcsa_compliance"],
        "mcsa_quantity": payload["mcsa_quantity"],
        "mcsa_active": payload["mcsa_active"],
        "mcsa_days_to_end": payload["mcsa_days_to_end"],
        "mcsa_update": payload["mcsa_update"],
    }

    if current_quantity != new_quantity:
        update_data["mcsa_track"] = 0

    repo_mcsa.update(
        data=update_data,
        where={"mcsa_id": existing["mcsa_id"]}
    )


def _upsert_metering_record(
    row: Dict[str, Any],
    customer_cache: Dict[str, Optional[int]],
    product_cache: Dict[str, Optional[int]]
) -> None:
    customer_id = _resolve_customer_id(row.get("Client"), customer_cache)
    if not customer_id:
        raise ValueError(f"Cliente não encontrado: {row.get('Client')}")

    product_id = _resolve_product_id(row.get("License"), product_cache)
    if not product_id:
        raise ValueError(f"Produto não encontrado para License: {row.get('License')}")

    payload = _build_metering_payload_from_row(row, customer_id, product_id)
    existing = _find_existing_metering(payload)

    if not existing:
        repo_mcsa.insert(payload)
        return

    update_data = {
        "mcsa_available_to_use": payload["mcsa_available_to_use"],
        "mcsa_in_use": payload["mcsa_in_use"],
        "mcsa_balance": payload["mcsa_balance"],
        "mcsa_compliance": payload["mcsa_compliance"],
        "mcsa_active": payload["mcsa_active"],
        "mcsa_days_to_end": payload["mcsa_days_to_end"],
        "mcsa_update": payload["mcsa_update"],
    }

    repo_mcsa.update(
        data=update_data,
        where={"mcsa_id": existing["mcsa_id"]}
    )


# ============================================================================
# FAILURE HELPERS
# ============================================================================

def _build_failed_row(
    row: Dict[str, Any],
    excel_row_number: int,
    error_message: str
) -> Dict[str, Any]:
    failed = dict(row)
    failed["import_error_message"] = error_message
    failed["import_error_column"] = None
    failed["import_error_value"] = None
    failed["import_original_row"] = excel_row_number
    failed["import_processed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return failed


# ============================================================================
# MAIN FILE PROCESSING
# ============================================================================

def _process_file(source_file: Path) -> None:
    _ensure_directories()

    lock_path = _build_lock_path(source_file)
    log_path = _build_execution_log_path(source_file)

    _acquire_lock(lock_path)

    try:
        result = ProcessingResult()
        _safe_log(log_path, result, f"Iniciando processamento de {source_file.name}")

        headers, source_rows = _read_source_rows(source_file)

        if not headers:
            _safe_log(log_path, result, "Arquivo sem cabeçalho ou vazio. Nada a processar.")
            _register_import_log(source_file, "warning", "Arquivo vazio ou sem cabeçalho.")
            return

        if len(source_rows) > MAX_ROWS_SOFT_WARNING:
            _safe_log(
                log_path,
                result,
                f"Aviso: arquivo com {len(source_rows)} linhas de dados. Processamento seguirá com proteção operacional."
            )

        original_data_by_row = {excel_row_number: row for excel_row_number, row in source_rows}

        customer_cache: Dict[str, Optional[int]] = {}
        product_cache: Dict[str, Optional[int]] = {}

        # ------------------------------------------------------------------
        # 1) Consolidação global de quantity
        # ------------------------------------------------------------------
        quantity_groups, quantity_pre_failures = _aggregate_quantity_groups(
            source_rows=source_rows,
            customer_cache=customer_cache,
            product_cache=product_cache,
            log_path=log_path,
            result=result
        )

        # Linhas quantity que falharam já devem sair do original
        for failure in quantity_pre_failures:
            row_number = failure.get("import_original_row")
            if row_number:
                result.failed_rows.append(failure)

        # ------------------------------------------------------------------
        # 2) Persistência de quantity consolidado
        # ------------------------------------------------------------------
        quantity_processed_rows = set()

        for group in quantity_groups:
            try:
                _upsert_quantity_group(group, customer_cache, product_cache)

                for excel_row_number, _ in group.original_rows:
                    quantity_processed_rows.add(excel_row_number)
                    result.success_row_numbers.add(excel_row_number)

            except Exception as e:
                error_message = f"Falha ao processar grupo quantity: {str(e)}"

                for excel_row_number, row in group.original_rows:
                    result.failed_rows.append(
                        _build_failed_row(
                            row=row,
                            excel_row_number=excel_row_number,
                            error_message=error_message
                        )
                    )

                _safe_log(log_path, result, error_message)

        _safe_log(
            log_path,
            result,
            f"Processamento quantity concluído. Linhas marcadas como sucesso: {len(quantity_processed_rows)}"
        )

        # ------------------------------------------------------------------
        # 3) Persistência de metering
        # ------------------------------------------------------------------
        metering_success = 0
        metering_failure = 0

        for excel_row_number, row in source_rows:
            if not _has_meaningful_metering(row):
                continue

            try:
                _upsert_metering_record(row, customer_cache, product_cache)
                result.success_row_numbers.add(excel_row_number)
                metering_success += 1

            except Exception as e:
                result.failed_rows.append(
                    _build_failed_row(
                        row=row,
                        excel_row_number=excel_row_number,
                        error_message=f"Falha ao processar metering: {str(e)}"
                    )
                )
                metering_failure += 1

        _safe_log(
            log_path,
            result,
            f"Processamento metering concluído. Sucessos: {metering_success}, falhas: {metering_failure}"
        )

        # ------------------------------------------------------------------
        # 4) Determinar linhas lidas
        # ------------------------------------------------------------------
        failed_row_numbers = {
            row.get("import_original_row")
            for row in result.failed_rows
            if row.get("import_original_row") is not None
        }

        read_row_numbers = set(result.success_row_numbers) | failed_row_numbers

        remaining_rows = [
            row
            for excel_row_number, row in source_rows
            if excel_row_number not in read_row_numbers
        ]

        # ------------------------------------------------------------------
        # 5) Persistir arquivo de falhas
        # ------------------------------------------------------------------
        failed_headers = headers + [
            "import_error_message",
            "import_error_column",
            "import_error_value",
            "import_original_row",
            "import_processed_at",
        ]

        failed_file = _build_failed_file_path(source_file)
        _append_failed_rows(failed_file, failed_headers, result.failed_rows)

        # ------------------------------------------------------------------
        # 6) Regravar arquivo original só com não lidas
        # ------------------------------------------------------------------
        _write_remaining_source_rows(
            source_file=source_file,
            headers=headers,
            remaining_rows=remaining_rows
        )

        # ------------------------------------------------------------------
        # 7) Logs finais
        # ------------------------------------------------------------------
        total_failures = len(result.failed_rows)
        total_success = len(result.success_row_numbers)

        summary = (
            f"Arquivo processado. "
            f"Total linhas origem={len(source_rows)} | "
            f"lidas={len(read_row_numbers)} | "
            f"sucesso={total_success} | "
            f"falhas={total_failures} | "
            f"remanescente={len(remaining_rows)}"
        )

        _safe_log(log_path, result, summary)

        if total_failures > 0 and total_success == 0:
            status = "error"
        elif total_failures > 0:
            status = "warning"
        else:
            status = "success"

        _register_import_log(source_file, status, summary)

    except Exception as e:
        error_message = f"Falha fatal no processamento de {source_file.name}: {str(e)}"
        _append_execution_log(log_path, error_message)
        _append_execution_log(log_path, traceback.format_exc())
        _register_import_log(source_file, "error", error_message)
        raise

    finally:
        _release_lock(lock_path)


# ============================================================================
# PUBLIC ENTRYPOINT
# ============================================================================

def run_import(file_name: Optional[str] = None, user_id: Optional[str] = None) -> Dict[str, str]:
    """
    Executa a importação de um arquivo específico ou de todos os .xlsx do input.

    Compatível com o scheduler em src/services/import_scheduling.py,
    que chama todas as funções registradas em IMPORT_DISPATCHER com a
    assinatura:
        run_import(file_name: str, user_id: Optional[str]) -> Dict[str, str]

    Parâmetros:
    - file_name: nome exato do arquivo dentro de .storage/input
      Se None, processa todos os arquivos .xlsx encontrados.
    - user_id: identificador opcional do usuário que disparou a importação.
      Não é utilizado aqui, mas é mantido para compatibilidade de interface.

    Regras operacionais:
    - cada arquivo recebe lock individual
    - falhas são acumuladas em arquivo fixo por origem
    - toda linha lida sai do arquivo original

    Retorno:
    - dicionário com pelo menos:
        {"status": "...", "message": "..."}
    """
    _ensure_directories()

    if file_name:
        source_files = [BASE_INPUT_PATH / file_name]
    else:
        source_files = sorted(BASE_INPUT_PATH.glob("*.xlsx"))

    processed = 0
    last_summary_message = None

    for source_file in source_files:
        if not source_file.exists():
            continue

        _process_file(source_file)

        # Após processamento, tentar ler última linha do log
        log_path = BASE_LOGS_PATH / f"{source_file.stem}.log"
        if log_path.exists():
            try:
                with open(log_path, "r", encoding="utf-8") as f:
                    lines = f.readlines()
                    if lines:
                        raw_line = lines[-1].strip()
                        # Remove prefixo de timestamp: "[YYYY-MM-DD HH:MM:SS] "
                        if raw_line.startswith("[") and "]" in raw_line:
                            last_summary_message = raw_line.split("] ", 1)[1]
                        else:
                            last_summary_message = raw_line
            except Exception:
                pass

        processed += 1

    if processed == 0:
        return {
            "status": "WARNING",
            "message": "Nenhum arquivo encontrado para processar."
        }

    return {
        "status": "FINISHED",
        "message": last_summary_message or f"Arquivos processados: {processed}."
    }


if __name__ == "__main__":
    run_import()
