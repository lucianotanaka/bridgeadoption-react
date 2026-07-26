import logging
from pathlib import Path
from dataclasses import dataclass
from typing import Optional, List, Dict, Tuple, Any
from datetime import datetime, timedelta

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.database.repositories.cisco_ea_repository import CiscoEARepository
from src.infrastructure.database.repositories.product_repository import ProductRepository
from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
from src.infrastructure.database.repositories.task_repository import TaskRepository
from src.infrastructure.database.repositories.task_activity_repository import TaskActivityRepository
from src.infrastructure.database.repositories.task_history_repository import TaskHistoryRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository
from src.infrastructure.database.repositories.cisco_web_order_repository import CiscoWebOrderRepository
from src.infrastructure.database.repositories.subscription_ignored_repository import SubscriptionIgnoredRepository
from src.services.matching.company_name_matching import generate_and_store_suggestions
from src.domain.import_schemas import IMPORT_SCHEMAS

BASE_STORAGE_PATH = Path("/home/bridgeadoption/storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"

IMPORT_SOURCE = "CiscoSubscriptionCCW"
ERROR_EXTRA_COLUMNS = [
    "import_error_message",
    "import_error_column",
    "import_error_value",
    "import_original_row",
    "import_processed_at",
    "import_source",
    "import_original_file",
]

logger = logging.getLogger(__name__)

repo_ea = CiscoEARepository()
repo_prod = ProductRepository()
repo_log = ImportLogRepository()
repo_task = TaskRepository()
repo_activity = TaskActivityRepository()
repo_history = TaskHistoryRepository()
repo_company = CompanyListNameRepository()
repo_weborder = CiscoWebOrderRepository()
repo_subscription_ignored = SubscriptionIgnoredRepository()


@dataclass
class ImportSummary:
    file_path: str
    total_rows: int
    imported_rows: int
    error_rows: int


@dataclass
class RowProcessResult:
    success: bool
    error_message: Optional[str] = None
    error_column: Optional[str] = None
    error_value: Optional[Any] = None


def _ensure_directories():
    """
    Garante que os diretórios necessários existam.
    """
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)


def _build_execution_log_path(input_path: Path) -> Path:
    """
    Monta o caminho do arquivo de log da execução.
    Agora segue o padrão: nome_do_arquivo_origem.log
    """
    return BASE_LOGS_PATH / f"{input_path.stem}.log"


def _append_execution_log(log_path: Path, message: str):
    """
    Escreve uma linha no log texto da execução.
    """
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def _safe_log(
    file_path: str,
    row_number: int,
    message: str,
    column_name: Optional[str] = None,
    value: Optional[Any] = None,
):
    """
    Grava log de importação com proteção contra falha no próprio logging.
    """
    try:
        repo_log.create(
            IMPORT_SOURCE,
            file_path,
            row_number,
            message,
            column_name,
            value,
        )
    except Exception as log_ex:
        logger.exception(
            "Falha ao gravar tbImportLog. file=%s row=%s column=%s value=%s error=%s",
            file_path,
            row_number,
            column_name,
            value,
            log_ex,
        )


def _safe_int(value: Any, default: int = 0, max_digits: int = 30) -> int:
    """
    Converte valor para int com validação defensiva para evitar:
    - strings gigantes
    - formatos inválidos
    - valores não numéricos
    """
    if value is None:
        return default

    if isinstance(value, bool):
        return default

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return default

    s = str(value).strip()
    if not s:
        return default

    if s.endswith(".0"):
        s = s[:-2]

    signless = s.lstrip("+-")
    if len(signless) > max_digits:
        return default

    if not signless.isdigit():
        return default

    try:
        return int(s)
    except Exception:
        return default


def _safe_repo_int(
    value: Any,
    context: str,
    file_path: str,
    row_number: int,
    execution_log_path: Optional[Path] = None,
) -> int:
    """
    Converte retorno de repository para int de forma segura e loga quando inválido.
    """
    result = _safe_int(value, default=0)
    if value is not None and result == 0:
        _safe_log(
            file_path=file_path,
            row_number=row_number,
            message=f"Valor inválido ao converter retorno para inteiro em {context}",
            column_name=context,
            value=str(value)[:500],
        )
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"WARN invalid integer conversion in context={context}, row={row_number}, value={str(value)[:500]}",
            )
    return result


def _normalize_str(value) -> Optional[str]:
    """
    Normaliza qualquer valor para string stripada.
    Colapsa múltiplos espaços internos em um único espaço.
    Retorna None para valores vazios ou NaN.
    """
    import pandas as pd

    if pd.isna(value):
        return None
    s = " ".join(str(value).split())
    return s if s else None


def _normalize_identifier(value: Any, max_length: int = 255) -> Optional[str]:
    """
    Normaliza identificadores de negócio como string.
    Não converte para inteiro.
    Protege contra conteúdo gigante.
    """
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    if len(text) > max_length:
        return text[:max_length]

    return text


def _to_float(value) -> Optional[float]:
    """
    Converte números para float Python, removendo caracteres não numéricos,
    exceto o ponto decimal e o sinal.
    """
    import pandas as pd
    import re

    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None

    s = str(value).strip()
    if not s:
        return None

    cleaned = re.sub(r"[^0-9\.\-+]", "", s)
    if not cleaned or cleaned in {".", "-", "+", "-.", "+."}:
        return None

    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_date(value) -> Optional[datetime]:
    """
    Converte diferentes formatos de datas em datetime.
    """
    from datetime import datetime as dt, date
    import pandas as pd

    if value is None or (isinstance(value, float) and pd.isna(value)) or value == "":
        return None
    if isinstance(value, dt):
        return value
    if isinstance(value, date):
        return dt(value.year, value.month, value.day)

    s = str(value).strip()[:50]
    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
    ):
        try:
            return dt.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _normalize_weborder(value: Optional[str]) -> Optional[str]:
    """
    Normaliza a coluna WebOrderID garantindo um identificador válido.
    """
    text = _normalize_identifier(value, max_length=255)
    if text is None:
        return None

    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if text.lower() in invalid_markers:
        return None

    if len(text) < 3:
        return None

    return text


def _get_customer_id_from_end_customer(
    value: Optional[str],
    file_path: str = "",
    row_number: int = 0,
    execution_log_path: Optional[Path] = None,
) -> int:
    """
    Recupera o ID do cliente a partir do texto do End Customer.
    """
    if value is None:
        return 0

    # Colapsa múltiplos espaços internos antes de buscar no banco.
    # Exemplo: "ANBIMA  ASSOCIACAO" → "ANBIMA ASSOCIACAO"
    name = " ".join(str(value).split())
    if not name:
        return 0

    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if name.lower() in invalid_markers:
        return 0

    if len(name) < 2:
        return 0

    try:
        company_id = repo_company.get_company_id_by_name(name)
    except Exception:
        generate_and_store_suggestions([name], user="from Cisco Subscription CCW")
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"WARN customer suggestion generated for name={name}, row={row_number}",
            )
        return 0

    if not company_id:
        generate_and_store_suggestions([name], user="from Cisco Subscription CCW")
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"WARN customer not found, suggestion generated for name={name}, row={row_number}",
            )
        return 0

    return _safe_repo_int(company_id, "company_id", file_path, row_number, execution_log_path)


def _get_or_create_product_id_from_offer_name(
    value: Optional[str],
    file_path: str = "",
    row_number: int = 0,
    execution_log_path: Optional[Path] = None,
) -> int:
    """
    Busca ou cria um produto Cisco (vendor_id=1) com base na coluna Offer Name.
    """
    if value is None:
        return 0

    name = str(value).strip()
    if not name:
        return 0

    invalid_markers = {"n/a", "na", "none", "null", "-", "--", "0"}
    if name.lower() in invalid_markers:
        return 0

    if len(name) < 2:
        return 0

    CISCO_VENDOR_ID = 1

    try:
        product_ids = repo_prod.find_ids_by_name_or_partnumber(
            name_or_part=name,
            vendor_id=CISCO_VENDOR_ID,
        )
    except AttributeError:
        where_fallback = {"product_vendor_id": CISCO_VENDOR_ID}
        where_fallback["product_name"] = name
        try:
            product_ids = repo_prod.find_ids_by(where_fallback)
        except Exception:
            product_ids = []

        if not product_ids:
            where_fallback.pop("product_name", None)
            where_fallback["product_part_number"] = name
            try:
                product_ids = repo_prod.find_ids_by(where_fallback)
            except Exception:
                product_ids = []
    except Exception:
        product_ids = []

    if product_ids:
        product_id = _safe_repo_int(product_ids[0], "product_id", file_path, row_number, execution_log_path)
        if product_id:
            return product_id

    data = {
        "product_vendor_id": CISCO_VENDOR_ID,
        "product_name": name,
        "product_part_number": name,
    }

    try:
        new_id = repo_prod.insert(data)
        new_product_id = _safe_repo_int(new_id, "product_insert_id", file_path, row_number, execution_log_path)
        if new_product_id and execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO product created for offer_name={name}, row={row_number}, product_id={new_product_id}",
            )
        return new_product_id
    except Exception:
        return 0


def _split_to_list(value: Optional[str]) -> List[str]:
    """
    Divide uma string com vírgulas em lista de valores stripados.
    """
    if not value:
        return []

    parts = [p.strip() for p in str(value).split(",")]
    return [p for p in parts if p]


def _build_ea_payload_from_row(
    row: Dict[str, Any],
    customer_id: int,
    product_id: int,
    weborder_str: str,
) -> Dict[str, Any]:
    """
    Constrói o payload final para tbCiscoEA.
    """
    norm = _normalize_str
    ident = _normalize_identifier
    to_float = _to_float
    to_dt = _to_date

    return {
        "ea_end_customer_id": customer_id,
        "ea_product_id": product_id,
        "ea_web_order_id": ident(weborder_str, max_length=255),
        "ea_consumption_status": norm(row.get("Consumption Status")),
        "ea_over_consumed_tf_groups": norm(row.get("Over Consumed TF Groups")),
        "ea_tf_groups": norm(row.get("TF Groups")),
        "ea_tf_effective_date": to_dt(row.get("True Forward Effective Date")),
        "ea_next_tf": to_dt(row.get("Next True Forward")),
        "ea_subscription_id": ident(row.get("Subscription ID"), max_length=255),
        "ea_ccw_line_status": norm(row.get("Status")),
        "ea_start_date": to_dt(row.get("Start Date")),
        "ea_end_date": to_dt(row.get("End Date")),
        "ea_inicial_term": to_float(row.get("Initial Term")),
        "ea_renewal_date": to_dt(row.get("Renewal Date")),
        "ea_currency": norm(row.get("Currency")),
        "ea_mrc": to_float(row.get("Monthly Charge")),
        "ea_tf_overage": to_float(row.get("TF Overage")),
        "ea_po": ident(row.get("Purchase Order Number"), max_length=255),
        "ea_buying_program_id": ident(row.get("Buying Program ID"), max_length=255),
        "ea_site_url": norm(row.get("Site URL")),
        "ea_customer_success_manager": norm(row.get("Customer Success Manager")),
        "ea_sales_specialist": norm(row.get("Sales Specialist")),
        "ea_customer_success_manager_email": norm(row.get("Customer Success Manager Email")),
        "ea_sales_specialist_email": norm(row.get("Sales Specialist Email")),
        "ea_primary_billing_contact_name": norm(row.get("Primary Billing Contact Name")),
        "ea_primary_billing_contact_email": norm(row.get("Primary Billing Contact Email")),
        "ea_service_contact_name": norm(row.get("Service To Contact Name")),
        "ea_service_contact_email": norm(row.get("Service To Contact Email")),
        "ea_end_customer_contact_name": norm(row.get("End Customer Contact Name")),
        "ea_end_customer_contact_email": norm(row.get("End Customer Contact Email")),
        "ea_end_customer_contact_phone": ident(row.get("End Customer Contact Phone"), max_length=100),
        "ea_order_submit_date": to_dt(row.get("Order Submitted Date")),
        "ea_smart_account_name": norm(row.get("Smart Account Name")),
        "ea_renewal_manager": norm(row.get("Renewal Manager")),
        "ea_renewal_manager_email": norm(row.get("Renewal Manager Email")),
        "ea_provisioning_status": norm(row.get("Provisioning Status")),
    }


def _resolve_input_path(file_path: str) -> Path:
    """
    Resolve caminho do arquivo de entrada.
    """
    path = Path(file_path)
    if path.is_absolute():
        return path
    return BASE_INPUT_PATH / path.name


def _build_failed_output_path(input_path: Path) -> Path:
    """
    Monta o caminho fixo do arquivo de falhas baseado no nome do original.
    Agora segue o padrão: nome_do_arquivo_failed_rows.xlsx
    """
    return BASE_OUTPUT_PATH / f"{input_path.stem}_failed_rows.xlsx"


def _open_workbook_rw(file_path: str) -> Tuple[Any, Worksheet, Dict[str, int], List[str]]:
    """
    Abre o arquivo Excel em modo leitura/escrita e monta o índice das colunas.
    """
    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h.strip() if h is not None else "" for h in header_row]
    idx_map = {h: i for i, h in enumerate(headers)}
    return wb, ws, idx_map, headers


def _ensure_failed_workbook(
    failed_path: Path,
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
):
    """
    Cria (ou recria) o arquivo XLSX de falhas com apenas o cabeçalho.

    Sempre sobrescreve o arquivo se já existir, garantindo que não haja
    resíduos de execuções anteriores. A criação só será feita de fato
    se esta função for chamada (ou seja, somente quando houver falhas
    a serem escritas).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "failed_rows"
    ws.append(original_headers + ERROR_EXTRA_COLUMNS)
    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO failed file created (overwritten if existed): {str(failed_path)}",
        )


def _append_failed_rows(
    failed_path: Path,
    rows_to_append: List[List[Any]],
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
):
    """
    Faz append das linhas de falha no arquivo XLSX fixo.

    Regras:
    - Se rows_to_append estiver vazio, não faz nada e não cria arquivo.
    - Se o arquivo ainda não existir, ele é criado/recriado com header
      (sobrescrevendo conteúdo anterior) via _ensure_failed_workbook.
    """
    if not rows_to_append:
        return

    _ensure_failed_workbook(failed_path, original_headers, execution_log_path)

    wb = load_workbook(str(failed_path))
    ws = wb.active

    for row in rows_to_append:
        ws.append(row)

    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO appended failed rows count={len(rows_to_append)} to file={str(failed_path)}",
        )


def _get_excel_validation_error_info(
    message: str,
    row_dict: Dict[str, Any],
    offer_name: Optional[str],
    weborder_str: Optional[str],
) -> Tuple[Optional[str], Optional[Any]]:
    """
    Mapeia erro técnico para coluna da planilha.
    """
    msg_err = str(message)
    log_column = "Offer Name"
    log_value = offer_name if offer_name is not None else weborder_str

    mappings = {
        "ea_primary_billing_contact_email": "Primary Billing Contact Email",
        "ea_service_contact_name": "Service To Contact Name",
        "ea_over_consumed_tf_groups": "Over Consumed TF Groups",
        "ea_tf_groups": "TF Groups",
        "ea_smart_account_name": "Smart Account Name",
        "ea_end_customer_contact_name": "End Customer Contact Name",
        "ea_end_customer_contact_email": "End Customer Contact Email",
        "ea_service_contact_email": "Service To Contact Email",
        "ea_end_customer_contact_phone": "End Customer Contact Phone",
        "ea_customer_success_manager_email": "Customer Success Manager Email",
        "ea_sales_specialist_email": "Sales Specialist Email",
        "ea_subscription_id": "Subscription ID",
        "ea_web_order_id": "WebOrderID",
        "ea_po": "Purchase Order Number",
    }

    for key_db, key_excel in mappings.items():
        if key_db in msg_err:
            return key_excel, row_dict.get(key_excel)

    return log_column, log_value


def _row_dict_to_failed_excel_row(
    headers: List[str],
    row_dict: Dict[str, Any],
    original_row_number: int,
    input_file_path: str,
    result: RowProcessResult,
) -> List[Any]:
    """
    Converte uma linha original + dados do erro para a estrutura do XLSX de falhas.
    """
    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    original_values = [row_dict.get(h) for h in headers]

    return original_values + [
        result.error_message,
        result.error_column,
        str(result.error_value)[:2000] if result.error_value is not None else None,
        original_row_number,
        processed_at,
        IMPORT_SOURCE,
        input_file_path,
    ]


def _handle_over_consumed(
    payload,
    row_dict,
    file_path,
    row_number,
    offer_name,
    weborder_str,
    ea_subscription_id,
    task_customer_id,
    execution_log_path: Optional[Path] = None,
):
    """
    Implementa o PASSO 7 para casos de Over Consumed.
    """
    now = datetime.now()

    def build_non_null_ref(*args) -> Optional[str]:
        parts = [str(x).strip() for x in args if x and str(x).strip()]
        return ", ".join(parts) if parts else None

    task_references = [
        str(row_dict.get("Next True Forward")).strip() if row_dict.get("Next True Forward") else None,
        ea_subscription_id,
        row_dict.get("Purchase Order Number"),
        weborder_str,
    ]

    offer_name_val = (offer_name or "").strip().upper()
    over_consumed_tf_groups = payload.get("ea_over_consumed_tf_groups")
    tf_groups = payload.get("ea_tf_groups")

    track_keywords = [offer_name_val]
    if over_consumed_tf_groups:
        track_keywords += _split_to_list(over_consumed_tf_groups)

    subtrack_keywords = _split_to_list(tf_groups)

    found_task_info = None

    for ref in task_references:
        if not ref:
            continue
        for track_kw in track_keywords:
            if not track_kw:
                continue
            for subtrack_kw in subtrack_keywords or [None]:
                found_task_info = repo_ea.find_task_over_consumed(
                    task_customer_id=task_customer_id,
                    task_ws=ea_subscription_id,
                    task_reference=ref,
                    track_keyword=track_kw,
                    subtrack_keyword=subtrack_kw,
                )
                if found_task_info:
                    break
            if found_task_info:
                break
        if found_task_info:
            break

    def parse_date_str(value):
        if isinstance(value, datetime):
            return value.date()
        elif isinstance(value, str):
            try:
                return datetime.strptime(value.strip(), "%Y-%m-%d").date()
            except Exception:
                return None
        return None

    true_forward_date = parse_date_str(row_dict.get("True Forward Effective Date"))
    next_true_forward_date = parse_date_str(row_dict.get("Next True Forward"))

    task_start_date = now.date()
    if true_forward_date and true_forward_date > task_start_date:
        task_end_date = true_forward_date - timedelta(days=7)
    elif next_true_forward_date:
        task_end_date = next_true_forward_date - timedelta(days=14)
    else:
        task_end_date = task_start_date + timedelta(days=30)

    monthly_charge = payload.get("ea_mrc")
    try:
        task_value = float(monthly_charge) if monthly_charge else 0.0
    except Exception:
        task_value = 0.0

    currency = payload.get("ea_currency") or "USD"

    if not found_task_info:
        try:
            owner_id = repo_task.get_last_task_owner_by_company(
                company_id=task_customer_id,
                user_type="CSM",
            )
        except Exception:
            owner_id = 0

        owner_id = _safe_int(owner_id, default=0)

        tf_effective_raw = row_dict.get("True Forward Effective Date")
        next_tf_raw = row_dict.get("Next True Forward")
        sub_id_raw = row_dict.get("Subscription ID")
        po_raw = row_dict.get("Purchase Order Number")

        parts_ref = []
        if tf_effective_raw not in (None, ""):
            parts_ref.append(f"TF Effective Date: {str(tf_effective_raw).strip()}")
        if next_tf_raw not in (None, ""):
            parts_ref.append(f"Next TF: {str(next_tf_raw).strip()}")
        if sub_id_raw not in (None, ""):
            parts_ref.append(f"Subscription: {str(sub_id_raw).strip()}")
        if po_raw not in (None, ""):
            parts_ref.append(f"PO: {str(po_raw).strip()}")
        if weborder_str:
            parts_ref.append(f"WebOrderID: {weborder_str}")

        task_reference_35 = ", ".join(parts_ref) if parts_ref else None

        task_data = {
            "task_tasktype_id": 35,
            "task_reference": task_reference_35,
            "task_owner_id": owner_id,
            "task_temp_owner_id": None,
            "task_customer_id": task_customer_id,
            "task_cr_party_id": 0,
            "task_cr_party_name": None,
            "task_created_in": now,
            "task_created_by": 0,
            "task_priority": "HIGH",
            "task_project_id": 0,
            "task_status": 1,
            "task_status_justification": None,
            "task_start": task_start_date,
            "task_end": task_end_date,
            "task_start_performed": task_start_date,
            "task_end_performed": task_end_date,
            "task_value": task_value,
            "task_forecast": 0,
            "task_backlog": 0,
            "task_rate": 1,
            "task_currency": currency,
            "task_ws": ea_subscription_id if ea_subscription_id else None,
            "task_deal_id": None,
            "task_track": build_non_null_ref(offer_name_val, over_consumed_tf_groups),
            "task_subtrack": tf_groups if tf_groups else None,
            "task_highlight": 1,
            "task_remark": None,
            "task_description": None,
            "task_ea_flag": 0,
            "task_telemetry_flag": 0,
            "task_opt_in_flag": 0,
            "task_completed": 0,
            "task_architecture": None,
            "task_solution_domain": None,
            "task_eligible": "Y",
            "task_end_fy": None,
            "task_booking_date": None,
            "task_booking_amount": None,
        }

        try:
            task_id_raw = repo_task.insert(task_data)
            task_id = _safe_repo_int(task_id_raw, "task_id_insert_type_35", file_path, row_number, execution_log_path)
        except Exception as e:
            _safe_log(
                file_path=file_path,
                row_number=row_number,
                message=f"Erro ao criar tarefa tipo 35 (Over Consumed): {e}",
                column_name="Offer Name",
                value=offer_name,
            )
            return

        activity_data = {
            "activity_task_id": task_id,
            "activity_name": f"Over Consumed: {offer_name_val}",
            "activity_status": 1,
            "activity_ws": ea_subscription_id if ea_subscription_id else None,
            "activity_track": build_non_null_ref(offer_name_val, over_consumed_tf_groups),
            "activity_sub_track": tf_groups if tf_groups else None,
            "activity_start": task_start_date,
            "activity_end": task_end_date,
            "activity_start_performed": task_start_date,
            "activity_end_performed": task_end_date,
            "activity_value": task_value,
            "activity_currency": currency,
        }

        try:
            activity_id_raw = repo_activity.insert(activity_data)
            activity_id = _safe_repo_int(activity_id_raw, "activity_id_insert_type_35", file_path, row_number, execution_log_path)
        except Exception as e:
            _safe_log(
                file_path=file_path,
                row_number=row_number,
                message=f"Erro ao criar atividade para task tipo 35: {e}",
                column_name="Offer Name",
                value=offer_name,
            )
            return

        if execution_log_path and task_id:
            _append_execution_log(
                execution_log_path,
                f"INFO over_consumed task created row={row_number}, task_id={task_id}, activity_id={activity_id}, subscription_id={ea_subscription_id}",
            )

        try:
            remark_str = now.strftime("%Y-%b-%d")
            followup_date = (now.date() + timedelta(days=30)).strftime("%Y-%m-%d")
            history_task = {
                "taskrecord_task_id": task_id,
                "taskrecord_activity_id": 0,
                "taskrecord_remark": f"Task created at {remark_str}",
                "taskrecord_next_followup": followup_date,
                "taskrecord_updated_by": "System BA",
            }
            repo_history.insert(history_task)
        except Exception:
            pass

        try:
            history_activity = {
                "taskrecord_task_id": task_id,
                "taskrecord_activity_id": activity_id,
                "taskrecord_remark": f"Activity created at {remark_str}",
                "taskrecord_next_followup": (now.date() + timedelta(days=7)).strftime("%Y-%m-%d"),
                "taskrecord_updated_by": "System BA",
            }
            repo_history.insert(history_activity)
        except Exception:
            pass

    else:
        task_id = _safe_repo_int(found_task_info.get("task_id"), "task_id_found", file_path, row_number, execution_log_path)
        task_status_id = _safe_repo_int(found_task_info.get("task_status_id"), "task_status_id_found", file_path, row_number, execution_log_path)
        activity_id = _safe_repo_int(found_task_info.get("activity_id"), "activity_id_found", file_path, row_number, execution_log_path)
        activity_status_id = _safe_repo_int(found_task_info.get("activity_status_id"), "activity_status_id_found", file_path, row_number, execution_log_path)

        def union_update_list(current: str, new: str) -> Optional[str]:
            current_list = [v.strip() for v in (current or "").split(",") if v.strip()]
            new_list = [v.strip() for v in (new or "").split(",") if v.strip()]
            union = current_list.copy()
            for v in new_list:
                if v not in union:
                    union.append(v)
            if set(union) != set(current_list):
                return ", ".join(union)
            return None

        try:
            task_columns = repo_task.get_columns_by_task_id(
                task_id=task_id,
                columns=["task_track", "task_subtrack"],
                as_df=False,
            )
        except Exception:
            task_columns = None

        if activity_id:
            try:
                activity_columns = repo_activity.get_columns_by_activity_id(
                    activity_id=activity_id,
                    columns=["activity_status", "activity_value", "activity_track", "activity_sub_track"],
                    as_df=False,
                )
            except Exception:
                activity_columns = None
        else:
            activity_columns = None

        task_track = task_columns.get("task_track", "") if task_columns else ""
        task_subtrack = task_columns.get("task_subtrack", "") if task_columns else ""

        activity_track = activity_columns.get("activity_track", "") if activity_columns else ""
        activity_subtrack = activity_columns.get("activity_sub_track", "") if activity_columns else ""

        monthly_charge_val = task_value

        need_new_activity = False
        update_task_fields = {}
        update_activity_fields = {}

        combined_track = build_non_null_ref(offer_name_val, over_consumed_tf_groups)
        combined_subtrack = tf_groups if tf_groups else ""

        new_task_track = union_update_list(task_track, combined_track)
        new_task_subtrack = union_update_list(task_subtrack, combined_subtrack)

        if new_task_track:
            update_task_fields["task_track"] = new_task_track

        if new_task_subtrack:
            update_task_fields["task_subtrack"] = new_task_subtrack

        new_activity_track = union_update_list(activity_track, combined_track)
        new_activity_subtrack = union_update_list(activity_subtrack, combined_subtrack)

        activity_value_current = activity_columns.get("activity_value", 0) if activity_columns else 0
        try:
            activity_value_current_num = float(activity_value_current) if activity_value_current is not None else 0.0
        except Exception:
            activity_value_current_num = 0.0

        if new_activity_track is None and monthly_charge_val > activity_value_current_num:
            need_new_activity = True

        if new_activity_track:
            update_activity_fields["activity_track"] = new_activity_track

        if new_activity_subtrack:
            update_activity_fields["activity_sub_track"] = new_activity_subtrack

        if update_task_fields and task_id:
            update_task_fields["task_status"] = 2
            try:
                repo_task.update(update_task_fields, where={"task_id": task_id})
                repo_history.insert(
                    {
                        "taskrecord_task_id": task_id,
                        "taskrecord_activity_id": 0,
                        "taskrecord_remark": "Task reopened with updated tracks",
                        "taskrecord_updated_by": "System BA",
                    }
                )
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO over_consumed task updated row={row_number}, task_id={task_id}",
                    )
            except Exception:
                pass

        if task_status_id not in (4, 5, 6, 10) and activity_id == 0:
            need_new_activity = True
        elif task_status_id not in (4, 5, 6, 10) and activity_status_id not in (4, 5, 6, 10):
            need_new_activity = False
        elif task_status_id not in (4, 5, 6, 10) and activity_status_id in (0, 4, 5, 6, 10):
            if not new_activity_track:
                need_new_activity = True

        if need_new_activity and task_id:
            try:
                activity_name = f"Over Consumed: {offer_name_val}"
                new_activity_data = {
                    "activity_task_id": task_id,
                    "activity_name": activity_name,
                    "activity_status": 1,
                    "activity_ws": ea_subscription_id if ea_subscription_id else None,
                    "activity_track": combined_track,
                    "activity_sub_track": combined_subtrack,
                    "activity_start": task_start_date,
                    "activity_end": task_end_date,
                    "activity_start_performed": task_start_date,
                    "activity_end_performed": task_end_date,
                    "activity_value": task_value,
                    "activity_currency": currency,
                }
                new_activity_id_raw = repo_activity.insert(new_activity_data)
                new_activity_id = _safe_repo_int(new_activity_id_raw, "new_activity_id_insert", file_path, row_number, execution_log_path)

                if new_activity_id:
                    repo_history.insert(
                        {
                            "taskrecord_task_id": task_id,
                            "taskrecord_activity_id": new_activity_id,
                            "taskrecord_remark": f"Activity created at {now.strftime('%Y-%b-%d')}",
                            "taskrecord_next_followup": (now.date() + timedelta(days=7)).strftime("%Y-%m-%d"),
                            "taskrecord_updated_by": "System BA",
                        }
                    )
                    if execution_log_path:
                        _append_execution_log(
                            execution_log_path,
                            f"INFO over_consumed new activity created row={row_number}, task_id={task_id}, activity_id={new_activity_id}",
                        )
            except Exception as e:
                _safe_log(
                    file_path=file_path,
                    row_number=row_number,
                    message=f"Erro ao criar nova atividade para task tipo 35: {e}",
                    column_name="Offer Name",
                    value=offer_name,
                )
                return

        if update_activity_fields and activity_id:
            update_activity_fields["activity_status"] = 2
            try:
                repo_activity.update(update_activity_fields, where={"activity_id": activity_id})
                repo_history.insert(
                    {
                        "taskrecord_task_id": task_id,
                        "taskrecord_activity_id": activity_id,
                        "taskrecord_remark": "Activity reopened with updated tracks",
                        "taskrecord_updated_by": "System BA",
                    }
                )
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO over_consumed activity updated row={row_number}, activity_id={activity_id}",
                    )
            except Exception:
                pass


def _process_single_row(
    row_dict: Dict[str, Any],
    file_path: str,
    row_number: int,
    execution_log_path: Optional[Path] = None,
) -> RowProcessResult:
    """
    Processa uma única linha e retorna sucesso/erro com detalhes.
    """
    end_customer_name = None
    weborder_str = None
    offer_name = None

    try:
        weborder_raw = row_dict.get("WebOrderID")
        weborder_str = _normalize_weborder(weborder_raw)
        if not weborder_str:
            message = "WebOrder inválida ou ausente em 'WebOrderID'"
            _safe_log(file_path, row_number, message, "WebOrderID", weborder_raw)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} invalid WebOrderID value={str(weborder_raw)[:500]}",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="WebOrderID",
                error_value=weborder_raw,
            )

        end_customer_raw = row_dict.get("End Customer")
        end_customer_name = _normalize_str(end_customer_raw)
        end_customer_id = _get_customer_id_from_end_customer(
            end_customer_name,
            file_path,
            row_number,
            execution_log_path,
        )
        if not end_customer_id:
            message = "Customer não encontrado a partir de 'End Customer'"
            _safe_log(file_path, row_number, message, "End Customer", end_customer_name)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} customer not found end_customer={str(end_customer_name)[:500]}",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="End Customer",
                error_value=end_customer_name,
            )

        try:
            ciscoweborder_id = repo_weborder.find_id_by_code_and_customer(
                code=weborder_str,
                customer_id=end_customer_id,
            )
        except Exception as e:
            _safe_log(
                file_path,
                row_number,
                f"Erro ao buscar CiscoWebOrder: {e}",
                "WebOrderID",
                weborder_str,
            )
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"WARN row={row_number} error finding CiscoWebOrder weborder={weborder_str} error={str(e)[:500]}",
                )
            ciscoweborder_id = None

        if ciscoweborder_id:
            ciscoweborder_id = _safe_repo_int(
                ciscoweborder_id,
                "ciscoweborder_id_found",
                file_path,
                row_number,
                execution_log_path,
            )

        if not ciscoweborder_id:
            try:
                inserted_weborder_id = repo_weborder.insert(
                    {
                        "weborder_number": weborder_str,
                        "weborder_customer_id": end_customer_id,
                    }
                )
                ciscoweborder_id = _safe_repo_int(
                    inserted_weborder_id,
                    "ciscoweborder_id_insert",
                    file_path,
                    row_number,
                    execution_log_path,
                )
                if not ciscoweborder_id:
                    raise ValueError("ID inválido retornado ao criar CiscoWebOrder")

                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} CiscoWebOrder created weborder_id={ciscoweborder_id} weborder={weborder_str}",
                    )
            except Exception as e:
                message = f"Falha ao criar WebOrder na tbCiscoWebOrder: {e}"
                _safe_log(file_path, row_number, message, "WebOrderID", weborder_str)
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"ERROR row={row_number} failed creating CiscoWebOrder weborder={weborder_str} error={str(e)[:500]}",
                    )
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="WebOrderID",
                    error_value=weborder_str,
                )

        offer_name_raw = row_dict.get("Offer Name")
        offer_name = _normalize_str(offer_name_raw)
        product_id = _get_or_create_product_id_from_offer_name(
            offer_name,
            file_path,
            row_number,
            execution_log_path,
        )
        if not product_id:
            message = "Product não encontrado/criado a partir de 'Offer Name'"
            _safe_log(file_path, row_number, message, "Offer Name", offer_name)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} product not found/created offer_name={str(offer_name)[:500]}",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Offer Name",
                error_value=offer_name,
            )

        payload = _build_ea_payload_from_row(
            row=row_dict,
            customer_id=end_customer_id,
            product_id=product_id,
            weborder_str=weborder_str,
        )

        offer_name_for_task = (offer_name or "").strip().upper()

        ea_subscription_id = payload.get("ea_subscription_id")
        ea_start_date = payload.get("ea_start_date")
        ea_end_date = payload.get("ea_end_date")
        ea_renewal_date = payload.get("ea_renewal_date")

        ignore_subscription = False
        if ea_subscription_id:
            try:
                ignored_id = repo_subscription_ignored.get_id(
                    subscription_id=ea_subscription_id,
                    customer_id=end_customer_id,
                )
                ignored_id_int = (
                    _safe_repo_int(
                        ignored_id,
                        "subscription_ignored_id",
                        file_path,
                        row_number,
                        execution_log_path,
                    )
                    if ignored_id
                    else 0
                )
                if ignored_id_int:
                    ignore_subscription = True
            except Exception:
                ignore_subscription = False

        where: Dict[str, Any] = {
            "ea_web_order_id": weborder_str,
            "ea_end_customer_id": end_customer_id,
            "ea_product_id": product_id,
            "ea_start_date": ea_start_date,
            "ea_end_date": ea_end_date,
            "ea_renewal_date": ea_renewal_date,
        }

        if ea_subscription_id:
            where["ea_subscription_id"] = ea_subscription_id
        else:
            where["ea_subscription_id"] = None

        try:
            ea_ids = repo_ea.find_ids_by(where)
        except Exception as e:
            message = f"Erro ao buscar EA existente: {e}"
            _safe_log(file_path, row_number, message, "Offer Name", offer_name)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} failed searching existing EA error={str(e)[:500]}",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Offer Name",
                error_value=offer_name,
            )

        if ea_ids:
            ea_id = _safe_repo_int(
                ea_ids[0],
                "ea_id_found",
                file_path,
                row_number,
                execution_log_path,
            )
            if not ea_id:
                message = "ID inválido retornado ao buscar EA existente"
                _safe_log(file_path, row_number, message, "Offer Name", offer_name)
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"ERROR row={row_number} invalid existing EA id",
                    )
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Offer Name",
                    error_value=offer_name,
                )

            try:
                repo_ea.update(payload, where={"ea_id": ea_id})
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"INFO row={row_number} EA updated ea_id={ea_id} subscription_id={str(ea_subscription_id)[:255] if ea_subscription_id else None}",
                    )
            except Exception as e:
                log_column, log_value = _get_excel_validation_error_info(
                    message=str(e),
                    row_dict=row_dict,
                    offer_name=offer_name,
                    weborder_str=weborder_str,
                )
                message = f"Erro ao atualizar EA existente (ea_id={ea_id}): {e}"
                _safe_log(file_path, row_number, message, log_column, log_value)
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"ERROR row={row_number} failed updating EA ea_id={ea_id} error={str(e)[:500]}",
                    )
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column=log_column,
                    error_value=log_value,
                )

            consumption_status = payload.get("ea_consumption_status")
            if (
                not ignore_subscription
                and consumption_status
                and consumption_status.upper() == "OVER CONSUMED"
            ):
                _handle_over_consumed(
                    payload=payload,
                    row_dict=row_dict,
                    file_path=file_path,
                    row_number=row_number,
                    offer_name=offer_name,
                    weborder_str=weborder_str,
                    ea_subscription_id=ea_subscription_id,
                    task_customer_id=end_customer_id,
                    execution_log_path=execution_log_path,
                )

            return RowProcessResult(success=True)

        try:
            ea_id_raw = repo_ea.insert(payload)
            ea_id = _safe_repo_int(
                ea_id_raw,
                "ea_id_insert",
                file_path,
                row_number,
                execution_log_path,
            )
        except Exception as e:
            log_column, log_value = _get_excel_validation_error_info(
                message=str(e),
                row_dict=row_dict,
                offer_name=offer_name,
                weborder_str=weborder_str,
            )
            message = f"Erro ao inserir novo EA: {e}"
            _safe_log(file_path, row_number, message, log_column, log_value)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} failed inserting EA error={str(e)[:500]}",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column=log_column,
                error_value=log_value,
            )

        if not ea_id:
            message = "Falha ao obter ea_id válido após inserir novo EA"
            _safe_log(file_path, row_number, message, "Offer Name", offer_name)
            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"ERROR row={row_number} invalid EA id after insert",
                )
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Offer Name",
                error_value=offer_name,
            )

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO row={row_number} EA inserted ea_id={ea_id} subscription_id={str(ea_subscription_id)[:255] if ea_subscription_id else None}",
            )

        if (
            not ignore_subscription
            and offer_name_for_task in {"EA3-M", "ELA2-M", "A-FLEX", "A-FLEX-3"}
        ):
            try:
                owner_id = repo_task.get_last_task_owner_by_company(
                    company_id=end_customer_id,
                    user_type="CSM",
                )
            except Exception:
                owner_id = 0

            owner_id = _safe_int(owner_id, default=0)

            now = datetime.now()
            task_start = now.date()
            task_end = task_start + timedelta(days=45)

            sub_id = ea_subscription_id
            po_num = _normalize_identifier(row_dict.get("Purchase Order Number"), max_length=255)

            ref_parts = []
            if sub_id:
                ref_parts.append(f"Subscription ID: {sub_id}")
            if po_num:
                ref_parts.append(f"PO: {po_num}")
            if weborder_str:
                ref_parts.append(f"WEB ORDER: {weborder_str}")

            task_reference = ", ".join(ref_parts) if ref_parts else None

            task_data = {
                "task_tasktype_id": 1,
                "task_reference": task_reference,
                "task_owner_id": owner_id,
                "task_temp_owner_id": None,
                "task_customer_id": end_customer_id,
                "task_cr_party_id": 0,
                "task_cr_party_name": None,
                "task_created_in": now,
                "task_created_by": 0,
                "task_priority": "LOW",
                "task_project_id": 0,
                "task_status": 1,
                "task_status_justification": None,
                "task_start": task_start,
                "task_end": task_end,
                "task_start_performed": task_start,
                "task_end_performed": task_end,
                "task_value": 0,
                "task_forecast": 0,
                "task_backlog": 0,
                "task_rate": 1,
                "task_currency": "USD",
                "task_ws": sub_id if sub_id else None,
                "task_deal_id": None,
                "task_track": offer_name_for_task,
                "task_subtrack": offer_name_for_task,
                "task_highlight": 0,
                "task_remark": None,
                "task_description": None,
                "task_ea_flag": 0,
                "task_telemetry_flag": 0,
                "task_opt_in_flag": 0,
                "task_completed": 0,
                "task_architecture": None,
                "task_solution_domain": None,
                "task_eligible": "Y",
                "task_end_fy": None,
                "task_booking_date": None,
                "task_booking_amount": None,
            }

            try:
                task_id_raw = repo_task.insert(task_data)
                task_id = _safe_repo_int(
                    task_id_raw,
                    "task_id_insert_type_1",
                    file_path,
                    row_number,
                    execution_log_path,
                )
                if not task_id:
                    raise ValueError("ID inválido retornado ao criar task tipo 1")
            except Exception as e:
                message = f"Erro ao criar tarefa tipo 1 (New CISCO EA): {e}"
                _safe_log(file_path, row_number, message, "Offer Name", offer_name)
                if execution_log_path:
                    _append_execution_log(
                        execution_log_path,
                        f"ERROR row={row_number} failed creating task type 1 error={str(e)[:500]}",
                    )
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Offer Name",
                    error_value=offer_name,
                )

            try:
                repo_ea.update({"ea_new_task_id": task_id}, where={"ea_id": ea_id})
            except Exception:
                pass

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} task type 1 created task_id={task_id} ea_id={ea_id}",
                )

            try:
                remark_date_str = now.strftime("%Y-%b-%d")
                next_followup_date = (now.date() + timedelta(days=45)).strftime("%Y-%m-%d")
                history_data = {
                    "taskrecord_task_id": task_id,
                    "taskrecord_activity_id": 0,
                    "taskrecord_remark": f"Task created at {remark_date_str}",
                    "taskrecord_next_followup": next_followup_date,
                    "taskrecord_updated_by": "System BA",
                }
                repo_history.insert(history_data)
            except Exception:
                pass

        return RowProcessResult(success=True)

    except Exception as ex:
        message = f"Unexpected error: {ex}"
        _safe_log(file_path, row_number, message, "End Customer", end_customer_name)
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"ERROR row={row_number} unexpected error={str(ex)[:1000]}",
            )
        return RowProcessResult(
            success=False,
            error_message=message,
            error_column="End Customer",
            error_value=end_customer_name,
        )


def _process_chunk(
    ws: Worksheet,
    idx_map: Dict[str, int],
    headers: List[str],
    start_row: int,
    end_row: int,
    file_path: str,
    failed_output_path: Path,
    execution_log_path: Optional[Path] = None,
    chunk_number: int = 0,
) -> Tuple[int, int]:
    """
    Processa um bloco de linhas.

    Regras:
    - toda linha lida é removida do arquivo original
    - se falhar, vai para o XLSX fixo de falhas
    - se tiver sucesso, apenas é removida do original
    """
    imported_in_chunk = 0
    error_in_chunk = 0
    failed_rows_buffer: List[List[Any]] = []

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_start chunk={chunk_number} start_row={start_row} end_row={end_row}",
        )

    for r in range(end_row, start_row - 1, -1):
        row_cells = ws[r]
        raw_values = [cell.value for cell in row_cells]
        row_dict: Dict[str, Any] = {}

        for h, idx in idx_map.items():
            row_dict[h] = raw_values[idx] if idx < len(raw_values) else None

        result = _process_single_row(
            row_dict=row_dict,
            file_path=file_path,
            row_number=r,
            execution_log_path=execution_log_path,
        )

        if result.success:
            imported_in_chunk += 1
        else:
            error_in_chunk += 1
            failed_rows_buffer.append(
                _row_dict_to_failed_excel_row(
                    headers=headers,
                    row_dict=row_dict,
                    original_row_number=r,
                    input_file_path=file_path,
                    result=result,
                )
            )

        ws.delete_rows(r, 1)

    if failed_rows_buffer:
        failed_rows_buffer.reverse()
        _append_failed_rows(
            failed_path=failed_output_path,
            rows_to_append=failed_rows_buffer,
            original_headers=headers,
            execution_log_path=execution_log_path,
        )

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_finish chunk={chunk_number} success={imported_in_chunk} error={error_in_chunk} total_processed={imported_in_chunk + error_in_chunk}",
        )

    return imported_in_chunk, error_in_chunk


def run_import(file_path: str, user_id: Optional[str] = None):
    """
    Fluxo refatorado:

    1. Lê o XLSX original em .storage/input
    2. Processa linhas em chunks
    3. Remove do original toda linha lida, com sucesso ou erro
    4. Salva linhas com erro em um XLSX fixo em .storage/output com append
    5. Em retomada, o original contém apenas linhas ainda não lidas
    6. Gera um arquivo texto de log por execução em .storage/logs

    Correção adicional:
    - protege conversões para inteiro contra strings gigantes
    """
    CHUNK_SIZE = 2000
    imported_rows = 0
    error_rows = 0
    chunk_number = 0
    started_at = datetime.now()

    _ensure_directories()

    input_path = _resolve_input_path(file_path)
    failed_output_path = _build_failed_output_path(input_path)
    execution_log_path = _build_execution_log_path(input_path)

    _append_execution_log(execution_log_path, "START import")
    _append_execution_log(execution_log_path, f"INFO import_source={IMPORT_SOURCE}")
    _append_execution_log(execution_log_path, f"INFO user_id={user_id}")
    _append_execution_log(execution_log_path, f"INFO input_file={str(input_path)}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")

    try:
        wb, ws, idx_map, headers = _open_workbook_rw(str(input_path))
    except Exception as e:
        msg = f"Erro ao abrir arquivo Excel: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR failed opening workbook error={str(e)[:1000]}")
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": 0,
                "imported_rows": 0,
                "error_rows": 0,
                "execution_log_path": str(execution_log_path),
            },
        }

    total_rows = ws.max_row - 1
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO chunk_size={CHUNK_SIZE}")

    required_schema = IMPORT_SCHEMAS.get(IMPORT_SOURCE, {})
    required_columns = required_schema.get("required", [])

    missing_cols = [c for c in required_columns if c not in headers]
    if missing_cols:
        msg = "Missing columns: " + ", ".join(missing_cols)
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR missing_columns={', '.join(missing_cols)}")
        wb.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "imported_rows": 0,
                "error_rows": total_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    # Removido _ensure_failed_workbook aqui para que o arquivo de falhas
    # só seja criado quando realmente houver linhas com erro (dentro de _append_failed_rows).

    try:
        while ws.max_row > 1:
            chunk_number += 1
            chunk_end = min(1 + CHUNK_SIZE, ws.max_row)

            imported_chunk, error_chunk = _process_chunk(
                ws=ws,
                idx_map=idx_map,
                headers=headers,
                start_row=2,
                end_row=chunk_end,
                file_path=str(input_path),
                failed_output_path=failed_output_path,
                execution_log_path=execution_log_path,
                chunk_number=chunk_number,
            )

            imported_rows += imported_chunk
            error_rows += error_chunk

            wb.save(str(input_path))

            _append_execution_log(
                execution_log_path,
                f"INFO checkpoint_after_chunk chunk={chunk_number} imported_total={imported_rows} error_total={error_rows} remaining_rows_current={max(ws.max_row - 1, 0)}",
            )
    except Exception as e:
        msg = f"Erro durante o processamento da importação: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR processing failed error={str(e)[:1000]}")
        try:
            wb.save(str(input_path))
            _append_execution_log(execution_log_path, "INFO workbook saved after processing exception")
        except Exception as save_ex:
            _append_execution_log(
                execution_log_path,
                f"ERROR failed saving workbook after exception error={str(save_ex)[:1000]}",
            )
        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED imported_rows={imported_rows} error_rows={error_rows} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "imported_rows": imported_rows,
                "error_rows": error_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        wb.save(str(input_path))
        _append_execution_log(execution_log_path, "INFO workbook saved successfully at end")
    except Exception as e:
        msg = f"Erro ao salvar arquivo Excel após processamento: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR final workbook save failed error={str(e)[:1000]}")
        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED imported_rows={imported_rows} error_rows={error_rows} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "imported_rows": imported_rows,
                "error_rows": error_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    wb.close()

    remaining_rows_in_input = 0
    try:
        wb_check = load_workbook(str(input_path), read_only=True, data_only=True)
        ws_check = wb_check.active
        remaining_rows_in_input = max(ws_check.max_row - 1, 0)
        wb_check.close()
    except Exception as e:
        remaining_rows_in_input = -1
        _append_execution_log(
            execution_log_path,
            f"WARN failed checking remaining rows error={str(e)[:1000]}",
        )

    finished_at = datetime.now()
    duration_seconds = int((finished_at - started_at).total_seconds())

    _append_execution_log(execution_log_path, f"INFO total_chunks={chunk_number}")
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO imported_rows={imported_rows}")
    _append_execution_log(execution_log_path, f"INFO error_rows={error_rows}")
    _append_execution_log(execution_log_path, f"INFO remaining_rows_in_input={remaining_rows_in_input}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")
    _append_execution_log(execution_log_path, f"INFO duration_seconds={duration_seconds}")

    if error_rows > 0:
        status = "FAILED"
        msg = (
            f"CiscoSubscriptionCCW import concluído com falhas. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={imported_rows}, "
            f"erros={error_rows}, linhas_restantes_input={remaining_rows_in_input}, "
            f"arquivo_falhas={str(failed_output_path)}, log_execucao={str(execution_log_path)}."
        )
    else:
        status = "FINISHED"
        msg = (
            f"CiscoSubscriptionCCW import concluído com sucesso. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={imported_rows}, "
            f"erros={error_rows}, linhas_restantes_input={remaining_rows_in_input}, "
            f"arquivo_falhas={str(failed_output_path)}, log_execucao={str(execution_log_path)}."
        )

    _append_execution_log(
        execution_log_path,
        f"FINISH status={status} imported_rows={imported_rows} error_rows={error_rows} remaining_rows_in_input={remaining_rows_in_input} duration_seconds={duration_seconds}",
    )

    return {
        "status": status,
        "message": msg,
        "summary": {
            "file_path": str(input_path),
            "total_rows": total_rows,
            "imported_rows": imported_rows,
            "error_rows": error_rows,
            "failed_file_path": str(failed_output_path),
            "remaining_rows_in_input": remaining_rows_in_input,
            "execution_log_path": str(execution_log_path),
            "duration_seconds": duration_seconds,
            "total_chunks": chunk_number,
        },
    }
