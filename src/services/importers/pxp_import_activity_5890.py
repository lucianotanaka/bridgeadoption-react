"""
pxp_import_activity_5890.py

Importador PXP de activities Cisco LCI (processo 5890).
"""

from __future__ import annotations

import logging
import re
import traceback
import json
from dataclasses import dataclass
from datetime import datetime, date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.infrastructure.database.repositories.task_repository import TaskRepository
from src.infrastructure.database.repositories.task_activity_repository import TaskActivityRepository
from src.infrastructure.database.repositories.task_history_repository import TaskHistoryRepository
from src.infrastructure.database.repositories.company_list_name_repository import CompanyListNameRepository

try:
    from src.infrastructure.database.repositories.import_log_repository import ImportLogRepository
except Exception:
    ImportLogRepository = None

try:
    from src.infrastructure.database.repositories.import_control_repository import (
        ImportControlRepository,
        ImportStatus,
    )
except Exception:
    ImportControlRepository = None
    ImportStatus = None


BASE_STORAGE_PATH = Path("/home/bridgeadoption/storage")
BASE_INPUT_PATH = BASE_STORAGE_PATH / "input"
BASE_OUTPUT_PATH = BASE_STORAGE_PATH / "output"
BASE_LOGS_PATH = BASE_STORAGE_PATH / "logs"

IMPORT_SOURCE = "PxpImportActivity5890"

ERROR_EXTRA_COLUMNS = [
    "import_error_message",
    "import_error_column",
    "import_error_value",
    "import_original_row",
    "import_processed_at",
    "import_source",
    "import_original_file",
]

logger = logging.getLogger(__name__)

repo_task = TaskRepository()
repo_activity = TaskActivityRepository()
repo_history = TaskHistoryRepository()
repo_company = CompanyListNameRepository()
repo_log = ImportLogRepository() if ImportLogRepository else None
repo_import_control = ImportControlRepository() if ImportControlRepository else None

IMPORT_CONTROL_PROGRESS_INTERVAL_CHUNKS = 1


CHUNK_SIZE = 1000

ACTIVITY_STATUS_OPEN = 1
ACTIVITY_STATUS_IN_PROGRESS = 2
ACTIVITY_STATUS_ON_HOLD = 3
ACTIVITY_STATUS_CANCELLED = 4
ACTIVITY_STATUS_DECLINED = 5
ACTIVITY_STATUS_EXPIRED = 6
ACTIVITY_STATUS_SUBMITTED = 7
ACTIVITY_STATUS_RESUBMITTED = 8
ACTIVITY_STATUS_APPROVED_TO_CLOSE = 9
ACTIVITY_STATUS_CLOSED = 10

CLOSED_ACTIVITY_STATUSES = {4, 5, 6, 10}
CLOSED_TASK_STATUSES = {4, 5, 6, 10}

TASK_TYPE_21 = 21
TASK_TYPE_22 = 22

REQUIRED_COLUMNS = [
    "Track",
    "Sub-Track",
    "Activity Id",
    "CR Party ID",
    "Deal CR Party Name",
    "Activity Type",
    "Activity Group",
    "Stage",
    "Activity Start Date",
    "Activity Expiration Date",
    "Claim Approval Status",
    "Claim Approved Amount",
    "Claim Approved Amount (Currency)",
    "Claim Submitted By",
    "Claim Submitted Date",
    "Deal ID",
    "Activity Requested Amount",
    "Activity Requested Amount (Currency)",
    "Onboard Stage Completion Date",
    "Adopt Stage Completion Date",
    "Engage Stage Completion Date",
    "Use Stage Completion Date",
    "Claim Approval Date",
    "L1 Claim Decision Date",
]


@dataclass
class RowProcessResult:
    success: bool
    error_message: Optional[str] = None
    error_column: Optional[str] = None
    error_value: Optional[Any] = None
    ignored: bool = False
    discarded_old_row: bool = False
    created: bool = False
    updated: bool = False
    task_updated: bool = False


def _ensure_directories() -> None:
    BASE_INPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_OUTPUT_PATH.mkdir(parents=True, exist_ok=True)
    BASE_LOGS_PATH.mkdir(parents=True, exist_ok=True)


def _build_execution_log_path(input_path: Path) -> Path:
    return BASE_LOGS_PATH / f"{input_path.stem}.log"


def _report_progress(
    import_control_id: Optional[int],
    total_rows: int,
    processed_success: int,
    failed_rows: int,
    ignored_rows: int,
    remaining_rows: int,
    execution_log_path: Optional[Path] = None,
) -> None:
    if repo_import_control is None or import_control_id is None:
        return

    lidas = total_rows - remaining_rows
    msg = (
        f"Arquivo processado. "
        f"Total linhas origem={total_rows} | "
        f"lidas={lidas} | "
        f"sucesso={processed_success} | "
        f"erros={failed_rows} | "
        f"ignoradas={ignored_rows} | "
        f"restantes={remaining_rows}"
    )

    try:
        repo_import_control.update_status(
            importctrl_id=import_control_id,
            status=ImportStatus.RUNNING,
            message=msg,
        )
    except Exception as e:
        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"WARN failed updating import control progress error={str(e)[:500]}",
            )


def _append_execution_log(log_path: Path, message: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{ts}] {message}\n")


def _safe_log(
    file_path: str,
    row_number: int,
    message: str,
    column_name: Optional[str] = None,
    value: Optional[Any] = None,
) -> None:
    if repo_log is None:
        return

    try:
        normalized_value = value

        # Converte estruturas complexas para JSON antes de persistir no banco.
        if isinstance(value, (dict, list, tuple)):
            normalized_value = json.dumps(value, ensure_ascii=False)

        # Converte tipos simples não textuais para string.
        elif value is not None and not isinstance(value, str):
            normalized_value = str(value)

        repo_log.create(
            IMPORT_SOURCE,
            file_path,
            row_number,
            message,
            column_name,
            normalized_value,
        )
    except Exception as log_ex:
        logger.exception(
            "Falha ao gravar tbImportLog. file=%s row=%s column=%s value=%s error=%s",
            file_path,
            row_number,
            column_name,
            value,
            log_ex,
        )


def _normalize_str(value: Any, max_length: int = 1000) -> Optional[str]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    return text[:max_length]


def _safe_int(value: Any, default: int = 0) -> int:
    if value is None:
        return default

    if isinstance(value, bool):
        return default

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value) if value.is_integer() else default

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]

    if text.lstrip("+-").isdigit():
        try:
            return int(text)
        except Exception:
            return default

    return default


def _to_decimal(value: Any) -> Optional[Decimal]:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    text = text.replace(",", ".")
    text = re.sub(r"[^0-9\.\-+]", "", text)

    if text in {"", ".", "-", "+", "-.", "+."}:
        return None

    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _to_float(value: Any) -> Optional[float]:
    dec = _to_decimal(value)
    return float(dec) if dec is not None else None


def _to_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()[:50]

    for fmt in (
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%m/%d/%Y",
        "%d-%m-%Y",
        "%Y/%m/%d",
        "%d %b %Y",
        "%d %B %Y",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def _normalize_currency(value: Any, default: str = "USD") -> str:
    text = _normalize_str(value, max_length=50)
    if not text:
        return default

    text = text.replace("($)", "").replace("$", "").replace(" ", "").strip()
    return text or default


def _normalize_stage_key(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=200)
    if not text:
        return None

    text = text.replace("–", "-")
    text = text.replace("—", "-")
    text = re.sub(r"\s+", "", text).lower()
    return text


def _normalize_activity_group(value: Any) -> Optional[str]:
    text = _normalize_str(value, max_length=255)
    if not text:
        return None
    return text


def _resolve_task_type_priority_from_activity_group(activity_group: Any) -> List[int]:
    normalized = _normalize_activity_group(activity_group)

    if normalized is None:
        return [TASK_TYPE_22, TASK_TYPE_21]

    if normalized.strip().lower() == "lci2.0":
        return [TASK_TYPE_22]

    return [TASK_TYPE_21, TASK_TYPE_22]


def _fiscal_year_of_date(dt: Optional[date]) -> int:
    if dt is None:
        return datetime.now().year

    if dt.month >= 8:
        return dt.year + 1
    return dt.year


def _fiscal_year_apr_mar(dt: Optional[date]) -> Optional[int]:
    """
    Fiscal year de Abril até Março.
    Ex:
    22-06-2026 -> FY 2026
    22-03-2027 -> FY 2026
    """
    if dt is None:
        return None

    if dt.month >= 4:
        return dt.year
    return dt.year - 1


def _current_fiscal_year() -> int:
    return _fiscal_year_of_date(datetime.now().date())


def _resolve_input_path(file_path: str) -> Path:
    path = Path(file_path)
    if path.is_absolute():
        return path
    return BASE_INPUT_PATH / path.name


def _build_failed_output_path(input_path: Path) -> Path:
    return BASE_OUTPUT_PATH / f"{input_path.stem}_failed_rows.xlsx"


def _open_workbook_rw(file_path: str) -> Tuple[Any, Worksheet, Dict[str, int], List[str]]:
    wb = load_workbook(filename=file_path, read_only=False, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    idx_map = {h: i for i, h in enumerate(headers)}
    return wb, ws, idx_map, headers


def _ensure_failed_workbook(
    failed_path: Path,
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
) -> None:
    """
    Cria (ou recria) o arquivo de falhas com apenas o header.

    Sempre sobrescreve o arquivo se já existir, garantindo que não haja
    resíduos de execuções anteriores. A criação só será feita de fato
    se esta função for chamada (ou seja, somente quando houver falhas
    a serem escritas).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "failed_rows"
    ws.append(original_headers + ERROR_EXTRA_COLUMNS)
    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO failed file created (overwritten if existed): {str(failed_path)}",
        )


def _append_failed_rows(
    failed_path: Path,
    rows_to_append: List[List[Any]],
    original_headers: List[str],
    execution_log_path: Optional[Path] = None,
) -> None:
    """
    Acrescenta linhas de falha ao arquivo de falhas.

    - Se rows_to_append estiver vazio, não faz nada e não cria arquivo.
    - Se o arquivo ainda não existir, ele é criado/recriado com header
      (sobrescrevendo conteúdo anterior) via _ensure_failed_workbook.
    """
    if not rows_to_append:
        return

    _ensure_failed_workbook(failed_path, original_headers, execution_log_path)

    wb = load_workbook(str(failed_path))
    ws = wb.active

    for row in rows_to_append:
        ws.append(row)

    wb.save(str(failed_path))
    wb.close()

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO appended failed rows count={len(rows_to_append)} to file={str(failed_path)}",
        )


def _row_dict_to_failed_excel_row(
    headers: List[str],
    row_dict: Dict[str, Any],
    original_row_number: int,
    input_file_path: str,
    result: RowProcessResult,
) -> List[Any]:
    processed_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    original_values = [row_dict.get(h) for h in headers]

    return original_values + [
        result.error_message,
        result.error_column,
        str(result.error_value)[:2000] if result.error_value is not None else None,
        original_row_number,
        processed_at,
        IMPORT_SOURCE,
        input_file_path,
    ]


def _validate_required_columns(headers: List[str]) -> List[str]:
    return [col for col in REQUIRED_COLUMNS if col not in headers]


def _resolve_customer_id(company_name: Optional[str], deal_id: Optional[str]) -> int:
    if company_name:
        try:
            company_id = repo_company.get_company_id_by_name(company_name)
            company_id_int = _safe_int(company_id, default=0)
            if company_id_int:
                return company_id_int
        except Exception:
            pass

    if deal_id:
        try:
            task_ids = repo_task.find_ids_by({"task_deal_id": deal_id})
            customer_ids = set()

            for task_id in task_ids or []:
                tid = _safe_int(task_id, default=0)
                if not tid:
                    continue

                rows = repo_task.get_task(task_id=tid, as_df=False)
                if not rows:
                    continue

                customer_id = _safe_int(rows[0].get("task_customer_id"), default=0)
                if customer_id:
                    customer_ids.add(customer_id)

            if len(customer_ids) == 1:
                return list(customer_ids)[0]
        except Exception:
            pass

    return 0


def _resolve_seq_and_stage_completion(row_dict: Dict[str, Any]) -> Tuple[int, Optional[date], Optional[str]]:
    activity_type = _normalize_str(row_dict.get("Activity Type"), max_length=255)

    if not activity_type:
        return 0, None, None

    activity_type_lower = activity_type.lower()

    if "onboard" in activity_type_lower:
        return 1, _to_date(row_dict.get("Onboard Stage Completion Date")), activity_type
    if "use" in activity_type_lower:
        return 2, _to_date(row_dict.get("Use Stage Completion Date")), activity_type
    if "engage" in activity_type_lower:
        return 3, _to_date(row_dict.get("Engage Stage Completion Date")), activity_type
    if "adopt" in activity_type_lower:
        return 4, _to_date(row_dict.get("Adopt Stage Completion Date")), activity_type

    return 0, None, activity_type


def _resolve_status_from_stage(stage: Optional[str]) -> Tuple[int, Dict[str, Any], str]:
    stage_key = _normalize_stage_key(stage)
    if not stage_key:
        raise ValueError("Stage vazio ou inválido")

    if stage_key.startswith("payment-type"):
        return (
            ACTIVITY_STATUS_CLOSED,
            {"activity_approved": 1, "activity_completed": 1},
            f"{stage} - status mapped to CLOSED",
        )

    mapping = {
        "activity-approved": (
            ACTIVITY_STATUS_OPEN,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to OPEN",
        ),
        "activity-declined": (
            ACTIVITY_STATUS_DECLINED,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to DECLINED",
        ),
        "activity-expired": (
            ACTIVITY_STATUS_EXPIRED,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to EXPIRED",
        ),
        "activity-cancelled": (
            ACTIVITY_STATUS_CANCELLED,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to CANCELLED",
        ),
        "claim-submitted": (
            ACTIVITY_STATUS_SUBMITTED,
            {"activity_approved": 0, "activity_completed": 0.75},
            f"{stage} - status mapped to SUBMITTED TO APPROVAL",
        ),
        "claim-declined": (
            ACTIVITY_STATUS_DECLINED,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to DECLINED",
        ),
        "claim-resubmitted": (
            ACTIVITY_STATUS_RESUBMITTED,
            {"activity_approved": 0, "activity_completed": 0.75},
            f"{stage} - status mapped to RESUBMITTED TO APPROVAL",
        ),
        "claim-expired": (
            ACTIVITY_STATUS_EXPIRED,
            {"activity_approved": 0, "activity_completed": 0},
            f"{stage} - status mapped to EXPIRED",
        ),
        "claim-approved": (
            ACTIVITY_STATUS_CLOSED,
            {"activity_approved": 1, "activity_completed": 1},
            f"{stage} - status mapped to CLOSED",
        ),
        "claim-paid": (
            ACTIVITY_STATUS_CLOSED,
            {"activity_approved": 1, "activity_completed": 1},
            f"{stage} - status mapped to CLOSED",
        ),
    }

    if stage_key not in mapping:
        raise ValueError(f"Stage não mapeado: {stage}")

    return mapping[stage_key]


def _get_task_columns_for_match(task_id: int) -> Optional[Dict[str, Any]]:
    required_columns = [
        "task_id",
        "task_tasktype_id",
        "task_status",
        "task_customer_id",
        "task_cr_party_id",
        "task_cr_party_name",
        "task_deal_id",
        "task_track",
        "task_subtrack",
        "task_completed",
    ]

    try:
        if hasattr(repo_task, "get_columns_by_task_id"):
            row = repo_task.get_columns_by_task_id(task_id, required_columns, as_df=False)

            if isinstance(row, dict):
                return row

            if isinstance(row, list) and row:
                if isinstance(row[0], dict):
                    return row[0]

        rows = repo_task.get_task(task_id=task_id, as_df=False)
        if isinstance(rows, list) and rows:
            if isinstance(rows[0], dict):
                return rows[0]
    except Exception:
        return None

    return None


def _resolve_existing_task(
    customer_id: int,
    deal_id: Optional[str],
    track: Optional[str],
    subtrack: Optional[str],
    cr_party_id: Optional[str],
    task_type_priority: List[int],
) -> Optional[Dict[str, Any]]:
    if not customer_id or not deal_id or not track or not subtrack or not cr_party_id:
        return None

    cr_party_id_int = _safe_int(cr_party_id, default=0)
    if not cr_party_id_int:
        return None

    matched_by_type: Dict[int, List[Dict[str, Any]]] = {TASK_TYPE_21: [], TASK_TYPE_22: []}

    for task_type_id in task_type_priority:
        where = {
            "task_customer_id": customer_id,
            "task_deal_id": deal_id,
            "task_track": track,
            "task_subtrack": subtrack,
            "task_cr_party_id": cr_party_id_int,
            "task_tasktype_id": task_type_id,
        }

        try:
            task_ids = repo_task.find_ids_by(where)
        except Exception:
            task_ids = []

        for found_id in task_ids or []:
            task_id = _safe_int(found_id, default=0)
            if not task_id:
                continue

            task_row = _get_task_columns_for_match(task_id)
            if not task_row:
                continue

            real_type = _safe_int(task_row.get("task_tasktype_id"), default=0)
            if real_type != task_type_id:
                continue

            matched_by_type.setdefault(task_type_id, []).append(task_row)

    for task_type_id in task_type_priority:
        matches = matched_by_type.get(task_type_id, [])
        if len(matches) > 1:
            raise ValueError(
                f"Mais de uma task tipo {task_type_id} encontrada para os critérios informados. "
                f"task_ids={[t.get('task_id') for t in matches]}"
            )
        if len(matches) == 1:
            return matches[0]

    return None


def _resolve_existing_activity(
    activity_ws: Optional[str],
    where_fallback: Dict[str, Any],
) -> Optional[Dict[str, Any]]:
    if activity_ws:
        try:
            ids_by_ws = repo_activity.find_ids_by({"activity_ws": activity_ws})
            if ids_by_ws:
                activity_id = _safe_int(ids_by_ws[0], default=0)
                if activity_id:
                    data = repo_activity.get_activity_by_id(
                        activity_id=activity_id,
                        as_df=False,
                    )
                    if data:
                        return data
                    return {"activity_id": activity_id}
        except Exception:
            pass

    try:
        ids = repo_activity.find_ids_by(where_fallback)
        if ids:
            activity_id = _safe_int(ids[0], default=0)
            if activity_id:
                data = repo_activity.get_activity_by_id(
                    activity_id=activity_id,
                    as_df=False,
                )
                if data:
                    if activity_ws and not data.get("activity_ws"):
                        try:
                            repo_activity.update(
                                data={"activity_ws": activity_ws},
                                where={"activity_id": activity_id},
                            )
                            data["activity_ws"] = activity_ws
                        except Exception:
                            pass
                    return data
                return {"activity_id": activity_id}
    except Exception:
        return None

    return None


def _insert_history(
    task_id: int,
    activity_id: int,
    remark: str,
    updated_by: str = "System BA",
) -> None:
    repo_history.insert(
        {
            "taskrecord_task_id": task_id,
            "taskrecord_activity_id": activity_id,
            "taskrecord_remark": remark,
            "taskrecord_updated_by": updated_by,
        }
    )


def _format_money(value: Any) -> str:
    try:
        return f"{float(value):.2f}"
    except Exception:
        return "0.00"


def _recalculate_task_completed(task_id: int) -> bool:
    try:
        task_rows = repo_task.get_task(task_id=task_id, as_df=False)
        if not task_rows:
            return False

        task_data = task_rows[0]
        current_task_completed = float(task_data.get("task_completed") or 0)

        activities = repo_activity.get_activities_by_task_id(task_id, as_df=False)
        if not activities:
            if current_task_completed != 0:
                repo_task.update(data={"task_completed": 0}, where={"task_id": task_id})
                return True
            return False

        count_activities = len(activities)
        sum_completed = 0.0

        for activity in activities:
            try:
                sum_completed += float(activity.get("activity_completed") or 0)
            except Exception:
                continue

        percent_completed = 0.0 if count_activities == 0 else sum_completed / count_activities

        if percent_completed > 1:
            percent_completed = 1.0

        percent_completed = round(percent_completed, 2)

        if float(current_task_completed) != float(percent_completed):
            repo_task.update(
                data={"task_completed": percent_completed},
                where={"task_id": task_id},
            )
            return True

        return False
    except Exception:
        return False


def _update_task_cr_party_if_needed(
    task_id: int,
    cr_party_id: Optional[str],
    cr_party_name: Optional[str],
) -> bool:
    updates: Dict[str, Any] = {}

    cr_party_id_int = _safe_int(cr_party_id, default=0)
    if cr_party_id_int:
        updates["task_cr_party_id"] = cr_party_id_int

    if cr_party_name:
        updates["task_cr_party_name"] = cr_party_name

    if not updates:
        return False

    try:
        repo_task.update(data=updates, where={"task_id": task_id})
        return True
    except Exception:
        return False


def _build_insert_payload(
    row_dict: Dict[str, Any],
    task_id: int,
    deal_id: Optional[str],
    seq_id: int,
    activity_type: Optional[str],
    activity_ws: Optional[str],
    amount: float,
    currency: str,
    start_date: Optional[date],
    end_date: Optional[date],
) -> Dict[str, Any]:
    payload: Dict[str, Any] = {
        "activity_track": _normalize_str(row_dict.get("Track")),
        "activity_sub_track": _normalize_str(row_dict.get("Sub-Track")),
        "activity_deal_id": deal_id,
        "activity_task_id": task_id,
        "activity_name": activity_type,
        "activity_ws": activity_ws,
        "activity_seq": seq_id,
        "activity_value": amount,
        "activity_currency": currency,
    }

    if start_date:
        payload["activity_start"] = start_date
        payload["activity_start_performed"] = start_date

    if end_date:
        payload["activity_end"] = end_date
        payload["activity_end_performed"] = end_date

    return payload


def _apply_amount_update_if_needed(
    task_id: int,
    activity_id: int,
    current_activity: Dict[str, Any],
    new_amount: float,
) -> bool:
    current_amount = float(current_activity.get("activity_value") or 0)

    if float(new_amount) == float(current_amount):
        return False

    repo_activity.update(
        data={"activity_value": new_amount},
        where={"activity_id": activity_id},
    )

    _insert_history(
        task_id=task_id,
        activity_id=activity_id,
        remark=f"Activity value changed from ${_format_money(current_amount)} to ${_format_money(new_amount)}",
    )
    return True


def _apply_date_updates_if_needed(
    activity_id: int,
    current_activity: Dict[str, Any],
    start_date: Optional[date],
    end_date: Optional[date],
) -> bool:
    updates: Dict[str, Any] = {}

    current_start = _to_date(current_activity.get("activity_start"))
    current_start_perf = _to_date(current_activity.get("activity_start_performed"))
    current_end = _to_date(current_activity.get("activity_end"))
    current_end_perf = _to_date(current_activity.get("activity_end_performed"))

    if start_date:
        if current_start and current_start_perf:
            if start_date < current_start_perf:
                updates["activity_start_performed"] = start_date
        elif current_start and not current_start_perf:
            if start_date < current_start:
                updates["activity_start_performed"] = start_date
        elif not current_start and not current_start_perf:
            updates["activity_start"] = start_date
            updates["activity_start_performed"] = start_date

    if end_date:
        if current_end and current_end_perf:
            if end_date > current_end_perf:
                updates["activity_end_performed"] = end_date
        elif current_end and not current_end_perf:
            if end_date > current_end:
                updates["activity_end_performed"] = end_date
        elif not current_end and not current_end_perf:
            updates["activity_end"] = end_date
            updates["activity_end_performed"] = end_date

    if not updates:
        return False

    repo_activity.update(
        data=updates,
        where={"activity_id": activity_id},
    )
    return True


def _resolve_approval_date(
    claim_approval_date: Optional[date],
    l1_decision_date: Optional[date],
) -> Optional[date]:
    if claim_approval_date and l1_decision_date:
        return claim_approval_date if claim_approval_date >= l1_decision_date else l1_decision_date

    if claim_approval_date:
        return claim_approval_date

    if l1_decision_date:
        return l1_decision_date

    return None


def _apply_status_update_if_needed(
    row_dict: Dict[str, Any],
    task_id: int,
    activity_id: int,
    current_activity: Dict[str, Any],
    activity_already_exists: bool,
) -> bool:
    stage = _normalize_str(row_dict.get("Stage"))
    status_id, extra_fields, history_text = _resolve_status_from_stage(stage)

    stage_key = _normalize_stage_key(stage)
    current_status = _safe_int(current_activity.get("activity_status"), default=0)

    closed_statuses = {4, 5, 6, 10}
    final_close_stages = {"claim-approved", "claim-paid"}

    updates: Dict[str, Any] = {}
    history_parts: List[str] = []

    if stage_key and stage_key.startswith("payment-type"):
        final_close_stages.add(stage_key)

    # Regra 1:
    # Activity - Approved não muda status se a activity já existir
    if stage_key == "activity-approved" and activity_already_exists:
        return False

    # Regra 3:
    # Claim - Approved / Claim - Paid / Payment - Type* só fecham se current_status NOT IN (4,5,6,10)
    if stage_key in final_close_stages:
        if current_status in closed_statuses:
            return False

        updates["activity_status"] = ACTIVITY_STATUS_CLOSED
        updates["activity_approved"] = 1
        updates["activity_completed"] = 1

        # Nova regra: usar Claim Approval Date para fechamento
        claim_approval_date = _to_date(row_dict.get("Claim Approval Date"))
        if claim_approval_date:
            updates["activity_end_performed"] = claim_approval_date
            updates["activity_end_fy"] = _fiscal_year_apr_mar(claim_approval_date)

        history_parts.append(history_text)
    else:
        # Regra geral: não alterar se já estiver em status fechado
        if current_status in closed_statuses:
            return False

        # Somente evolui se houver avanço de status
        if not (status_id > current_status):
            return False

        updates["activity_status"] = status_id
        updates.update(extra_fields)
        history_parts.append(history_text)

    if 7 <= updates.get("activity_status", 0) <= 10:
        claim_submitted_by = _normalize_str(row_dict.get("Claim Submitted By"), max_length=255)
        claim_submitted_date = _to_date(row_dict.get("Claim Submitted Date"))

        if claim_submitted_by:
            history_parts.append(f"Claim Submitted by {claim_submitted_by}")

        if claim_submitted_date:
            history_parts.append(f"at {claim_submitted_date.strftime('%Y-%b-%d')}")
            updates["activity_approval_request_date"] = claim_submitted_date

        _, stage_completion_date, _ = _resolve_seq_and_stage_completion(row_dict)

        if updates.get("activity_status") == ACTIVITY_STATUS_CLOSED:
            approved_amount = _to_float(row_dict.get("Claim Approved Amount")) or 0.0

            if stage_completion_date:
                history_parts.append(
                    f"Stage completion date at {stage_completion_date.strftime('%Y-%b-%d')}"
                )

            if approved_amount > 0:
                updates["activity_approved_value"] = approved_amount
                history_parts.append(f"Claim approved ${_format_money(approved_amount)}")

                approval_date = _resolve_approval_date(
                    claim_approval_date=_to_date(row_dict.get("Claim Approval Date")),
                    l1_decision_date=_to_date(row_dict.get("L1 Claim Decision Date")),
                )

                if approval_date:
                    updates["activity_approval_date"] = approval_date
                    history_parts.append(f"approved at {approval_date.strftime('%Y-%b-%d')}")

    repo_activity.update(
        data=updates,
        where={"activity_id": activity_id},
    )

    _insert_history(
        task_id=task_id,
        activity_id=activity_id,
        remark="; ".join(history_parts),
    )

    return True


def _process_single_row(
    row_dict: Dict[str, Any],
    file_path: str,
    row_number: int,
    execution_log_path: Optional[Path] = None,
) -> RowProcessResult:
    try:
        end_date = _to_date(row_dict.get("Activity Expiration Date"))
        current_fy = _current_fiscal_year()
        cutoff_fy = current_fy - 1
        end_date_fy = _fiscal_year_of_date(end_date) if end_date else current_fy

        activity_ws = _normalize_str(row_dict.get("Activity Id"), max_length=255)
        track = _normalize_str(row_dict.get("Track"), max_length=255)
        subtrack = _normalize_str(row_dict.get("Sub-Track"), max_length=255)
        activity_type = _normalize_str(row_dict.get("Activity Type"), max_length=255)
        activity_group = _normalize_activity_group(row_dict.get("Activity Group"))
        stage = _normalize_str(row_dict.get("Stage"), max_length=255)
        deal_id = _normalize_str(row_dict.get("Deal ID"), max_length=255)
        cr_party_id = _normalize_str(row_dict.get("CR Party ID"), max_length=255)
        cr_party_name = _normalize_str(row_dict.get("Deal CR Party Name"), max_length=500)

        if not activity_ws:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No WS"
            _safe_log(file_path, row_number, message, "Activity Id", row_dict.get("Activity Id"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Activity Id",
                error_value=row_dict.get("Activity Id"),
            )

        if not track:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "TRACK is empty"
            _safe_log(file_path, row_number, message, "Track", row_dict.get("Track"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Track",
                error_value=row_dict.get("Track"),
            )

        if not subtrack:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "SUBTRACK is empty"
            _safe_log(file_path, row_number, message, "Sub-Track", row_dict.get("Sub-Track"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Sub-Track",
                error_value=row_dict.get("Sub-Track"),
            )

        if not activity_type:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No Activity Type"
            _safe_log(file_path, row_number, message, "Activity Type", row_dict.get("Activity Type"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Activity Type",
                error_value=row_dict.get("Activity Type"),
            )

        if not deal_id:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No Deal Id"
            _safe_log(file_path, row_number, message, "Deal ID", row_dict.get("Deal ID"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Deal ID",
                error_value=row_dict.get("Deal ID"),
            )

        if not cr_party_id:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No CR Party ID"
            _safe_log(file_path, row_number, message, "CR Party ID", row_dict.get("CR Party ID"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="CR Party ID",
                error_value=row_dict.get("CR Party ID"),
            )

        seq_id, _, resolved_activity_type = _resolve_seq_and_stage_completion(row_dict)
        if not resolved_activity_type:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No Activity Type"
            _safe_log(file_path, row_number, message, "Activity Type", row_dict.get("Activity Type"))
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Activity Type",
                error_value=row_dict.get("Activity Type"),
            )

        customer_id = _resolve_customer_id(cr_party_name, deal_id)
        if not customer_id:
            if end_date_fy <= cutoff_fy:
                return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

            message = "No Customer name or Customer Id"
            _safe_log(file_path, row_number, message, "Deal CR Party Name", cr_party_name)
            return RowProcessResult(
                success=False,
                error_message=message,
                error_column="Deal CR Party Name",
                error_value=cr_party_name,
            )

        task_type_priority = _resolve_task_type_priority_from_activity_group(activity_group)

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO row={row_number} activity_group={activity_group} task_type_priority={task_type_priority}",
            )

        amount = _to_float(row_dict.get("Activity Requested Amount")) or 0.0
        currency = _normalize_currency(row_dict.get("Activity Requested Amount (Currency)"), default="USD")
        start_date = _to_date(row_dict.get("Activity Start Date"))

        # Tenta resolver a activity pelo activity_ws ANTES de buscar a task.
        # Isso evita o erro "Mais de uma task encontrada" quando a activity já
        # existe no banco — nesse caso o task_id é obtido direto do activity_task_id.
        pre_existing_activity = _resolve_existing_activity(
            activity_ws=activity_ws,
            where_fallback={},
        )

        if pre_existing_activity:
            # Activity já cadastrada: obtém task_id diretamente do registro existente.
            task_id = _safe_int(pre_existing_activity.get("activity_task_id"), default=0)
            if not task_id:
                message = "No TASK ID (activity exists but has no activity_task_id)"
                _safe_log(file_path, row_number, message, "Activity Id", activity_ws)
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Activity Id",
                    error_value=activity_ws,
                )

            existing_activity = pre_existing_activity

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} activity found by ws={activity_ws} activity_id={pre_existing_activity.get('activity_id')} task_id={task_id}",
                )
        else:
            # Activity não existe: resolve a task para saber onde inserir.
            task_data = _resolve_existing_task(
                customer_id=customer_id,
                deal_id=deal_id,
                track=track,
                subtrack=subtrack,
                cr_party_id=cr_party_id,
                task_type_priority=task_type_priority,
            )

            if not task_data:
                if end_date_fy <= cutoff_fy:
                    return RowProcessResult(success=True, discarded_old_row=True, ignored=True)

                message = "No TASK ID"
                _safe_log(
                    file_path,
                    row_number,
                    message,
                    "Deal ID",
                    {
                        "deal_id": deal_id,
                        "track": track,
                        "subtrack": subtrack,
                        "cr_party_id": cr_party_id,
                        "customer_id": customer_id,
                        "activity_group": activity_group,
                        "task_type_priority": task_type_priority,
                    },
                )
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Deal ID",
                    error_value=deal_id,
                )

            task_id = _safe_int(task_data.get("task_id"), default=0)
            if not task_id:
                message = "No TASK ID"
                _safe_log(file_path, row_number, message, "Deal ID", deal_id)
                return RowProcessResult(
                    success=False,
                    error_message=message,
                    error_column="Deal ID",
                    error_value=deal_id,
                )

            existing_activity = None

        _update_task_cr_party_if_needed(
            task_id=task_id,
            cr_party_id=cr_party_id,
            cr_party_name=cr_party_name,
        )

        where_fallback = {
            "activity_track": track,
            "activity_sub_track": subtrack,
            "activity_deal_id": deal_id,
            "activity_name": resolved_activity_type,
            "activity_task_id": task_id,
        }

        if existing_activity is None:
            existing_activity = _resolve_existing_activity(
                activity_ws=activity_ws,
                where_fallback=where_fallback,
            )

        created = False
        updated = False
        task_updated = False

        if existing_activity is None:
            payload = _build_insert_payload(
                row_dict=row_dict,
                task_id=task_id,
                deal_id=deal_id,
                seq_id=seq_id,
                activity_type=resolved_activity_type,
                activity_ws=activity_ws,
                amount=amount,
                currency=currency,
                start_date=start_date,
                end_date=end_date,
            )

            activity_id = repo_activity.insert(payload)
            activity_id = _safe_int(activity_id, default=0)

            if not activity_id:
                raise ValueError("Falha ao inserir activity")

            _insert_history(
                task_id=task_id,
                activity_id=activity_id,
                remark=f"Activity created by System BA at {datetime.now().strftime('%Y-%b-%d')}",
            )

            existing_activity = repo_activity.get_activity_by_id(
                activity_id=activity_id,
                as_df=False,
            ) or {"activity_id": activity_id}

            created = True

            if execution_log_path:
                _append_execution_log(
                    execution_log_path,
                    f"INFO row={row_number} activity created activity_id={activity_id} task_id={task_id}",
                )

        activity_id = _safe_int(existing_activity.get("activity_id"), default=0)
        if not activity_id:
            raise ValueError("activity_id inválido")

        if _apply_amount_update_if_needed(
            task_id=task_id,
            activity_id=activity_id,
            current_activity=existing_activity,
            new_amount=amount,
        ):
            updated = True

        refreshed_activity = repo_activity.get_activity_by_id(
            activity_id=activity_id,
            as_df=False,
        ) or existing_activity

        if _apply_date_updates_if_needed(
            activity_id=activity_id,
            current_activity=refreshed_activity,
            start_date=start_date,
            end_date=end_date,
        ):
            updated = True

        refreshed_activity = repo_activity.get_activity_by_id(
            activity_id=activity_id,
            as_df=False,
        ) or refreshed_activity

        if stage:
            if _apply_status_update_if_needed(
                row_dict=row_dict,
                task_id=task_id,
                activity_id=activity_id,
                current_activity=refreshed_activity,
                activity_already_exists=not created,
            ):
                updated = True
        else:
            return RowProcessResult(
                success=True,
                created=created,
                updated=updated,
                task_updated=task_updated,
                ignored=False,
            )

        if _recalculate_task_completed(task_id):
            task_updated = True

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"INFO row={row_number} processed task_id={task_id} activity_id={activity_id} created={created} updated={updated} task_updated={task_updated}",
            )

        return RowProcessResult(
            success=True,
            created=created,
            updated=updated,
            task_updated=task_updated,
        )

    except Exception as ex:
        message = f"Unexpected error: {ex}"
        _safe_log(file_path, row_number, message, None, None)

        if execution_log_path:
            _append_execution_log(
                execution_log_path,
                f"ERROR row={row_number} unexpected error={str(ex)[:1000]} traceback={traceback.format_exc()[:2000]}",
            )

        return RowProcessResult(
            success=False,
            error_message=message,
            error_column=None,
            error_value=None,
        )


def _process_chunk(
    ws: Worksheet,
    idx_map: Dict[str, int],
    headers: List[str],
    start_row: int,
    end_row: int,
    file_path: str,
    failed_output_path: Path,
    execution_log_path: Optional[Path] = None,
    chunk_number: int = 0,
) -> Dict[str, int]:
    metrics = {
        "success": 0,
        "error": 0,
        "ignored": 0,
        "discarded_old_row": 0,
        "created": 0,
        "updated": 0,
        "task_updated": 0,
    }

    failed_rows_buffer: List[List[Any]] = []

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_start chunk={chunk_number} start_row={start_row} end_row={end_row}",
        )

    for r in range(end_row, start_row - 1, -1):
        row_cells = ws[r]
        raw_values = [cell.value for cell in row_cells]
        row_dict: Dict[str, Any] = {}

        for h, idx in idx_map.items():
            row_dict[h] = raw_values[idx] if idx < len(raw_values) else None

        result = _process_single_row(
            row_dict=row_dict,
            file_path=file_path,
            row_number=r,
            execution_log_path=execution_log_path,
        )

        if result.success:
            metrics["success"] += 1
            if result.ignored:
                metrics["ignored"] += 1
            if result.discarded_old_row:
                metrics["discarded_old_row"] += 1
            if result.created:
                metrics["created"] += 1
            if result.updated:
                metrics["updated"] += 1
            if result.task_updated:
                metrics["task_updated"] += 1
        else:
            metrics["error"] += 1
            failed_rows_buffer.append(
                _row_dict_to_failed_excel_row(
                    headers=headers,
                    row_dict=row_dict,
                    original_row_number=r,
                    input_file_path=file_path,
                    result=result,
                )
            )

        ws.delete_rows(r, 1)

    if failed_rows_buffer:
        failed_rows_buffer.reverse()
        _append_failed_rows(
            failed_path=failed_output_path,
            rows_to_append=failed_rows_buffer,
            original_headers=headers,
            execution_log_path=execution_log_path,
        )

    if execution_log_path:
        _append_execution_log(
            execution_log_path,
            f"INFO chunk_finish chunk={chunk_number} success={metrics['success']} error={metrics['error']} ignored={metrics['ignored']} discarded_old_row={metrics['discarded_old_row']} created={metrics['created']} updated={metrics['updated']} task_updated={metrics['task_updated']}",
        )

    return metrics


def run_import(
    file_path: str,
    user_id: Optional[str] = None,
    import_control_id: Optional[int] = None,
) -> Dict[str, Any]:
    _ensure_directories()

    input_path = _resolve_input_path(file_path)
    failed_output_path = _build_failed_output_path(input_path)
    execution_log_path = _build_execution_log_path(input_path)

    if import_control_id is None and repo_import_control is not None:
        try:
            import_control_id = repo_import_control.get_id_by_file(input_path.name)
        except Exception:
            import_control_id = None

    started_at = datetime.now()
    total_rows = 0
    chunk_number = 0

    summary_metrics = {
        "processed_success": 0,
        "failed_rows": 0,
        "ignored_rows": 0,
        "discarded_old_rows": 0,
        "activities_created": 0,
        "activities_updated": 0,
        "tasks_updated": 0,
    }

    _append_execution_log(execution_log_path, "START import")
    _append_execution_log(execution_log_path, f"INFO import_source={IMPORT_SOURCE}")
    _append_execution_log(execution_log_path, f"INFO user_id={user_id}")
    _append_execution_log(execution_log_path, f"INFO input_file={str(input_path)}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")

    try:
        wb, ws, idx_map, headers = _open_workbook_rw(str(input_path))
    except Exception as e:
        msg = f"Erro ao abrir arquivo Excel: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR failed opening workbook error={str(e)[:1000]}")
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": 0,
                "processed_success": 0,
                "failed_rows": 0,
                "execution_log_path": str(execution_log_path),
            },
        }

    total_rows = max(ws.max_row - 1, 0)
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO chunk_size={CHUNK_SIZE}")

    missing_cols = _validate_required_columns(headers)
    if missing_cols:
        msg = "Missing columns: " + ", ".join(missing_cols)
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR missing_columns={', '.join(missing_cols)}")
        wb.close()
        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": 0,
                "failed_rows": total_rows,
                "execution_log_path": str(execution_log_path),
            },
        }

    # Removido _ensure_failed_workbook aqui para que o arquivo de falhas
    # só seja criado quando realmente houver linhas com erro (dentro de _append_failed_rows).

    try:
        while ws.max_row > 1:
            chunk_number += 1
            chunk_end = min(1 + CHUNK_SIZE, ws.max_row)

            chunk_metrics = _process_chunk(
                ws=ws,
                idx_map=idx_map,
                headers=headers,
                start_row=2,
                end_row=chunk_end,
                file_path=str(input_path),
                failed_output_path=failed_output_path,
                execution_log_path=execution_log_path,
                chunk_number=chunk_number,
            )

            summary_metrics["processed_success"] += chunk_metrics["success"]
            summary_metrics["failed_rows"] += chunk_metrics["error"]
            summary_metrics["ignored_rows"] += chunk_metrics["ignored"]
            summary_metrics["discarded_old_rows"] += chunk_metrics["discarded_old_row"]
            summary_metrics["activities_created"] += chunk_metrics["created"]
            summary_metrics["activities_updated"] += chunk_metrics["updated"]
            summary_metrics["tasks_updated"] += chunk_metrics["task_updated"]

            wb.save(str(input_path))

            remaining_rows_current = max(ws.max_row - 1, 0)

            _append_execution_log(
                execution_log_path,
                f"INFO checkpoint_after_chunk chunk={chunk_number} success_total={summary_metrics['processed_success']} error_total={summary_metrics['failed_rows']} ignored_total={summary_metrics['ignored_rows']} remaining_rows_current={remaining_rows_current}",
            )

            if chunk_number % IMPORT_CONTROL_PROGRESS_INTERVAL_CHUNKS == 0:
                _report_progress(
                    import_control_id=import_control_id,
                    total_rows=total_rows,
                    processed_success=summary_metrics["processed_success"],
                    failed_rows=summary_metrics["failed_rows"],
                    ignored_rows=summary_metrics["ignored_rows"],
                    remaining_rows=remaining_rows_current,
                    execution_log_path=execution_log_path,
                )

    except Exception as e:
        msg = f"Erro durante o processamento da importação: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR processing failed error={str(e)[:1000]}")

        try:
            wb.save(str(input_path))
            _append_execution_log(execution_log_path, "INFO workbook saved after processing exception")
        except Exception as save_ex:
            _append_execution_log(
                execution_log_path,
                f"ERROR failed saving workbook after exception error={str(save_ex)[:1000]}",
            )

        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "discarded_old_rows": summary_metrics["discarded_old_rows"],
                "activities_created": summary_metrics["activities_created"],
                "activities_updated": summary_metrics["activities_updated"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    try:
        wb.save(str(input_path))
        _append_execution_log(execution_log_path, "INFO workbook saved successfully at end")
    except Exception as e:
        msg = f"Erro ao salvar arquivo Excel após processamento: {e}"
        _safe_log(str(input_path), 0, msg, None, None)
        _append_execution_log(execution_log_path, f"ERROR final workbook save failed error={str(e)[:1000]}")
        wb.close()

        finished_at = datetime.now()
        duration_seconds = int((finished_at - started_at).total_seconds())

        _append_execution_log(
            execution_log_path,
            f"FINISH status=FAILED success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} duration_seconds={duration_seconds}",
        )

        return {
            "status": "FAILED",
            "message": msg,
            "summary": {
                "file_path": str(input_path),
                "total_rows": total_rows,
                "processed_success": summary_metrics["processed_success"],
                "failed_rows": summary_metrics["failed_rows"],
                "ignored_rows": summary_metrics["ignored_rows"],
                "discarded_old_rows": summary_metrics["discarded_old_rows"],
                "activities_created": summary_metrics["activities_created"],
                "activities_updated": summary_metrics["activities_updated"],
                "tasks_updated": summary_metrics["tasks_updated"],
                "failed_file_path": str(failed_output_path),
                "execution_log_path": str(execution_log_path),
            },
        }

    wb.close()

    remaining_rows_in_input = 0
    try:
        wb_check = load_workbook(str(input_path), read_only=True, data_only=True)
        ws_check = wb_check.active
        remaining_rows_in_input = max(ws_check.max_row - 1, 0)
        wb_check.close()
    except Exception as e:
        remaining_rows_in_input = -1
        _append_execution_log(
            execution_log_path,
            f"WARN failed checking remaining rows error={str(e)[:1000]}",
        )

    finished_at = datetime.now()
    duration_seconds = int((finished_at - started_at).total_seconds())

    _append_execution_log(execution_log_path, f"INFO total_chunks={chunk_number}")
    _append_execution_log(execution_log_path, f"INFO total_rows_initial={total_rows}")
    _append_execution_log(execution_log_path, f"INFO processed_success={summary_metrics['processed_success']}")
    _append_execution_log(execution_log_path, f"INFO failed_rows={summary_metrics['failed_rows']}")
    _append_execution_log(execution_log_path, f"INFO ignored_rows={summary_metrics['ignored_rows']}")
    _append_execution_log(execution_log_path, f"INFO discarded_old_rows={summary_metrics['discarded_old_rows']}")
    _append_execution_log(execution_log_path, f"INFO activities_created={summary_metrics['activities_created']}")
    _append_execution_log(execution_log_path, f"INFO activities_updated={summary_metrics['activities_updated']}")
    _append_execution_log(execution_log_path, f"INFO tasks_updated={summary_metrics['tasks_updated']}")
    _append_execution_log(execution_log_path, f"INFO remaining_rows_in_input={remaining_rows_in_input}")
    _append_execution_log(execution_log_path, f"INFO failed_file={str(failed_output_path)}")
    _append_execution_log(execution_log_path, f"INFO duration_seconds={duration_seconds}")

    if summary_metrics["failed_rows"] > 0:
        status = "FAILED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com falhas. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"descartadas_antigas={summary_metrics['discarded_old_rows']}, criadas={summary_metrics['activities_created']}, "
            f"atualizadas={summary_metrics['activities_updated']}, tasks_atualizadas={summary_metrics['tasks_updated']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )
    else:
        status = "FINISHED"
        msg = (
            f"{IMPORT_SOURCE} import concluído com sucesso. "
            f"arquivo={str(input_path)}, total={total_rows}, sucesso={summary_metrics['processed_success']}, "
            f"erros={summary_metrics['failed_rows']}, ignoradas={summary_metrics['ignored_rows']}, "
            f"descartadas_antigas={summary_metrics['discarded_old_rows']}, criadas={summary_metrics['activities_created']}, "
            f"atualizadas={summary_metrics['activities_updated']}, tasks_atualizadas={summary_metrics['tasks_updated']}, "
            f"linhas_restantes_input={remaining_rows_in_input}, arquivo_falhas={str(failed_output_path)}, "
            f"log_execucao={str(execution_log_path)}."
        )

    _append_execution_log(
        execution_log_path,
        f"FINISH status={status} success={summary_metrics['processed_success']} error={summary_metrics['failed_rows']} ignored={summary_metrics['ignored_rows']} discarded_old_rows={summary_metrics['discarded_old_rows']} remaining_rows_in_input={remaining_rows_in_input} duration_seconds={duration_seconds}",
    )

    return {
        "status": status,
        "message": msg,
        "summary": {
            "file_path": str(input_path),
            "total_rows": total_rows,
            "processed_success": summary_metrics["processed_success"],
            "failed_rows": summary_metrics["failed_rows"],
            "ignored_rows": summary_metrics["ignored_rows"],
            "discarded_old_rows": summary_metrics["discarded_old_rows"],
            "activities_created": summary_metrics["activities_created"],
            "activities_updated": summary_metrics["activities_updated"],
            "tasks_updated": summary_metrics["tasks_updated"],
            "failed_file_path": str(failed_output_path),
            "remaining_rows_in_input": remaining_rows_in_input,
            "execution_log_path": str(execution_log_path),
            "duration_seconds": duration_seconds,
            "total_chunks": chunk_number,
        },
    }
